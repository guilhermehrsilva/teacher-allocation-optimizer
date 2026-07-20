from __future__ import annotations

import json
from collections import Counter
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from .domain import MODULE_STAGE_1, MODULE_STAGE_2, MODULE_STAGES, AllocationResult, Problem


def contract_model_for(job_function: str, allocated: bool) -> str:
    if not allocated:
        return "A DEFINIR"
    normalized = " ".join(job_function.strip().upper().split())
    if normalized in {"PROFESSOR DE ENSINO SUPERIOR EAD", "PROFESSOR REGENTE"}:
        return "CLT EAD"
    if normalized == "PROFESSOR DE ENSINO SUPERIOR PRESENCIAL":
        return "CLT STRICTO"
    return "A DEFINIR"


def create_round_directory(base_dir: str | Path) -> Path:
    base = Path(base_dir)
    base.mkdir(parents=True, exist_ok=True)
    numbers = []
    for item in base.glob("rodada_*"):
        if item.is_dir() and item.name.removeprefix("rodada_").isdigit():
            numbers.append(int(item.name.removeprefix("rodada_")))
    round_dir = base / f"rodada_{max(numbers, default=0) + 1:03d}"
    round_dir.mkdir(parents=False, exist_ok=False)
    return round_dir


def _autosize(worksheet) -> None:
    for column in worksheet.columns:
        width = min(60, max(len(str(cell.value or "")) for cell in column) + 2)
        worksheet.column_dimensions[get_column_letter(column[0].column)].width = width
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    for cell in worksheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")


def write_results(problem: Problem, result: AllocationResult, output_dir: str | Path) -> tuple[Path, Path]:
    output = Path(output_dir)
    output.mkdir(parents=True, exist_ok=True)
    xlsx_path = output / "resultado_alocacao.xlsx"
    json_path = output / "resumo_alocacao.json"
    teachers = {teacher.id: teacher for teacher in problem.teachers}
    transmissions = {item.id: item for item in problem.transmissions}
    teacher_stage_counts = {
        teacher.id: {stage: 0 for stage in MODULE_STAGES}
        for teacher in problem.teachers
    }
    for assignment in result.assignments:
        if assignment.teacher_id is None:
            continue
        for stage in transmissions[assignment.transmission_id].module_stages:
            teacher_stage_counts[assignment.teacher_id][stage] += 1
    allocated_stage_hours = {
        stage: problem.hours_per_transmission
        * sum(counts[stage] for counts in teacher_stage_counts.values())
        for stage in MODULE_STAGES
    }
    average_weekly_allocated_hours = (
        sum(allocated_stage_hours.values()) / len(MODULE_STAGES)
    )

    workbook = Workbook()
    summary = workbook.active
    summary.title = "RESUMO"
    summary.append(("INDICADOR", "VALOR"))
    summary_rows = (
        ("Status do motor", result.status),
        ("Status CP-SAT", result.solver_status),
        ("Transmissões", len(problem.transmissions)),
        ("Alocadas", result.allocated_count),
        ("Não alocadas", result.unassigned_count),
        ("Horas alocadas", result.allocated_count * problem.hours_per_transmission),
        (
            "Horas semanais alocadas - 1ª etapa",
            allocated_stage_hours[MODULE_STAGE_1],
        ),
        (
            "Horas semanais alocadas - 2ª etapa",
            allocated_stage_hours[MODULE_STAGE_2],
        ),
        ("Horas semanais médias no módulo", average_weekly_allocated_hours),
        ("Docentes ativos", sum(teacher.is_active for teacher in problem.teachers)),
        ("Docentes com alocação", result.used_teacher_count),
        ("Docentes ativos sem alocação", result.zero_active_teacher_count),
        ("Score de ocupação por CH", result.high_capacity_score or 0),
        ("Penalidade de escassez após resgate", result.rescue_scarcity_score or 0),
        ("Tempo do solver (s)", round(result.wall_time_seconds, 3)),
    )
    for row in summary_rows:
        summary.append(row)
    _autosize(summary)

    allocations = workbook.create_sheet("ALOCACOES")
    allocations.append((
        "STATUS", "MOTIVO", "MOTIVO_ALOCACAO", "LINHA_ORIGEM", "CURSO",
        "NOME_CURSO", "CURRÍCULO",
        "COD_DISCIPLINA", "NOME_DISCIPLINA", "PERFIL_DISCIPLINA", "SINERGIA",
        "DIA_AULA", "HORÁRIO", "ORDEM", "CLUSTER", "COORDENADOR",
        "MODELO_CONTRATO_ORIGEM", "MODELO_CONTRATO", "DOCENTE", "CHAPA",
        "CANDIDATOS_ELEGÍVEIS",
    ))
    for assignment in result.assignments:
        transmission = transmissions[assignment.transmission_id]
        teacher = teachers.get(assignment.teacher_id)
        allocations.append((
            assignment.status, assignment.reason, assignment.allocation_reason,
            transmission.excel_row,
            transmission.course, transmission.course_name, transmission.curriculum,
            transmission.discipline_code, transmission.discipline_name,
            transmission.profile_text, transmission.synergy, transmission.day,
            transmission.start_time.strftime("%H:%M") if transmission.start_time else "",
            transmission.order, transmission.cluster, transmission.coordinator,
            transmission.contract_model,
            contract_model_for(teacher.job_function, True) if teacher else "A DEFINIR",
            teacher.name if teacher else "",
            teacher.badge if teacher else "", assignment.eligible_teacher_count,
        ))
    _autosize(allocations)

    loads = workbook.create_sheet("DOCENTES")
    loads.append((
        "NOME", "CHAPA", "NM_FUNCAO", "CH_CONTRATADA", "CH_LETIVA", "GESTOR",
        "STATUS", "PERFIL_DISCIPLINA", "CH_ALOCADA", "CH_DISPONÍVEL",
        "UTILIZAÇÃO", "QTD_TRANSMISSÕES", "SITUACAO_ALOCACAO",
        "QTD_DISCIPLINAS_1ª_ETAPA", "UTILIZAÇÃO_1ª_ETAPA",
        "QTD_DISCIPLINAS_2ª_ETAPA", "UTILIZAÇÃO_2ª_ETAPA",
    ))
    for teacher in sorted(problem.teachers, key=lambda item: item.name):
        raw_used = result.teacher_loads.get(teacher.id, 0)
        first_count = teacher_stage_counts[teacher.id][MODULE_STAGE_1]
        second_count = teacher_stage_counts[teacher.id][MODULE_STAGE_2]
        first_load = first_count * problem.hours_per_transmission
        second_load = second_count * problem.hours_per_transmission
        if teacher.is_stricto:
            # A base Stricto preserva CH_LETIVA zero e usa um limite especial de
            # disciplinas; por isso seus percentuais permanecem em 0%.
            used = raw_used
            available = 0
            utilization = first_utilization = second_utilization = 0
        else:
            # Média semanal ponderada pelas duas metades de cinco semanas.
            # ESTENDIDA participa das duas cargas; 1ª e 2ª de apenas uma.
            used = (first_load + second_load) / len(MODULE_STAGES)
            available = teacher.teaching_capacity - used
            first_utilization = (
                first_load / teacher.teaching_capacity
                if teacher.teaching_capacity
                else 0
            )
            second_utilization = (
                second_load / teacher.teaching_capacity
                if teacher.teaching_capacity
                else 0
            )
            utilization = (
                first_utilization + second_utilization
            ) / len(MODULE_STAGES)
        loads.append((
            teacher.name, teacher.badge, teacher.job_function, teacher.contracted_capacity,
            teacher.teaching_capacity, teacher.manager, teacher.status, teacher.profile_text,
            used, available, utilization,
            raw_used // problem.hours_per_transmission,
            "ALOCADO" if raw_used else "SEM ALOCACAO",
            first_count, first_utilization, second_count, second_utilization,
        ))
        for column in (11, 15, 17):
            loads.cell(loads.max_row, column).number_format = "0.0%"
    _autosize(loads)
    temporary_xlsx = xlsx_path.with_suffix(xlsx_path.suffix + ".tmp")
    workbook.save(temporary_xlsx)
    workbook.close()
    temporary_xlsx.replace(xlsx_path)

    reasons = Counter(item.reason for item in result.assignments if item.reason)
    allocation_reasons = Counter(
        item.allocation_reason for item in result.assignments if item.allocation_reason
    )
    contract_model_comparison = Counter(
        (
            transmissions[item.transmission_id].contract_model,
            contract_model_for(teachers[item.teacher_id].job_function, True),
        )
        for item in result.assignments
        if item.teacher_id is not None
    )
    payload = {
        "source": str(problem.source),
        "status": result.status,
        "solver_status": result.solver_status,
        "transmissions": len(problem.transmissions),
        "allocated": result.allocated_count,
        "unassigned": result.unassigned_count,
        "hours_per_transmission": problem.hours_per_transmission,
        "allocated_hours": result.allocated_count * problem.hours_per_transmission,
        "allocated_stage_hours": {
            "first_stage": allocated_stage_hours[MODULE_STAGE_1],
            "second_stage": allocated_stage_hours[MODULE_STAGE_2],
        },
        "average_weekly_allocated_hours": average_weekly_allocated_hours,
        "unassigned_reasons": dict(reasons),
        "allocation_reasons": dict(allocation_reasons),
        "contract_model_comparison": [
            {"original": original, "suggested": suggested, "count": count}
            for (original, suggested), count in sorted(contract_model_comparison.items())
        ],
        "active_teachers": sum(teacher.is_active for teacher in problem.teachers),
        "used_teachers": result.used_teacher_count,
        "zero_active_teachers": result.zero_active_teacher_count,
        "high_capacity_score": result.high_capacity_score,
        "rescue_scarcity_score": result.rescue_scarcity_score,
        "wall_time_seconds": result.wall_time_seconds,
        "diagnostics": result.diagnostics,
        "decisions": [
            {
                "transmission_id": assignment.transmission_id,
                "source_row": transmissions[assignment.transmission_id].excel_row,
                "curriculum": transmissions[assignment.transmission_id].curriculum,
                "discipline_code": transmissions[assignment.transmission_id].discipline_code,
                "discipline_name": transmissions[assignment.transmission_id].discipline_name,
                "status": assignment.status,
                "unassigned_reason": assignment.reason,
                "allocation_reason": assignment.allocation_reason,
                "eligible_teacher_count": assignment.eligible_teacher_count,
                "original_contract_model": (
                    transmissions[assignment.transmission_id].contract_model
                ),
                "suggested_contract_model": (
                    contract_model_for(
                        teachers[assignment.teacher_id].job_function,
                        True,
                    )
                    if assignment.teacher_id is not None
                    else "A DEFINIR"
                ),
                "teacher_id": assignment.teacher_id,
                "teacher_badge": (
                    teachers[assignment.teacher_id].badge
                    if assignment.teacher_id is not None
                    else None
                ),
                "teacher_name": (
                    teachers[assignment.teacher_id].name
                    if assignment.teacher_id is not None
                    else None
                ),
            }
            for assignment in result.assignments
        ],
    }
    temporary_json = json_path.with_suffix(json_path.suffix + ".tmp")
    temporary_json.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    temporary_json.replace(json_path)
    return xlsx_path, json_path
