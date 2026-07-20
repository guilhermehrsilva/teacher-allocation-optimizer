from __future__ import annotations

import json
import shutil
import sys
import unittest
import uuid
from datetime import time
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "src"))

from motor_alocacao.domain import (  # noqa: E402
    AllocationResult,
    Assignment,
    Problem,
    Teacher,
    Transmission,
)
from motor_alocacao.reporting import write_results  # noqa: E402


PROFILE = frozenset({"gestão - administração"})


def teacher(identifier: int, *, stricto: bool = False) -> Teacher:
    return Teacher(
        id=identifier,
        name=f"Docente {identifier}",
        badge=f"{identifier + 1:012d}",
        job_function=(
            "PROFESSOR DE ENSINO SUPERIOR PRESENCIAL"
            if stricto
            else "PROFESSOR REGENTE"
        ),
        contracted_capacity=0 if stricto else 40,
        teaching_capacity=0 if stricto else 4,
        manager="Gestor",
        status="ATIVO",
        profiles=PROFILE,
        profile_text="Gestão - Administração",
    )


def transmission(
    identifier: int,
    *,
    stricto: bool = False,
    order: str = "1ª",
) -> Transmission:
    return Transmission(
        id=identifier,
        excel_row=identifier + 2,
        course=f"C{identifier}",
        course_name="Curso",
        curriculum=f"CURR{identifier}",
        discipline_code=f"D{identifier}",
        discipline_name=f"Disciplina {identifier}",
        profiles=PROFILE,
        profile_text="Gestão - Administração",
        synergy="Curso Único",
        day="SEGUNDA",
        start_time=time(19 + identifier, 0),
        order=order,
        cluster="GESTÃO",
        coordinator="Coordenador",
        contract_model="CLT STRICTO" if stricto else "CLT EAD",
    )


def reporting_fixture() -> tuple[Problem, AllocationResult]:
    problem = Problem(
        source=Path("base_m53.xlsx"),
        teachers=(teacher(0, stricto=True), teacher(1)),
        transmissions=(transmission(0), transmission(1)),
    )
    result = AllocationResult(
        status="PARCIAL",
        solver_status="OPTIMAL",
        assignments=[
            Assignment(
                transmission_id=0,
                teacher_id=0,
                status="ALOCADA",
                reason="",
                eligible_teacher_count=1,
                allocation_reason="Regra Stricto",
            ),
            Assignment(
                transmission_id=1,
                teacher_id=None,
                status="NAO_ALOCADA",
                reason="SEM_DOCENTE_COM_PERFIL_E_CARGA",
                eligible_teacher_count=0,
                allocation_reason="",
            ),
        ],
        teacher_loads={0: 2, 1: 0},
        objective_unassigned=1,
        used_teacher_count=1,
        zero_active_teacher_count=1,
        high_capacity_score=0,
        rescue_scarcity_score=0,
        wall_time_seconds=0.25,
        diagnostics={"all_lexicographic_phases_optimal": True},
    )
    return problem, result


class ReportingTests(unittest.TestCase):
    def setUp(self) -> None:
        self.output_dir = ROOT / "tests" / f".reporting-{uuid.uuid4().hex}"
        self.addCleanup(shutil.rmtree, self.output_dir, True)

    def test_generates_xlsx_and_json_with_line_level_decisions(self) -> None:
        problem, result = reporting_fixture()

        xlsx_path, json_path = write_results(problem, result, self.output_dir)

        self.assertEqual(self.output_dir / "resultado_alocacao.xlsx", xlsx_path)
        self.assertEqual(self.output_dir / "resumo_alocacao.json", json_path)
        self.assertTrue(xlsx_path.is_file())
        self.assertTrue(json_path.is_file())

        payload = json.loads(json_path.read_text(encoding="utf-8"))
        self.assertEqual(2, payload["transmissions"])
        self.assertEqual(1, payload["allocated"])
        self.assertEqual(1, payload["unassigned"])
        self.assertEqual(
            {"first_stage": 2, "second_stage": 0},
            payload["allocated_stage_hours"],
        )
        self.assertEqual(1, payload["average_weekly_allocated_hours"])
        self.assertEqual(
            {"SEM_DOCENTE_COM_PERFIL_E_CARGA": 1},
            payload["unassigned_reasons"],
        )
        self.assertEqual(
            [{"original": "CLT EAD", "suggested": "CLT STRICTO", "count": 1}],
            payload["contract_model_comparison"],
        )
        self.assertIsInstance(payload["decisions"], list)
        self.assertEqual(2, len(payload["decisions"]))

        allocated, unassigned = payload["decisions"]
        self.assertEqual(
            {
                "transmission_id": 0,
                "source_row": 2,
                "curriculum": "CURR0",
                "discipline_code": "D0",
                "discipline_name": "Disciplina 0",
                "status": "ALOCADA",
                "unassigned_reason": "",
                "allocation_reason": "Regra Stricto",
                "eligible_teacher_count": 1,
                "original_contract_model": "CLT EAD",
                "suggested_contract_model": "CLT STRICTO",
                "teacher_id": 0,
                "teacher_badge": "000000000001",
                "teacher_name": "Docente 0",
            },
            allocated,
        )
        self.assertEqual("NAO_ALOCADA", unassigned["status"])
        self.assertEqual("SEM_DOCENTE_COM_PERFIL_E_CARGA", unassigned["unassigned_reason"])
        self.assertEqual("CLT EAD", unassigned["original_contract_model"])
        self.assertEqual("A DEFINIR", unassigned["suggested_contract_model"])
        self.assertIsNone(unassigned["teacher_id"])
        self.assertIsNone(unassigned["teacher_badge"])
        self.assertIsNone(unassigned["teacher_name"])

        workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
        try:
            self.assertEqual(["RESUMO", "ALOCACOES", "DOCENTES"], workbook.sheetnames)
            allocations = workbook["ALOCACOES"]
            headers = [cell.value for cell in allocations[1]]
            first_row = dict(zip(headers, next(allocations.iter_rows(min_row=2, values_only=True))))
            self.assertEqual("ALOCADA", first_row["STATUS"])
            self.assertEqual("Regra Stricto", first_row["MOTIVO_ALOCACAO"])
            self.assertEqual("CLT EAD", first_row["MODELO_CONTRATO_ORIGEM"])
            self.assertEqual("CLT STRICTO", first_row["MODELO_CONTRATO"])
        finally:
            workbook.close()

    def test_preserves_stricto_base_loads_and_removes_temporary_files(self) -> None:
        problem, result = reporting_fixture()

        xlsx_path, _ = write_results(problem, result, self.output_dir)

        workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
        try:
            worksheet = workbook["DOCENTES"]
            headers = [cell.value for cell in worksheet[1]]
            rows = [dict(zip(headers, values)) for values in worksheet.iter_rows(min_row=2, values_only=True)]
            stricto = next(row for row in rows if row["CHAPA"] == "000000000001")
            self.assertEqual(0, stricto["CH_CONTRATADA"])
            self.assertEqual(0, stricto["CH_LETIVA"])
            self.assertEqual(2, stricto["CH_ALOCADA"])
            self.assertEqual(0, stricto["CH_DISPONÍVEL"])
            self.assertEqual(0, stricto["UTILIZAÇÃO"])
            self.assertEqual(1, stricto["QTD_TRANSMISSÕES"])
            self.assertEqual(1, stricto["QTD_DISCIPLINAS_1ª_ETAPA"])
            self.assertEqual(0, stricto["UTILIZAÇÃO_1ª_ETAPA"])
            self.assertEqual(0, stricto["QTD_DISCIPLINAS_2ª_ETAPA"])
            self.assertEqual(0, stricto["UTILIZAÇÃO_2ª_ETAPA"])
            self.assertNotIn("MODELO_CONTRATO", headers)
        finally:
            workbook.close()

        self.assertEqual([], list(self.output_dir.rglob("*.tmp")))

    def test_reports_stage_counts_and_weighted_general_utilization(self) -> None:
        problem = Problem(
            source=Path("base_m53.xlsx"),
            teachers=(teacher(0),),
            transmissions=(
                transmission(0, order="1ª"),
                transmission(1, order="ESTENDIDA"),
            ),
        )
        assignments = [
            Assignment(
                transmission_id=identifier,
                teacher_id=0,
                status="ALOCADA",
                reason="",
                eligible_teacher_count=1,
                allocation_reason="Regra por etapa",
            )
            for identifier in range(2)
        ]
        result = AllocationResult(
            status="COMPLETA",
            solver_status="OPTIMAL",
            assignments=assignments,
            teacher_loads={0: 4},
            objective_unassigned=0,
            used_teacher_count=1,
            zero_active_teacher_count=0,
            high_capacity_score=8,
            rescue_scarcity_score=0,
            wall_time_seconds=0.1,
            teacher_stage_loads={
                0: {"PRIMEIRA_ETAPA": 4, "SEGUNDA_ETAPA": 2},
            },
        )

        xlsx_path, _ = write_results(problem, result, self.output_dir)

        workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
        try:
            worksheet = workbook["DOCENTES"]
            headers = [cell.value for cell in worksheet[1]]
            values = next(worksheet.iter_rows(min_row=2, values_only=True))
            row = dict(zip(headers, values))
            self.assertEqual(
                [
                    "NOME", "CHAPA", "NM_FUNCAO", "CH_CONTRATADA", "CH_LETIVA",
                    "GESTOR", "STATUS", "PERFIL_DISCIPLINA", "CH_ALOCADA",
                    "CH_DISPONÍVEL", "UTILIZAÇÃO", "QTD_TRANSMISSÕES",
                    "SITUACAO_ALOCACAO", "QTD_DISCIPLINAS_1ª_ETAPA",
                    "UTILIZAÇÃO_1ª_ETAPA", "QTD_DISCIPLINAS_2ª_ETAPA",
                    "UTILIZAÇÃO_2ª_ETAPA",
                ],
                headers,
            )
            self.assertNotIn("MODELO_CONTRATO", headers)
            self.assertEqual(2, row["QTD_TRANSMISSÕES"])
            self.assertEqual(2, row["QTD_DISCIPLINAS_1ª_ETAPA"])
            self.assertEqual(1, row["UTILIZAÇÃO_1ª_ETAPA"])
            self.assertEqual(1, row["QTD_DISCIPLINAS_2ª_ETAPA"])
            self.assertEqual(0.5, row["UTILIZAÇÃO_2ª_ETAPA"])
            self.assertEqual(3, row["CH_ALOCADA"])
            self.assertEqual(1, row["CH_DISPONÍVEL"])
            self.assertEqual(0.75, row["UTILIZAÇÃO"])
            for header in (
                "UTILIZAÇÃO",
                "UTILIZAÇÃO_1ª_ETAPA",
                "UTILIZAÇÃO_2ª_ETAPA",
            ):
                column = headers.index(header) + 1
                self.assertEqual("0.0%", worksheet.cell(2, column).number_format)
        finally:
            workbook.close()


if __name__ == "__main__":
    unittest.main()
