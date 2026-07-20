from __future__ import annotations

import csv
import hashlib
import json
import re
from collections import Counter, defaultdict
from dataclasses import asdict, dataclass, field
from datetime import datetime, time, timezone
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook


MAPA_SHEET = "MAPA PEDAGÓGICO"
DOCENTES_SHEET = "DOCENTES"
VALIDATION_CONTRACT_VERSION = "2.1"

MAPA_HEADERS = (
    "CURSO", "NOME_CURSO", "CURRÍCULO", "COD_DISCIPLINA", "NOME_DISCIPLINA",
    "PERFIL_DISCIPLINA", "MATRIZ", "MÓD", "ANO", "ORDEM", "VALIDADO",
    "METODOLOGIA", "FORMATO", "PROVA", "ENTURMAÇÃO", "PERFIL", "CLUSTER",
    "COORDENADOR", "MODELO_CONTRATO", "SINERGIA", "DIA_AULA", "HORÁRIO",
    "FORMATO_AULA",
)

DOCENTES_HEADERS = (
    "NOME", "CHAPA", "NM_FUNCAO", "CH_CONTRATADA", "CH_LETIVA", "GESTOR",
    "STATUS", "PERFIL_DISCIPLINA",
)

MAPA_ENUMS = {
    "ANO": {"1º", "2º", "3º", "4º", "5º"},
    "ORDEM": {"1ª", "2ª", "ESTENDIDA"},
    "VALIDADO": {"VALIDADO"},
    "METODOLOGIA": {"5 SEMANAS", "ESTENDIDA"},
    "FORMATO": {"ONLINE", "HÍBRIDA", "SEMIPRESENCIAL"},
    "PROVA": {"5", "10"},
    "ENTURMAÇÃO": {"N", "PORTFÓLIO"},
    "PERFIL": {"CLUSTER", "ESPECÍFICA", "INSTITUCIONAL"},
    "SINERGIA": {"SINÉRGICA", "CURSO ÚNICO", "CURSO RESPONSÁVEL", "NSA"},
    "DIA_AULA": {"SEGUNDA", "TERÇA", "QUARTA", "QUINTA", "SEXTA", "SÁBADO", "NSA"},
    "FORMATO_AULA": {"AO VIVO"},
}

DOCENTES_ENUMS = {
    "NM_FUNCAO": {
        "PROFESSOR REGENTE",
        "PROFESSOR DE ENSINO SUPERIOR EAD",
        "PROFESSOR DE ENSINO SUPERIOR PRESENCIAL",
    },
    "STATUS": {
        "ATIVO", "DEMITIDO", "LICENÇA MATER.", "APOS. POR INCAPACIDADE PERMANENTE",
        "LICENÇA MATER. COMPL. 180 DIAS",
    },
}


def _blank(value: Any) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


def _text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _norm(value: Any) -> str:
    return " ".join(_text(value).casefold().split())


def _norm_upper(value: Any) -> str:
    return " ".join(_text(value).upper().split())


def _profile_tokens(value: Any) -> set[str]:
    return {_norm(item) for item in _text(value).split(",") if item.strip()}


def _check_profile_format(
    report: "ValidationReport",
    sheet: str,
    records: list[tuple[int, dict[str, Any]]],
) -> None:
    without_tokens: list[int] = []
    empty_fragments: list[int] = []
    for row, record in records:
        text = _text(record["PERFIL_DISCIPLINA"])
        if not text:
            continue
        parts = text.split(",")
        if not _profile_tokens(text):
            without_tokens.append(row)
        elif len(parts) > 1 and any(not part.strip() for part in parts):
            empty_fragments.append(row)
    if without_tokens:
        report.add(
            "ALTA",
            "PERFIL_SEM_TOKENS",
            sheet,
            "PERFIL_DISCIPLINA",
            "O perfil informado não contém nenhuma alternativa utilizável.",
            without_tokens,
            blocking=True,
        )
    if empty_fragments:
        report.add(
            "MÉDIA",
            "PERFIL_MALFORMADO",
            sheet,
            "PERFIL_DISCIPLINA",
            "Há alternativas vazias entre vírgulas no perfil.",
            empty_fragments,
            blocking=False,
        )


def _is_integer(value: Any) -> bool:
    return isinstance(value, int) and not isinstance(value, bool) or (
        isinstance(value, float) and value.is_integer()
    )


def _is_time(value: Any) -> bool:
    if isinstance(value, (time, datetime)):
        return True
    if isinstance(value, str):
        return bool(re.fullmatch(r"(?:[01]\d|2[0-3]):[0-5]\d(?::[0-5]\d)?", value.strip()))
    return False


@dataclass
class Issue:
    severity: str
    code: str
    sheet: str | None
    column: str | None
    message: str
    blocking: bool = False
    count: int = 1
    rows: list[int] = field(default_factory=list)
    details: dict[str, Any] = field(default_factory=dict)


@dataclass
class ValidationReport:
    source: str
    status: str = "APROVADO"
    sheets: dict[str, dict[str, int]] = field(default_factory=dict)
    metadata: dict[str, Any] = field(default_factory=dict)
    issues: list[Issue] = field(default_factory=list)
    assumptions: list[str] = field(default_factory=list)

    def add(
        self,
        severity: str,
        code: str,
        sheet: str | None,
        column: str | None,
        message: str,
        rows: Iterable[int] = (),
        details: dict[str, Any] | None = None,
        blocking: bool | None = None,
    ) -> None:
        row_list = sorted(set(rows))
        is_blocking = severity in {"CRÍTICA", "ALTA"} if blocking is None else blocking
        self.issues.append(
            Issue(
                severity=severity,
                code=code,
                sheet=sheet,
                column=column,
                message=message,
                blocking=is_blocking,
                count=len(row_list) or 1,
                rows=row_list,
                details=details or {},
            )
        )

    def finalize(self) -> None:
        severities = {issue.severity for issue in self.issues}
        if any(issue.blocking for issue in self.issues):
            self.status = "REPROVADO"
        elif severities & {"ALTA", "MÉDIA"}:
            self.status = "APROVADO_COM_RESSALVAS"
        else:
            self.status = "APROVADO"

    @property
    def severity_counts(self) -> dict[str, int]:
        counts = Counter(issue.severity for issue in self.issues)
        return {key: counts.get(key, 0) for key in ("CRÍTICA", "ALTA", "MÉDIA", "BAIXA")}

    def to_dict(self) -> dict[str, Any]:
        return {
            "source": self.source,
            "status": self.status,
            "sheets": self.sheets,
            "metadata": self.metadata,
            "summary": {
                "issue_groups": len(self.issues),
                "blocking_issue_groups": sum(issue.blocking for issue in self.issues),
                "by_severity": self.severity_counts,
            },
            "assumptions": self.assumptions,
            "issues": [asdict(issue) for issue in self.issues],
        }

    def write(self, output_dir: str | Path) -> tuple[Path, Path]:
        output = Path(output_dir)
        output.mkdir(parents=True, exist_ok=True)
        json_path = output / "relatorio_validacao.json"
        csv_path = output / "inconsistencias.csv"
        json_path.write_text(
            json.dumps(self.to_dict(), ensure_ascii=False, indent=2, default=str),
            encoding="utf-8",
        )
        with csv_path.open("w", newline="", encoding="utf-8-sig") as handle:
            writer = csv.DictWriter(
                handle,
                fieldnames=(
                    "severity", "blocking", "code", "sheet", "column", "count",
                    "rows", "message", "details",
                ),
            )
            writer.writeheader()
            for issue in self.issues:
                row = asdict(issue)
                row["rows"] = ";".join(map(str, issue.rows))
                row["details"] = json.dumps(issue.details, ensure_ascii=False, default=str)
                writer.writerow(row)
        return json_path, csv_path


def _records(worksheet: Any, headers: tuple[str, ...]) -> list[tuple[int, dict[str, Any]]]:
    actual_headers = [_text(cell.value) for cell in worksheet[1]]
    indexes = {header: index for index, header in enumerate(actual_headers) if header}
    records: list[tuple[int, dict[str, Any]]] = []
    for excel_row, values in enumerate(
        worksheet.iter_rows(min_row=2, max_col=len(actual_headers), values_only=True), start=2
    ):
        if all(_blank(value) for value in values):
            continue
        records.append(
            (
                excel_row,
                {
                    header: values[indexes[header]]
                    if indexes[header] < len(values)
                    else None
                    for header in headers
                },
            )
        )
    return records


def _check_headers(report: ValidationReport, worksheet: Any, expected: tuple[str, ...]) -> bool:
    actual = tuple(_text(cell.value) for cell in worksheet[1])
    missing = [header for header in expected if header not in actual]
    extras = [header for header in actual if header and header not in expected]
    duplicates = [header for header, count in Counter(actual).items() if header and count > 1]
    if missing:
        report.add("CRÍTICA", "COLUNAS_AUSENTES", worksheet.title, None,
                   f"Colunas obrigatórias ausentes: {', '.join(missing)}",
                   details={"missing": missing})
    if extras:
        report.add("MÉDIA", "COLUNAS_EXTRAS", worksheet.title, None,
                   f"Colunas não previstas no contrato: {', '.join(extras)}",
                   details={"extra": extras})
    if duplicates:
        report.add("CRÍTICA", "COLUNAS_DUPLICADAS", worksheet.title, None,
                   f"Cabeçalhos duplicados: {', '.join(duplicates)}",
                   details={"duplicates": duplicates})
    if not missing and actual[:len(expected)] != expected:
        report.add("MÉDIA", "ORDEM_COLUNAS", worksheet.title, None,
                   "As colunas existem, mas não estão na ordem contratada.")
    return not missing and not duplicates


def _check_required(
    report: ValidationReport,
    sheet: str,
    records: list[tuple[int, dict[str, Any]]],
    columns: Iterable[str],
    severity: str = "ALTA",
    blocking: bool = True,
) -> None:
    for column in columns:
        rows = [row for row, record in records if _blank(record[column])]
        if rows:
            report.add(
                severity,
                "CAMPO_OBRIGATORIO_VAZIO",
                sheet,
                column,
                f"{len(rows)} linha(s) sem valor obrigatório em {column}.",
                rows,
                blocking=blocking,
            )


def _check_enums(
    report: ValidationReport,
    sheet: str,
    records: list[tuple[int, dict[str, Any]]],
    enums: dict[str, set[str]],
    blocking: bool = True,
) -> None:
    for column, allowed in enums.items():
        invalid: dict[str, list[int]] = defaultdict(list)
        for row, record in records:
            if not _blank(record[column]) and _norm_upper(record[column]) not in allowed:
                invalid[_text(record[column])].append(row)
        if invalid:
            rows = [row for values in invalid.values() for row in values]
            report.add("ALTA", "VALOR_FORA_DOMINIO", sheet, column,
                       f"{len(rows)} linha(s) com valor fora do domínio de {column}.", rows,
                       {"invalid_values": dict(invalid), "allowed": sorted(allowed)},
                       blocking=blocking)


def _check_unique(
    report: ValidationReport,
    sheet: str,
    records: list[tuple[int, dict[str, Any]]],
    columns: tuple[str, ...],
    severity: str = "CRÍTICA",
    blocking: bool = True,
) -> None:
    groups: dict[tuple[str, ...], list[int]] = defaultdict(list)
    for row, record in records:
        key = tuple(_norm(record[column]) for column in columns)
        if all(key):
            groups[key].append(row)
    duplicated = {" | ".join(key): rows for key, rows in groups.items() if len(rows) > 1}
    if duplicated:
        rows = [row for values in duplicated.values() for row in values]
        report.add(severity, "CHAVE_DUPLICADA", sheet, ", ".join(columns),
                   f"{len(duplicated)} chave(s) duplicada(s) no grão esperado.", rows,
                   {"duplicated_keys": duplicated}, blocking=blocking)


def _check_functional_dependency(
    report: ValidationReport,
    sheet: str,
    records: list[tuple[int, dict[str, Any]]],
    determinant: str,
    dependent: str,
) -> None:
    values: dict[str, dict[str, list[int]]] = defaultdict(lambda: defaultdict(list))
    for row, record in records:
        key, value = _norm(record[determinant]), _norm(record[dependent])
        if key and value:
            values[key][value].append(row)
    conflicts = {key: grouped for key, grouped in values.items() if len(grouped) > 1}
    if conflicts:
        rows = [row for grouped in conflicts.values() for group in grouped.values() for row in group]
        report.add("ALTA", "DEPENDENCIA_INCONSISTENTE", sheet, dependent,
                   f"{determinant} não determina um único {dependent}.", rows,
                   {"conflicting_keys": list(conflicts)}, blocking=False)


def _is_allocating_record(record: dict[str, Any]) -> bool:
    return (
        _norm_upper(record["FORMATO_AULA"]) == "AO VIVO"
        and _norm_upper(record["SINERGIA"]) in {"CURSO ÚNICO", "CURSO RESPONSÁVEL"}
    )


def _validate_mapa(
    report: ValidationReport,
    records: list[tuple[int, dict[str, Any]]],
    expected_module: int | None,
) -> None:
    allocating = [(row, record) for row, record in records if _is_allocating_record(record)]
    supporting = [(row, record) for row, record in records if not _is_allocating_record(record)]
    routing_columns = ("SINERGIA", "FORMATO_AULA")
    allocation_columns = (
        "CURRÍCULO",
        "COD_DISCIPLINA",
        "NOME_DISCIPLINA",
        "PERFIL_DISCIPLINA",
        "MÓD",
    )
    # SINERGIA e FORMATO_AULA determinam se a linha entra no motor e, por isso,
    # precisam existir em todo registro. Os demais campos só bloqueiam quando a
    # oferta realmente é alocável; linhas Sinérgicas/NSA são contexto e não
    # devem impedir a execução por um campo que o motor nunca consumirá.
    _check_required(
        report,
        MAPA_SHEET,
        records,
        routing_columns,
        blocking=True,
    )
    _check_required(
        report,
        MAPA_SHEET,
        allocating,
        allocation_columns,
        blocking=True,
    )
    _check_required(
        report,
        MAPA_SHEET,
        supporting,
        allocation_columns,
        severity="MÉDIA",
        blocking=False,
    )
    remaining_required = tuple(
        header
        for header in MAPA_HEADERS
        if header not in routing_columns + allocation_columns
        and header not in {"HORÁRIO", "MODELO_CONTRATO"}
    )
    # Defeitos restantes que o motor consegue representar como pendência são
    # ressalvas; campos que classificam a oferta foram bloqueados acima.
    _check_required(
        report,
        MAPA_SHEET,
        allocating,
        remaining_required,
        blocking=False,
    )
    _check_required(
        report,
        MAPA_SHEET,
        supporting,
        remaining_required,
        severity="MÉDIA",
        blocking=False,
    )
    _check_enums(report, MAPA_SHEET, records, MAPA_ENUMS)
    _check_profile_format(report, MAPA_SHEET, records)
    _check_unique(report, MAPA_SHEET, records, ("CURRÍCULO", "COD_DISCIPLINA"))
    for determinant, dependent in (
        ("CURSO", "NOME_CURSO"), ("CURSO", "COORDENADOR"),
        ("CURSO", "CLUSTER"), ("CURRÍCULO", "CURSO"),
        ("COD_DISCIPLINA", "NOME_DISCIPLINA"),
    ):
        _check_functional_dependency(report, MAPA_SHEET, records, determinant, dependent)

    for column in ("MATRIZ", "MÓD"):
        rows = [row for row, record in records if not _blank(record[column]) and not _is_integer(record[column])]
        if rows:
            report.add("ALTA", "TIPO_INVALIDO", MAPA_SHEET, column,
                       f"{column} deve ser um número inteiro.", rows)

    module_values = sorted(
        {int(record["MÓD"]) for _, record in records if _is_integer(record["MÓD"])}
    )
    report.metadata["modules"] = module_values
    report.metadata["allocating_rows"] = len(allocating)
    if not allocating:
        report.add(
            "CRÍTICA",
            "SEM_OFERTAS_ALOCAVEIS",
            MAPA_SHEET,
            None,
            "A base não possui Curso Único ou Curso Responsável ao vivo.",
            blocking=True,
        )
    if expected_module is not None:
        unexpected_module_rows = [
            row
            for row, record in records
            if _is_integer(record["MÓD"]) and int(record["MÓD"]) != expected_module
        ]
        if unexpected_module_rows:
            report.add(
                "ALTA",
                "MODULO_DIVERGENTE",
                MAPA_SHEET,
                "MÓD",
                f"A execução esperava somente o módulo {expected_module}.",
                unexpected_module_rows,
                {"expected": expected_module, "found": module_values},
                blocking=True,
            )

    invalid_matrix = [
        row for row, record in records
        if _is_integer(record["MATRIZ"]) and not 1900 <= int(record["MATRIZ"]) <= 2100
    ]
    if invalid_matrix:
        report.add("ALTA", "ANO_MATRIZ_INVALIDO", MAPA_SHEET, "MATRIZ",
                   "MATRIZ deve estar entre 1900 e 2100.", invalid_matrix)

    invalid_time = [
        row for row, record in records
        if not _blank(record["HORÁRIO"]) and not _is_time(record["HORÁRIO"])
    ]
    if invalid_time:
        report.add("ALTA", "HORARIO_INVALIDO", MAPA_SHEET, "HORÁRIO",
                   "HORÁRIO deve usar um horário válido (HH:MM).", invalid_time,
                   blocking=False)

    missing_schedule = [
        row for row, record in allocating
        if _norm_upper(record["DIA_AULA"]) != "NSA" and _blank(record["HORÁRIO"])
    ]
    if missing_schedule:
        report.add("ALTA", "OFERTA_SEM_HORARIO", MAPA_SHEET, "HORÁRIO",
                   "Oferta alocável com dia definido, mas sem horário.",
                   missing_schedule, blocking=False)

    live_without_day = [
        row for row, record in allocating
        if _norm_upper(record["DIA_AULA"]) == "NSA"
    ]
    if live_without_day:
        report.add("ALTA", "AULA_AO_VIVO_SEM_DIA", MAPA_SHEET, "DIA_AULA",
                   "Oferta alocável sem dia utilizável para verificar choque de agenda.",
                   live_without_day, blocking=False)

    proof_mismatch = []
    method_mismatch = []
    for row, record in records:
        order = _norm_upper(record["ORDEM"])
        proof = _norm_upper(record["PROVA"])
        method = _norm_upper(record["METODOLOGIA"])
        if order == "1ª" and proof != "5" or order in {"2ª", "ESTENDIDA"} and proof != "10":
            proof_mismatch.append(row)
        if (order == "ESTENDIDA") != (method == "ESTENDIDA"):
            method_mismatch.append(row)
    if proof_mismatch:
        report.add("ALTA", "PROVA_INCOMPATIVEL_COM_ORDEM", MAPA_SHEET, "PROVA",
                   "PROVA deve ser 5 na 1ª ordem e 10 nas demais.", proof_mismatch,
                   blocking=False)
    if method_mismatch:
        report.add("MÉDIA", "ORDEM_METODOLOGIA_DIVERGENTE", MAPA_SHEET, "METODOLOGIA",
                   "ORDEM e METODOLOGIA divergem quanto à classificação ESTENDIDA; confirmar regra de negócio.",
                   method_mismatch)


def _validate_docentes(report: ValidationReport, records: list[tuple[int, dict[str, Any]]]) -> None:
    _check_required(report, DOCENTES_SHEET, records, DOCENTES_HEADERS)
    _check_enums(report, DOCENTES_SHEET, records, DOCENTES_ENUMS)
    _check_profile_format(report, DOCENTES_SHEET, records)
    _check_unique(report, DOCENTES_SHEET, records, ("CHAPA",))
    _check_unique(
        report,
        DOCENTES_SHEET,
        records,
        ("NOME",),
        severity="MÉDIA",
        blocking=False,
    )

    invalid_badge = [
        row for row, record in records
        if not _blank(record["CHAPA"]) and not re.fullmatch(r"\d{12}", _text(record["CHAPA"]))
    ]
    if invalid_badge:
        report.add("ALTA", "CHAPA_INVALIDA", DOCENTES_SHEET, "CHAPA",
                   "CHAPA deve conter exatamente 12 dígitos, preservando zeros à esquerda.", invalid_badge)

    for column in ("CH_CONTRATADA", "CH_LETIVA"):
        invalid = [
            row for row, record in records
            if not _blank(record[column]) and (not _is_integer(record[column]) or record[column] < 0)
        ]
        if invalid:
            report.add("ALTA", "CARGA_HORARIA_INVALIDA", DOCENTES_SHEET, column,
                       f"{column} deve ser um inteiro não negativo.", invalid)

    overload = [
        row for row, record in records
        if _is_integer(record["CH_CONTRATADA"]) and _is_integer(record["CH_LETIVA"])
        and record["CH_LETIVA"] > record["CH_CONTRATADA"]
    ]
    if overload:
        report.add("ALTA", "CH_LETIVA_SUPERIOR_A_CONTRATADA", DOCENTES_SHEET, "CH_LETIVA",
                   "Carga letiva não pode superar a carga contratada.", overload)


def _check_profile_coverage(
    report: ValidationReport,
    mapa: list[tuple[int, dict[str, Any]]],
    docentes: list[tuple[int, dict[str, Any]]],
) -> None:
    active = [record for _, record in docentes if _norm_upper(record["STATUS"]) == "ATIVO"]
    active_tokens: set[str] = set()
    for record in active:
        active_tokens.update(_profile_tokens(record["PERFIL_DISCIPLINA"]))

    missing: dict[str, list[int]] = defaultdict(list)
    original: dict[str, str] = {}
    for row, record in mapa:
        if not _is_allocating_record(record):
            continue
        tokens = _profile_tokens(record["PERFIL_DISCIPLINA"])
        compatible = False
        for teacher in active:
            function = _norm_upper(teacher["NM_FUNCAO"])
            teacher_family = (
                "STRICTO"
                if function == "PROFESSOR DE ENSINO SUPERIOR PRESENCIAL"
                else "EAD_POOL"
                if function in {
                    "PROFESSOR DE ENSINO SUPERIOR EAD",
                    "PROFESSOR REGENTE",
                }
                else "UNKNOWN"
            )
            capacity_ok = teacher_family == "STRICTO" or (
                _is_integer(teacher["CH_LETIVA"]) and int(teacher["CH_LETIVA"]) >= 2
            )
            if (
                teacher_family != "UNKNOWN"
                and capacity_ok
                and tokens.intersection(_profile_tokens(teacher["PERFIL_DISCIPLINA"]))
            ):
                compatible = True
                break
        if tokens and not compatible:
            key = _norm(record["PERFIL_DISCIPLINA"])
            missing[key].append(row)
            original[key] = _text(record["PERFIL_DISCIPLINA"])
    if missing:
        rows = [row for values in missing.values() for row in values]
        details = {
            "profiles_without_active_teacher": [
                {"profile": original[key], "rows": values} for key, values in sorted(missing.items())
            ],
            "active_profile_token_count": len(active_tokens),
        }
        report.add("ALTA", "PERFIL_SEM_DOCENTE_ATIVO", MAPA_SHEET, "PERFIL_DISCIPLINA",
                   f"{len(rows)} oferta(s), em {len(missing)} perfil(is), não possuem docente ativo compatível.",
                   rows, details, blocking=False)


def validate_workbook(
    path: str | Path,
    expected_module: int | None = None,
) -> ValidationReport:
    source = Path(path)
    report = ValidationReport(str(source))
    report.assumptions = [
        "O grão do MAPA PEDAGÓGICO é CURRÍCULO + COD_DISCIPLINA.",
        "Perfis separados por vírgula são alternativas; basta uma interseção exata após normalizar caixa e espaços.",
        "Somente docentes com STATUS=ATIVO estão disponíveis para cobertura.",
        "MODELO_CONTRATO é aceito para compatibilidade estrutural, mas não é consumido; o contrato de saída deriva do docente selecionado.",
        "Divergência ESTENDIDA entre ORDEM e METODOLOGIA é alerta até confirmação da regra de negócio.",
        "Somente Curso Único e Curso Responsável ao vivo entram na alocação; linhas Sinérgicas são informativas.",
    ]
    report.metadata["expected_module"] = expected_module
    report.metadata["contract_version"] = VALIDATION_CONTRACT_VERSION
    report.metadata["validated_at_utc"] = datetime.now(timezone.utc).isoformat()
    if not source.exists():
        report.add("CRÍTICA", "ARQUIVO_INEXISTENTE", None, None,
                   f"Arquivo não encontrado: {source}")
        report.finalize()
        return report
    digest = hashlib.sha256()
    with source.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    report.metadata["source_sha256"] = digest.hexdigest()
    try:
        workbook = load_workbook(source, data_only=False, read_only=False)
    except Exception as exc:
        report.add("CRÍTICA", "ARQUIVO_ILEGIVEL", None, None,
                   "Não foi possível abrir a planilha.", details={"error": str(exc)})
        report.finalize()
        return report

    for sheet_name in (MAPA_SHEET, DOCENTES_SHEET):
        if sheet_name not in workbook.sheetnames:
            report.add("CRÍTICA", "ABA_AUSENTE", sheet_name, None,
                       f"Aba obrigatória ausente: {sheet_name}")
    if any(issue.code == "ABA_AUSENTE" for issue in report.issues):
        workbook.close()
        report.finalize()
        return report

    mapa_ws, docentes_ws = workbook[MAPA_SHEET], workbook[DOCENTES_SHEET]
    mapa_ok = _check_headers(report, mapa_ws, MAPA_HEADERS)
    docentes_ok = _check_headers(report, docentes_ws, DOCENTES_HEADERS)
    mapa = _records(mapa_ws, MAPA_HEADERS) if mapa_ok else []
    docentes = _records(docentes_ws, DOCENTES_HEADERS) if docentes_ok else []
    report.sheets = {
        MAPA_SHEET: {"rows": len(mapa), "columns": len(MAPA_HEADERS)},
        DOCENTES_SHEET: {"rows": len(docentes), "columns": len(DOCENTES_HEADERS)},
    }
    if not mapa:
        report.add("CRÍTICA", "ABA_SEM_DADOS", MAPA_SHEET, None, "Aba sem registros.")
    if not docentes:
        report.add("CRÍTICA", "ABA_SEM_DADOS", DOCENTES_SHEET, None, "Aba sem registros.")
    if mapa:
        _validate_mapa(report, mapa, expected_module)
    if docentes:
        _validate_docentes(report, docentes)
    if mapa and docentes:
        _check_profile_coverage(report, mapa, docentes)

    formula_rows: dict[str, list[int]] = defaultdict(list)
    for worksheet in (mapa_ws, docentes_ws):
        for row in worksheet.iter_rows(min_row=2):
            if any(cell.data_type == "f" for cell in row):
                formula_rows[worksheet.title].append(row[0].row)
    for sheet, rows in formula_rows.items():
        report.add(
            "ALTA",
            "FORMULA_NA_BASE",
            sheet,
            None,
            "Há fórmulas na área de dados; o motor exige valores materializados.",
            rows,
            blocking=True,
        )

    workbook.close()
    report.finalize()
    return report
