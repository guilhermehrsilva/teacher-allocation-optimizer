from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import load_workbook


def column_map(worksheet) -> dict[str, int]:
    return {
        str(cell.value or "").strip(): cell.column
        for cell in worksheet[1]
        if str(cell.value or "").strip()
    }


def replace_distinct(worksheet, column_name: str, prefix: str) -> None:
    headers = column_map(worksheet)
    if column_name not in headers:
        return
    column = headers[column_name]
    mapping: dict[str, str] = {}
    for row in range(2, worksheet.max_row + 1):
        cell = worksheet.cell(row, column)
        original = str(cell.value or "").strip()
        if not original:
            continue
        if original not in mapping:
            mapping[original] = f"{prefix} {len(mapping) + 1:03d}"
        cell.value = mapping[original]


def replace_badges(worksheet) -> None:
    """Substitui CHAPAs preservando o contrato de exatamente 12 dígitos."""
    headers = column_map(worksheet)
    column = headers.get("CHAPA")
    if column is None:
        return
    originals = {
        str(worksheet.cell(row, column).value or "").strip()
        for row in range(2, worksheet.max_row + 1)
    }
    mapping: dict[str, str] = {}
    next_number = 900_000_000_001
    for row in range(2, worksheet.max_row + 1):
        cell = worksheet.cell(row, column)
        original = str(cell.value or "").strip()
        if not original:
            continue
        if original not in mapping:
            while str(next_number) in originals:
                next_number += 1
            mapping[original] = str(next_number)
            next_number += 1
        cell.value = mapping[original]
        cell.number_format = "@"


def anonymize(source: Path, destination: Path) -> None:
    workbook = load_workbook(source, keep_links=False)
    try:
        replace_distinct(workbook["DOCENTES"], "NOME", "Docente")
        replace_badges(workbook["DOCENTES"])
        replace_distinct(workbook["DOCENTES"], "GESTOR", "Gestor")
        replace_distinct(workbook["MAPA PEDAGÓGICO"], "COORDENADOR", "Coordenação")

        for worksheet in workbook.worksheets:
            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.comment is not None:
                        cell.comment = None
                    if cell.hyperlink is not None:
                        cell.hyperlink = None

        workbook.properties.creator = "Equipe de Alocação Docente"
        workbook.properties.lastModifiedBy = "Equipe de Alocação Docente"
        workbook.properties.title = "Base de demonstração anonimizada"
        workbook.properties.subject = "Homologação interna da vBeta 1.0"
        workbook.properties.description = (
            "Estrutura sintética com identificadores pessoais substituídos."
        )
        destination.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(destination)
    finally:
        workbook.close()


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Anonimiza uma base para demonstração interna.")
    parser.add_argument("--origem", type=Path, required=True)
    parser.add_argument("--destino", type=Path, required=True)
    return parser.parse_args()


if __name__ == "__main__":
    arguments = parse_args()
    anonymize(arguments.origem.resolve(), arguments.destino.resolve())
    print(arguments.destino.resolve())
