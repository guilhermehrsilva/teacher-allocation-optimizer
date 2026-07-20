from __future__ import annotations

import asyncio
import hashlib
import os
import shutil
import json
import subprocess
import sys
import unittest
import uuid
from concurrent.futures import Future, ThreadPoolExecutor
from io import BytesIO
from pathlib import Path
from unittest.mock import patch

from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook

from app.config import Settings
from app.database import Database
from app.main import create_app
from app.security import LocalRequestSecurityMiddleware
from app.services.jobs import build_dashboard
from app.services.jobs import public_job
from app.services.insights import build_insights
from app.services.scenarios import ScenarioError, materialize_scenario
from app.services.uploads import write_validation_workbook


REPO_ROOT = Path(__file__).resolve().parents[2]


def _artifact(path: Path, round_dir: Path) -> dict[str, object]:
    return {
        "path": str(path.relative_to(round_dir)),
        "size_bytes": path.stat().st_size,
        "sha256": hashlib.sha256(path.read_bytes()).hexdigest(),
    }


def _write_manifest(
    round_dir: Path,
    artifacts: dict[str, Path],
    phases: dict | None = None,
) -> None:
    (round_dir / "manifesto.json").write_text(
        json.dumps(
            {
                "artifacts": {
                    key: _artifact(path, round_dir) for key, path in artifacts.items()
                },
                "phases": phases or {},
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )


class BackendApiTests(unittest.TestCase):
    def setUp(self) -> None:
        data_dir = REPO_ROOT / "data" / "test_runs" / uuid.uuid4().hex
        data_dir.mkdir(parents=True, exist_ok=False)
        self.data_dir = data_dir
        self.settings = Settings(
            repo_root=REPO_ROOT,
            tela_dir=REPO_ROOT,
            vbeta_dir=REPO_ROOT / "engines" / "primary",
            data_dir=data_dir,
            allowed_origins=(
                "http://localhost:5173",
                "http://127.0.0.1:5173",
            ),
        )
        self.app = create_app(self.settings)
        self.client_context = TestClient(self.app, base_url="http://localhost")
        self.client = self.client_context.__enter__()

    def tearDown(self) -> None:
        self.client_context.__exit__(None, None, None)
        shutil.rmtree(self.data_dir, ignore_errors=True)

    def test_health_reports_vbeta_and_data_directory(self) -> None:
        self.assertEqual("1.0.0", self.app.version)
        response = self.client.get("/api/health")
        self.assertEqual(200, response.status_code)
        self.assertEqual("ok", response.json()["status"])
        self.assertTrue(response.json()["vbeta_available"])
        self.assertTrue(response.json()["scenario_engine_available"])
        self.assertTrue(response.json()["data_dir_ready"])

    def test_local_request_security_rejects_dns_rebinding_and_browser_csrf(self) -> None:
        for local_host in ("localhost:8000", "127.0.0.1:8000", "[::1]:8000"):
            with self.subTest(local_host=local_host):
                self.assertEqual(
                    200,
                    self.client.get("/api/health", headers={"Host": local_host}).status_code,
                )
        with patch(
            "app.main.reset_saved_scenarios",
            side_effect=AssertionError("destructive endpoint must not run"),
        ):
            rebound = self.client.delete(
                "/api/scenarios?scope=all",
                headers={
                    "Host": "allocation.attacker.example:8000",
                    "Origin": "http://localhost:5173",
                },
            )
            self.assertEqual(400, rebound.status_code)
            self.assertIn("apenas acesso local", rebound.json()["detail"])

            untrusted_origin = self.client.delete(
                "/api/scenarios?scope=all",
                headers={
                    "Origin": "https://attacker.example",
                    "Sec-Fetch-Site": "cross-site",
                },
            )
            self.assertEqual(403, untrusted_origin.status_code)

            forged_fetch_context = self.client.delete(
                "/api/scenarios?scope=all",
                headers={
                    "Origin": "http://localhost",
                    "Sec-Fetch-Site": "cross-site",
                },
            )
            self.assertEqual(403, forged_fetch_context.status_code)

            missing_browser_origin = self.client.delete(
                "/api/scenarios?scope=all",
                headers={"Sec-Fetch-Site": "same-origin"},
            )
            self.assertEqual(403, missing_browser_origin.status_code)

        trusted_frontend = self.client.delete(
            "/api/scenarios?scope=all",
            headers={
                "Origin": "http://localhost:5173",
                "Sec-Fetch-Site": "cross-site",
            },
        )
        self.assertEqual(200, trusted_frontend.status_code)
        self.assertEqual(
            "http://localhost:5173",
            trusted_frontend.headers["access-control-allow-origin"],
        )

    def test_oversized_content_length_is_rejected_before_multipart_parsing(self) -> None:
        with patch(
            "starlette.formparsers.MultiPartParser.parse",
            side_effect=AssertionError("multipart parser must not run"),
        ):
            response = self.client.post(
                "/api/uploads",
                content=b"ignored",
                headers={
                    "Content-Type": "multipart/form-data; boundary=unused",
                    "Content-Length": str(self.settings.max_request_body_bytes + 1),
                },
            )
        self.assertEqual(413, response.status_code)
        self.assertIn("excede", response.json()["detail"])

    def test_scenario_materialization_is_isolated_from_primary_source(self) -> None:
        source = self.data_dir / "baseline.xlsx"
        workbook = Workbook()
        teachers = workbook.active
        teachers.title = "DOCENTES"
        teachers.append(["CHAPA", "NOME", "STATUS", "CH_LETIVA", "PERFIL_DISCIPLINA"])
        teachers.append(["100", "Docente Teste", "ATIVO", 8, "GESTÃO"])
        offers = workbook.create_sheet("MAPA PEDAGÓGICO")
        offers.append([
            "CURSO", "NOME_CURSO", "COD_DISCIPLINA", "NOME_DISCIPLINA",
            "DIA_AULA", "HORÁRIO", "ORDEM", "PERFIL_DISCIPLINA", "CLUSTER",
            "SINERGIA", "FORMATO_AULA",
        ])
        offers.append(["ADM", "Administração", "ADM01", "Gestão", "SEGUNDA", "19:00", "1ª", "GESTÃO", "GESTÃO E NEGÓCIOS", "Curso único", "Ao vivo"])
        offers.append(["ADM", "Administração", "ADM02", "Linha informativa", "SEGUNDA", "20:40", "1ª", "GESTÃO", "GESTÃO E NEGÓCIOS", "Sinérgica", "Ao vivo"])
        workbook.save(source)
        workbook.close()

        round_dir = self.data_dir / "rodada_001"
        source_copy = round_dir / "fonte" / source.name
        summary_path = round_dir / "alocacao" / "resumo_alocacao.json"
        source_copy.parent.mkdir(parents=True)
        summary_path.parent.mkdir(parents=True)
        shutil.copy2(source, source_copy)
        summary_path.write_text(
            json.dumps({
                "transmissions": 1,
                "allocated": 0,
                "unassigned": 1,
                "hours_per_transmission": 2,
                "decisions": [{
                    "source_row": 2,
                    "status": "NAO_ALOCADA",
                    "discipline_code": "ADM01",
                    "discipline_name": "Gestão",
                    "teacher_badge": None,
                    "teacher_name": None,
                }],
            }, ensure_ascii=False),
            encoding="utf-8",
        )
        _write_manifest(
            round_dir,
            {"source_copy": source_copy, "allocation_summary": summary_path},
        )

        now = "2026-07-15T00:00:00+00:00"
        database = self.app.state.database
        database.create_upload({
            "id": "upload-scenario", "original_name": source.name,
            "stored_path": str(source), "module": 52,
            "validation_status": "APROVADO",
            "validation_path": str(self.data_dir / "validation.json"),
            "validation_csv_path": str(self.data_dir / "validation.xlsx"),
            "created_at": now,
        })
        database.create_job({
            "id": "primary-baseline", "upload_id": "upload-scenario",
            "status": "CONCLUIDA", "message": "ok", "module": 52,
            "require_optimal": True, "time_limit_seconds": None,
            "created_at": now, "updated_at": now,
        })
        database.update_job(
            "primary-baseline", round_name="rodada_001", round_dir=str(round_dir),
        )

        created = self.client.post("/api/scenarios", json={
            "baseline_job_id": "primary-baseline",
            "name": "Capacidade futura",
            "description": "Simular aumento de capacidade sem tocar no solver principal.",
        })
        self.assertEqual(201, created.status_code)
        scenario_id = created.json()["id"]
        catalog = self.client.get(f"/api/scenarios/{scenario_id}/catalog")
        self.assertEqual(200, catalog.status_code)
        self.assertEqual([2], [item["row_number"] for item in catalog.json()["offers"]])
        change = self.client.post(f"/api/scenarios/{scenario_id}/changes", json={
            "change_type": "CAPACIDADE", "entity_type": "teacher",
            "row_number": 2, "field_name": "CH_LETIVA", "new_value": 12,
        })
        self.assertEqual(201, change.status_code)
        change_id = change.json()["id"]
        rejected_supporting_row = self.client.post(f"/api/scenarios/{scenario_id}/changes", json={
            "change_type": "AGENDA", "entity_type": "offer",
            "row_number": 3, "field_name": "DIA_AULA", "new_value": "TERÇA",
        })
        self.assertEqual(409, rejected_supporting_row.status_code)
        self.assertIn("ofertas tratadas pelo solver", rejected_supporting_row.json()["detail"])
        policy = self.client.post(f"/api/scenarios/{scenario_id}/policies", json={
            "policy_type": "ALOCAR_CLUSTER", "target_type": "CLUSTER",
            "target_value": "GESTÃO E NEGÓCIOS", "configuration": {},
        })
        self.assertEqual(201, policy.status_code)
        policy_id = policy.json()["id"]
        self.assertEqual([2], policy.json()["configuration"]["baseline_unassigned_rows"])

        unsupported = self.client.post(f"/api/scenarios/{scenario_id}/policies", json={
            "policy_type": "INTERNALIZACAO", "target_type": "CLUSTER",
            "target_value": "GESTÃO E NEGÓCIOS", "configuration": {},
        })
        self.assertEqual(409, unsupported.status_code)

        # A origem operacional pode mudar; o cenário deve partir da cópia
        # imutável registrada no manifesto da rodada-base.
        mutable = load_workbook(source)
        mutable["DOCENTES"]["D2"] = 99
        mutable.save(source)
        mutable.close()
        materialized, snapshot = materialize_scenario(
            self.settings, database, scenario_id, "scenario-job-test",
        )

        original_book = load_workbook(source, data_only=True)
        scenario_book = load_workbook(materialized, data_only=True)
        try:
            self.assertEqual(99, original_book["DOCENTES"]["D2"].value)
            self.assertEqual(12, scenario_book["DOCENTES"]["D2"].value)
        finally:
            original_book.close()
            scenario_book.close()
        self.assertTrue(snapshot.is_file())
        snapshot_payload = json.loads(snapshot.read_text(encoding="utf-8"))
        self.assertEqual("ALOCAR_CLUSTER", snapshot_payload["policies"][0]["policy_type"])
        self.assertEqual("PRIMARY", self.client.get("/api/jobs").json()[0]["kind"])

        # A reserva de execução é atômica dentro do processo único.
        manager = self.app.state.scenario_manager
        pending_future: Future[None] = Future()
        with patch.object(manager.executor, "submit", return_value=pending_future):
            def reserve() -> tuple[str, str]:
                try:
                    return ("ok", manager.create_run(scenario_id)["id"])
                except ScenarioError as exc:
                    return ("error", str(exc))

            with ThreadPoolExecutor(max_workers=2) as pool:
                reservations = list(pool.map(lambda _: reserve(), range(2)))
        self.assertEqual(1, sum(kind == "ok" for kind, _ in reservations))
        self.assertEqual(1, sum(kind == "error" for kind, _ in reservations))
        scenario_job_id = next(value for kind, value in reservations if kind == "ok")

        # Enquanto executa, premissas não podem ser removidas.
        blocked = self.client.delete(
            f"/api/scenarios/{scenario_id}/changes/{change_id}"
        )
        self.assertEqual(409, blocked.status_code)

        database.update_job(scenario_job_id, status="CONCLUIDA")
        database.update_scenario(scenario_id, status="CONCLUIDO")
        deleted_policy = self.client.delete(
            f"/api/scenarios/{scenario_id}/policies/{policy_id}"
        )
        self.assertEqual(204, deleted_policy.status_code)
        self.assertEqual("RASCUNHO", database.get_scenario(scenario_id)["status"])

        database.update_scenario(scenario_id, status="CONCLUIDO")
        deleted_change = self.client.delete(
            f"/api/scenarios/{scenario_id}/changes/{change_id}"
        )
        self.assertEqual(204, deleted_change.status_code)
        self.assertEqual("RASCUNHO", database.get_scenario(scenario_id)["status"])
        stale_comparison = self.client.get(f"/api/scenarios/{scenario_id}/comparison")
        self.assertEqual(409, stale_comparison.status_code)

        reset_scenario = self.client.delete("/api/scenarios", params={"scope": "latest"})
        self.assertEqual(200, reset_scenario.status_code)
        self.assertEqual(1, reset_scenario.json()["deleted_scenarios"])
        self.assertEqual([], self.client.get("/api/scenarios").json())
        self.assertTrue(source.is_file())

        reset_baseline = self.client.delete("/api/jobs/primary", params={"scope": "latest"})
        self.assertEqual(200, reset_baseline.status_code)
        self.assertEqual(1, reset_baseline.json()["deleted_rounds"])
        self.assertEqual([], self.client.get("/api/jobs").json())

    def test_frontend_html_disables_browser_cache(self) -> None:
        response = self.client.get("/")
        self.assertEqual(200, response.status_code)
        self.assertEqual(
            "no-store, no-cache, must-revalidate, max-age=0",
            response.headers["cache-control"],
        )

    def test_upload_rejects_non_xlsx_file(self) -> None:
        response = self.client.post(
            "/api/uploads",
            data={"module": "52"},
            files={"file": ("base.csv", b"x,y", "text/csv")},
        )
        self.assertEqual(422, response.status_code)
        self.assertIn(".xlsx", response.json()["detail"])

    def test_upload_requires_explicit_module(self) -> None:
        response = self.client.post(
            "/api/uploads",
            files={
                "file": (
                    "base.xlsx",
                    b"not-needed",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        self.assertEqual(422, response.status_code)
        self.assertIn("module", str(response.json()["detail"]))

    def test_invalid_workbook_does_not_leave_partial_upload(self) -> None:
        before = set(self.settings.upload_dir.iterdir())
        engine_src = str(self.settings.vbeta_dir / "src")
        if engine_src not in sys.path:
            sys.path.insert(0, engine_src)
        with patch(
            "alocacao_docente.validation.validate_workbook",
            side_effect=RuntimeError("arquivo corrompido"),
        ):
            response = self.client.post(
                "/api/uploads",
                data={"module": "52"},
                files={
                    "file": (
                        "base.xlsx",
                        b"this is not an xlsx archive",
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                },
            )
        self.assertEqual(422, response.status_code)
        self.assertEqual(before, set(self.settings.upload_dir.iterdir()))

    def test_rejected_workbook_keeps_report_but_discards_source(self) -> None:
        workbook = Workbook()
        buffer = BytesIO()
        workbook.save(buffer)
        workbook.close()

        response = self.client.post(
            "/api/uploads",
            data={"module": "52"},
            files={
                "file": (
                    "incompleta.xlsx",
                    buffer.getvalue(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )

        self.assertEqual(200, response.status_code)
        payload = response.json()
        self.assertEqual("REPROVADO", payload["validation"]["status"])
        upload = self.app.state.database.get_upload(payload["id"])
        self.assertIsNotNone(upload)
        assert upload is not None
        self.assertFalse(Path(upload["stored_path"]).exists())
        self.assertTrue(Path(upload["validation_path"]).is_file())
        report = self.client.get(f"/api/uploads/{payload['id']}/validation.xlsx")
        self.assertEqual(200, report.status_code)

    def test_upload_quota_and_startup_orphan_cleanup_are_enforced(self) -> None:
        data_dir = self.data_dir / "quota-instance"
        settings = Settings(
            repo_root=REPO_ROOT,
            tela_dir=REPO_ROOT,
            vbeta_dir=REPO_ROOT / "engines" / "primary",
            data_dir=data_dir,
            max_upload_bytes=100,
            max_upload_storage_bytes=10,
            multipart_overhead_bytes=1024,
            orphan_upload_retention_seconds=0,
        )
        settings.ensure_directories()
        orphan = settings.upload_dir / "not-in-database"
        orphan.mkdir()
        (orphan / "payload.bin").write_bytes(b"orphan")
        stale = settings.upload_dir / "stale-upload"
        stale.mkdir()
        stale_source = stale / "fonte.xlsx"
        stale_source.write_bytes(b"legacy rejected source")
        database = Database(settings.db_path)
        database.initialize()
        database.create_upload({
            "id": "stale-upload", "original_name": "stale.xlsx",
            "stored_path": str(stale_source), "module": 52,
            "validation_status": "REPROVADO",
            "validation_path": str(stale / "validation.json"),
            "validation_csv_path": str(stale / "validation.xlsx"),
            "created_at": "2026-07-01T00:00:00+00:00",
        })

        app = create_app(settings)
        self.assertFalse(orphan.exists())
        self.assertFalse(stale.exists())
        self.assertIsNone(app.state.database.get_upload("stale-upload"))
        with TestClient(app, base_url="http://localhost") as client:
            response = client.post(
                "/api/uploads",
                data={"module": "52"},
                files={
                    "file": (
                        "base.xlsx",
                        b"x" * 20,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                },
            )
        self.assertEqual(507, response.status_code)
        self.assertIn("cota", response.json()["detail"].casefold())
        self.assertEqual([], list(settings.upload_dir.iterdir()))

    def test_primary_and_scenario_queues_have_hard_caps(self) -> None:
        database = self.app.state.database
        now = "2026-07-16T00:00:00+00:00"
        database.create_upload({
            "id": "queue-upload", "original_name": "base.xlsx",
            "stored_path": str(self.data_dir / "base.xlsx"), "module": 52,
            "validation_status": "APROVADO",
            "validation_path": str(self.data_dir / "validation.json"),
            "validation_csv_path": str(self.data_dir / "validation.xlsx"),
            "created_at": now,
        })
        database.create_job({
            "id": "queue-baseline", "upload_id": "queue-upload",
            "status": "CONCLUIDA", "message": "ok", "module": 52,
            "require_optimal": True, "time_limit_seconds": None,
            "created_at": now, "updated_at": now,
        })
        database.create_scenario({
            "id": "queue-target-scenario",
            "baseline_job_id": "queue-baseline",
            "name": "Cenário com fila cheia",
            "description": "",
            "status": "RASCUNHO",
            "created_at": now,
            "updated_at": now,
        })
        for index in range(self.settings.max_primary_inflight_jobs):
            database.create_job({
                "id": f"queued-primary-{index}", "upload_id": "queue-upload",
                "status": "QUEUED", "message": "queued", "module": 52,
                "require_optimal": True, "time_limit_seconds": None,
                "created_at": now, "updated_at": now,
            })
        for index in range(self.settings.max_scenario_inflight_jobs):
            database.create_job({
                "id": f"queued-scenario-{index}", "upload_id": "queue-upload",
                "status": "QUEUED", "message": "queued", "module": 52,
                "require_optimal": True, "time_limit_seconds": None,
                "kind": "SCENARIO", "scenario_id": f"other-{index}",
                "created_at": now, "updated_at": now,
            })

        primary = self.client.post("/api/jobs", json={"upload_id": "queue-upload"})
        scenario = self.client.post("/api/scenarios/queue-target-scenario/runs")

        self.assertEqual(429, primary.status_code)
        self.assertEqual("5", primary.headers["retry-after"])
        self.assertEqual(429, scenario.status_code)
        self.assertEqual("5", scenario.headers["retry-after"])
        self.assertEqual(
            self.settings.max_primary_inflight_jobs,
            database.count_inflight_jobs("PRIMARY"),
        )
        self.assertEqual(
            self.settings.max_scenario_inflight_jobs,
            database.count_inflight_jobs("SCENARIO"),
        )

    def test_create_job_and_reset_are_one_atomic_lifecycle_operation(self) -> None:
        database = self.app.state.database
        upload_dir = self.settings.upload_dir / "race-upload"
        upload_dir.mkdir()
        source = upload_dir / "fonte.xlsx"
        source.write_bytes(b"source")
        now = "2026-07-16T00:00:00+00:00"
        database.create_upload({
            "id": "race-upload", "original_name": "base.xlsx",
            "stored_path": str(source), "module": 52,
            "validation_status": "APROVADO",
            "validation_path": str(upload_dir / "validation.json"),
            "validation_csv_path": str(upload_dir / "validation.xlsx"),
            "created_at": now,
        })
        database.create_job({
            "id": "race-old-job", "upload_id": "race-upload",
            "status": "CONCLUIDA", "message": "ok", "module": 52,
            "require_optimal": True, "time_limit_seconds": None,
            "created_at": now, "updated_at": now,
        })
        manager = self.app.state.job_manager
        pending: Future[None] = Future()

        with patch.object(manager.executor, "submit", return_value=pending):
            with ThreadPoolExecutor(max_workers=2) as pool:
                create_future = pool.submit(
                    self.client.post,
                    "/api/jobs",
                    json={"upload_id": "race-upload"},
                )
                reset_future = pool.submit(
                    self.client.delete,
                    "/api/jobs/primary?scope=all",
                )
                create_response = create_future.result()
                reset_response = reset_future.result()

        self.assertIn(
            (create_response.status_code, reset_response.status_code),
            {(202, 409), (404, 200)},
        )
        if create_response.status_code == 202:
            created_job = database.get_job(create_response.json()["id"])
            self.assertIsNotNone(created_job)
            self.assertEqual("QUEUED", created_job["status"])
            self.assertTrue(source.is_file())
        else:
            self.assertIsNone(database.get_upload("race-upload"))
            self.assertFalse(upload_dir.exists())

    def test_shutdown_terminates_children_before_returning(self) -> None:
        creationflags = subprocess.CREATE_NO_WINDOW if os.name == "nt" else 0
        primary_child = subprocess.Popen(
            [sys.executable, "-c", "import time; time.sleep(60)"],
            creationflags=creationflags,
        )
        scenario_child = subprocess.Popen(
            [sys.executable, "-c", "import time; time.sleep(60)"],
            creationflags=creationflags,
        )
        try:
            primary_manager = self.app.state.job_manager
            scenario_manager = self.app.state.scenario_manager
            with primary_manager._state_lock:
                primary_manager._processes["shutdown-proof"] = primary_child
            with scenario_manager._state_lock:
                scenario_manager._processes["shutdown-proof"] = scenario_child

            primary_manager.shutdown()
            scenario_manager.shutdown()

            self.assertIsNotNone(primary_child.poll())
            self.assertIsNotNone(scenario_child.poll())
        finally:
            for child in (primary_child, scenario_child):
                if child.poll() is None:
                    child.kill()
                    child.wait()

    def test_importing_factory_has_no_operational_database_side_effect(self) -> None:
        isolated_data = self.data_dir / "factory-import-must-not-create"
        environment = os.environ.copy()
        environment["ALOCACAO_DATA_DIR"] = str(isolated_data)
        result = subprocess.run(
            [sys.executable, "-c", "import app.main"],
            cwd=REPO_ROOT / "backend",
            env=environment,
            capture_output=True,
            text=True,
            timeout=30,
        )
        self.assertEqual(0, result.returncode, result.stderr)
        self.assertFalse(isolated_data.exists())

    def test_terminal_database_status_wins_over_stale_round_status(self) -> None:
        round_dir = self.data_dir / "stale-status"
        round_dir.mkdir()
        (round_dir / "status.json").write_text(
            json.dumps({"state": "RUNNING", "message": "antigo"}),
            encoding="utf-8",
        )
        job = {
            "id": "job-stale", "upload_id": "upload-stale",
            "original_name": "base.xlsx", "module": 52,
            "status": "INTERROMPIDA", "message": "interrompida no banco",
            "validation_status": "APROVADO", "require_optimal": 1,
            "time_limit_seconds": None, "kind": "PRIMARY",
            "scenario_id": None, "is_official": 0,
            "round_name": "rodada_001", "round_dir": str(round_dir),
            "exit_code": 1, "created_at": "2026-07-15T00:00:00+00:00",
            "updated_at": "2026-07-15T00:00:01+00:00",
        }
        result = public_job(job)
        self.assertEqual("INTERROMPIDA", result["status"])
        self.assertEqual("interrompida no banco", result["message"])
        self.assertTrue(result["terminal"])

    def test_upload_accepts_only_supported_modules(self) -> None:
        response = self.client.post(
            "/api/uploads",
            data={"module": "50"},
            files={"file": ("base.xlsx", b"not-needed", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        self.assertEqual(422, response.status_code)
        self.assertIn("greater than or equal to 51", str(response.json()["detail"]))

    def test_validation_issues_download_is_actionable_xlsx(self) -> None:
        report = {
            "status": "APROVADO_COM_RESSALVAS",
            "metadata": {"expected_module": 52, "allocating_rows": 10},
            "summary": {"issue_groups": 1, "blocking_issue_groups": 0},
            "issues": [{
                "severity": "ALTA",
                "blocking": False,
                "code": "PERFIL_SEM_DOCENTE_ATIVO",
                "sheet": "MAPA PEDAGÓGICO",
                "column": "PERFIL_DISCIPLINA",
                "count": 2,
                "rows": [4, 8],
                "message": "Duas ofertas não possuem docente ativo compatível.",
                "details": {
                    "profiles_without_active_teacher": [{
                        "profile": "Comunicação - Relações Públicas",
                        "rows": [4, 8],
                    }],
                    "active_profile_token_count": 111,
                },
            }],
        }
        workbook_path = write_validation_workbook(report, self.data_dir, "base.xlsx")
        self.app.state.database.create_upload({
            "id": "upload-xlsx", "original_name": "base.xlsx",
            "stored_path": str(self.data_dir / "base.xlsx"), "module": 52,
            "validation_status": "APROVADO_COM_RESSALVAS",
            "validation_path": str(self.data_dir / "validation.json"),
            "validation_csv_path": str(workbook_path),
            "created_at": "2026-07-14T00:00:00+00:00",
        })

        response = self.client.get("/api/uploads/upload-xlsx/validation.xlsx")

        self.assertEqual(200, response.status_code)
        self.assertIn("pendencias_validacao.xlsx", response.headers["content-disposition"])
        workbook = load_workbook(BytesIO(response.content))
        try:
            self.assertEqual(["PENDENCIAS", "RESUMO"], workbook.sheetnames)
            worksheet = workbook["PENDENCIAS"]
            self.assertEqual("STATUS DA CORREÇÃO", worksheet["A1"].value)
            self.assertEqual("PENDENTE", worksheet["A2"].value)
            self.assertEqual("MAPA PEDAGÓGICO", worksheet["F2"].value)
            self.assertEqual("4; 8", worksheet["I2"].value)
            self.assertEqual("INSIGHT ACIONÁVEL", worksheet["K1"].value)
            self.assertIn("Comunicação - Relações Públicas", worksheet["K2"].value)
            self.assertIn("Ação sugerida", worksheet["K2"].value)
            self.assertNotIn("profiles_without_active_teacher", worksheet["K2"].value)
            self.assertNotIn("{", worksheet["K2"].value)
            self.assertEqual(1, len(worksheet.data_validations.dataValidation))
        finally:
            workbook.close()

    def test_dashboard_route_accepts_repeated_filters(self) -> None:
        round_dir = self.data_dir / "fixture_round"
        allocation_dir = round_dir / "alocacao"
        allocation_dir.mkdir(parents=True)

        workbook = Workbook()
        allocations = workbook.active
        allocations.title = "ALOCACOES"
        allocations.append([
            "STATUS", "MOTIVO", "ORDEM", "CURSO", "NOME_CURSO",
            "CLUSTER", "DIA_AULA", "HORÁRIO", "CHAPA",
        ])
        allocations.append(["ALOCADA", None, "1ª", "ADM", "Administração", "GESTÃO", "SEGUNDA", "19:00", "1"])
        allocations.append(["ALOCADA", None, "2ª", "ADM", "Administração", "GESTÃO", "TERÇA", "19:00", "1"])
        allocations.append(["ALOCADA", None, "ESTENDIDA", "ENG", "Engenharia", "ENGENHARIAS", "SEGUNDA", "20:40", "2"])
        allocations.append(["NAO_ALOCADA", "SEM_CANDIDATO", "1ª", "ADM", "Administração", "GESTÃO", "QUARTA", "19:00", None])
        teachers = workbook.create_sheet("DOCENTES")
        teachers.append(["STATUS", "NM_FUNCAO", "CH_LETIVA"])
        teachers.append(["ATIVO", "PROFESSOR REGENTE", 20])
        workbook.save(allocation_dir / "resultado_alocacao.xlsx")

        workbook_path = allocation_dir / "resultado_alocacao.xlsx"
        summary_path = allocation_dir / "resumo_alocacao.json"
        summary_path.write_text(
            '{"hours_per_transmission":2,"solver_status":"OPTIMAL","unassigned":1}',
            encoding="utf-8",
        )
        _write_manifest(
            round_dir,
            {
                "allocation_workbook": workbook_path,
                "allocation_summary": summary_path,
            },
            {"validation": {"status": "APROVADO"}, "audit": {"status": "APROVADO"}},
        )

        database = self.app.state.database
        now = "2026-07-14T00:00:00+00:00"
        database.create_upload({
            "id": "upload-multifilter", "original_name": "base.xlsx",
            "stored_path": str(self.data_dir / "base.xlsx"), "module": 52,
            "validation_status": "APROVADO",
            "validation_path": str(self.data_dir / "validation.json"),
            "validation_csv_path": str(self.data_dir / "validation.csv"),
            "created_at": now,
        })
        database.create_job({
            "id": "job-multifilter", "upload_id": "upload-multifilter",
            "status": "CONCLUIDA", "message": "ok", "module": 52,
            "require_optimal": True, "time_limit_seconds": None,
            "created_at": now, "updated_at": now,
        })
        database.update_job(
            "job-multifilter", round_name="rodada_001", round_dir=str(round_dir),
        )

        response = self.client.get(
            "/api/dashboard/job-multifilter",
            params=[("order", "1ª"), ("order", "2ª"), ("course", "ADM")],
        )

        self.assertEqual(200, response.status_code)
        payload = response.json()
        self.assertEqual(3, payload["kpis"]["transmissions"])
        self.assertEqual(1, payload["kpis"]["unassigned"])
        self.assertEqual(
            {"order": ["1ª", "2ª"], "course": ["ADM"]},
            payload["filters"]["selected"],
        )

        summary_path.write_text('{"unassigned":999}', encoding="utf-8")
        tampered = self.client.get("/api/jobs/job-multifilter/summary")
        self.assertEqual(409, tampered.status_code)
        self.assertIn("manifesto", tampered.json()["detail"])


class SecurityMiddlewareTests(unittest.TestCase):
    def test_streamed_body_without_content_length_is_still_limited(self) -> None:
        async def exercise() -> list[dict]:
            async def downstream(scope, receive, send) -> None:
                while True:
                    message = await receive()
                    if not message.get("more_body", False):
                        break
                await send({"type": "http.response.start", "status": 204, "headers": []})
                await send({"type": "http.response.body", "body": b""})

            middleware = LocalRequestSecurityMiddleware(
                downstream,
                allowed_hosts=("localhost",),
                allowed_origins=(),
                max_body_bytes=8,
            )
            messages = iter([
                {"type": "http.request", "body": b"12345", "more_body": True},
                {"type": "http.request", "body": b"67890", "more_body": False},
            ])
            sent: list[dict] = []

            async def receive() -> dict:
                return next(messages)

            async def send(message: dict) -> None:
                sent.append(message)

            await middleware(
                {
                    "type": "http",
                    "http_version": "1.1",
                    "method": "POST",
                    "scheme": "http",
                    "path": "/api/jobs",
                    "raw_path": b"/api/jobs",
                    "query_string": b"",
                    "headers": [(b"host", b"localhost")],
                    "server": ("localhost", 80),
                    "client": ("127.0.0.1", 12345),
                },
                receive,
                send,
            )
            return sent

        sent = asyncio.run(exercise())
        start = next(message for message in sent if message["type"] == "http.response.start")
        self.assertEqual(413, start["status"])


class DashboardContractTests(unittest.TestCase):
    def test_dashboard_reconciles_core_metrics(self) -> None:
        round_dir = REPO_ROOT / "data" / "test_runs" / uuid.uuid4().hex
        allocation_dir = round_dir / "alocacao"
        allocation_dir.mkdir(parents=True, exist_ok=False)
        try:
            workbook = Workbook()
            allocations = workbook.active
            allocations.title = "ALOCACOES"
            allocations.append(["STATUS", "MOTIVO", "ORDEM", "CURSO", "NOME_CURSO", "CLUSTER", "DIA_AULA", "HORÁRIO", "CHAPA"])
            allocations.append(["ALOCADA", None, "1ª", "ADM", "Administração", "GESTÃO", "SEGUNDA", "19:00", "1"])
            allocations.append(["ALOCADA", None, "2ª", "ADM", "Administração", "GESTÃO", "TERÇA", "19:00", "1"])
            allocations.append(["ALOCADA", None, "ESTENDIDA", "ENG", "Engenharia", "ENGENHARIAS", "SEGUNDA", "20:40", "2"])
            allocations.append(["NAO_ALOCADA", "CHOQUE_DE_HORARIO", "1ª", "ADM", "Administração", "GESTÃO", "SEGUNDA", "19:00", None])
            for index in range(3, 8):
                allocations.append(["NAO_ALOCADA", "CHOQUE_DE_HORARIO", "1ª", "ADM", "Administração", f"CLUSTER {index}", "SEGUNDA", "19:00", None])
            teachers = workbook.create_sheet("DOCENTES")
            teachers.append(["STATUS", "NM_FUNCAO", "CH_LETIVA"])
            teachers.append(["ATIVO", "PROFESSOR REGENTE", 20])
            teachers.append(["ATIVO", "PROFESSOR DE ENSINO SUPERIOR PRESENCIAL", 9])
            workbook.save(allocation_dir / "resultado_alocacao.xlsx")
            _write_manifest(
                round_dir,
                {"allocation_workbook": allocation_dir / "resultado_alocacao.xlsx"},
                {
                    "validation": {"status": "APROVADO"},
                    "audit": {"status": "APROVADO"},
                },
            )
            job = {
                "id": "job-1",
                "upload_id": "upload-1",
                "original_name": "base.xlsx",
                "module": 52,
                "status": "CONCLUIDA",
                "message": "ok",
                "validation_status": "APROVADO",
                "require_optimal": 1,
                "time_limit_seconds": None,
                "round_name": "rodada_001",
                "round_dir": str(round_dir),
                "exit_code": 0,
                "created_at": "2026-07-14T00:00:00+00:00",
                "updated_at": "2026-07-14T00:00:01+00:00",
            }
            summary = {
                "hours_per_transmission": 2,
                "solver_status": "OPTIMAL",
                "wall_time_seconds": 12.5,
            }
            dashboard = build_dashboard(job, summary)
            self.assertEqual(33.33, dashboard["kpis"]["coverage_pct"])
            self.assertEqual(100.0, dashboard["kpis"]["teacher_use_pct"])
            self.assertEqual(6, dashboard["unassigned_reasons"][0]["count"])
            self.assertEqual(16, dashboard["kpis"]["first_stage_demand_hours"])
            self.assertEqual(28, dashboard["kpis"]["active_teaching_capacity_hours"])
            self.assertEqual(-12, dashboard["kpis"]["first_stage_capacity_delta_hours"])
            self.assertEqual(7, len(dashboard["charts"]["demand_hours_by_cluster"]))
            self.assertNotIn("OUTROS", {item["cluster"] for item in dashboard["charts"]["demand_hours_by_cluster"]})
            filtered = build_dashboard(job, summary, {"course": "ENG"})
            self.assertEqual(1, filtered["kpis"]["transmissions"])
            self.assertEqual(0, filtered["kpis"]["unassigned"])
            self.assertEqual(2, filtered["charts"]["demand_hours_by_cluster"][0]["hours"])
            self.assertEqual({"course": ["ENG"]}, filtered["filters"]["selected"])
            self.assertEqual("APROVADO", dashboard["guardrails"]["audit"])
        finally:
            shutil.rmtree(round_dir, ignore_errors=True)


class InsightsContractTests(unittest.TestCase):
    def test_insights_reconcile_pareto_risk_and_stage_exposure(self) -> None:
        round_dir = REPO_ROOT / "data" / "test_runs" / uuid.uuid4().hex
        allocation_dir = round_dir / "alocacao"
        allocation_dir.mkdir(parents=True, exist_ok=False)
        try:
            workbook = Workbook()
            allocations = workbook.active
            allocations.title = "ALOCACOES"
            allocations.append([
                "STATUS", "MOTIVO", "ORDEM", "CLUSTER", "COORDENADOR",
                "MODELO_CONTRATO", "DOCENTE", "DIA_AULA", "CANDIDATOS_ELEGÍVEIS",
                "CURSO", "NOME_CURSO", "HORÁRIO", "PERFIL_DISCIPLINA",
            ])
            for teacher, order, day, candidates in [
                ("A", "1ª", "SEGUNDA", 1), ("A", "1ª", "SEGUNDA", 2),
                ("A", "2ª", "TERÇA", 2), ("A", "ESTENDIDA", "QUARTA", 2),
                ("B", "1ª", "SEGUNDA", 2), ("B", "2ª", "TERÇA", 2),
                ("C", "1ª", "SEGUNDA", 2), ("C", "2ª", "TERÇA", 2),
                ("D", "1ª", "QUARTA", 2), ("E", "2ª", "QUINTA", 2),
            ]:
                allocations.append(["ALOCADA", None, order, "CLUSTER 1", "GESTOR", "CLT EAD", teacher, day, candidates, "CURSO 1", "Curso 1", "19:00", "PERFIL 1"])
            allocations.append(["NAO_ALOCADA", "SEM_CANDIDATO", "1ª", "CLUSTER 2", "GESTOR", "CLT EAD", None, "SEXTA", 0, "CURSO 2", "Curso 2", "21:00", "PERFIL 2"])

            teachers = workbook.create_sheet("DOCENTES")
            teachers.append([
                "NOME", "NM_FUNCAO", "STATUS", "UTILIZAÇÃO_1ª_ETAPA", "UTILIZAÇÃO_2ª_ETAPA",
            ])
            for name, p1, p2 in [("A", 1, .5), ("B", .5, .5), ("C", .4, .4), ("D", .2, 0), ("E", 0, .2)]:
                teachers.append([name, "PROFESSOR", "ATIVO", p1, p2])
            workbook.save(allocation_dir / "resultado_alocacao.xlsx")
            _write_manifest(
                round_dir,
                {"allocation_workbook": allocation_dir / "resultado_alocacao.xlsx"},
            )
            job = {
                "id": "job-insights", "upload_id": "upload-1", "original_name": "base.xlsx",
                "module": 52, "status": "CONCLUIDA", "message": "ok",
                "validation_status": "APROVADO", "require_optimal": 1,
                "time_limit_seconds": None, "round_name": "rodada_001",
                "round_dir": str(round_dir), "exit_code": 0,
                "created_at": "2026-07-14T00:00:00+00:00",
                "updated_at": "2026-07-14T00:00:01+00:00",
            }

            result = build_insights(job)

            self.assertEqual(40.0, result["kpis"]["top_20_teacher_share_pct"])
            self.assertEqual(3, result["kpis"]["pareto_teacher_count_80"])
            self.assertEqual(1, result["kpis"]["single_candidate_allocations"])
            self.assertEqual(6, result["stage_load"]["first_stage_allocations"])
            self.assertEqual(5, result["stage_load"]["second_stage_allocations"])
            self.assertEqual("SEM_CANDIDATO", result["breakdowns"]["unassigned_reasons"][0]["label"])
            capacity_opportunity = next(
                item for item in result["opportunities"] if item["kind"] == "capacity"
            )
            self.assertEqual("5 docentes sem alocação", capacity_opportunity["metric"])
            self.assertNotIn("saldo médio", capacity_opportunity["metric"].lower())
            self.assertEqual(
                [{"day": "SEXTA", "count": 1}],
                result["coverage"]["courses"][0]["gap_days"],
            )
            congested_slot = result["diagnostics"]["most_congested_unassigned_slots"][0]
            self.assertEqual("SEXTA · 21:00", congested_slot["slot"])
            self.assertEqual(1, congested_slot["affected_courses"])
            self.assertEqual("Curso 2", congested_slot["top_course"])
            self.assertEqual("PERFIL 2", congested_slot["top_profile"])
            self.assertEqual("CLUSTER 2", result["coverage"]["clusters"][0]["cluster"])
            self.assertEqual(1, result["coverage"]["clusters"][0]["unassigned"])
        finally:
            shutil.rmtree(round_dir, ignore_errors=True)

    def test_insights_use_badge_for_homonyms_and_pareto_reaches_eighty_percent(self) -> None:
        round_dir = REPO_ROOT / "data" / "test_runs" / uuid.uuid4().hex
        allocation_dir = round_dir / "alocacao"
        allocation_dir.mkdir(parents=True, exist_ok=False)
        try:
            workbook = Workbook()
            allocations = workbook.active
            allocations.title = "ALOCACOES"
            allocations.append([
                "STATUS", "MOTIVO", "ORDEM", "CLUSTER", "COORDENADOR",
                "MODELO_CONTRATO", "CHAPA", "DOCENTE", "DIA_AULA",
                "CANDIDATOS_ELEGÍVEIS", "CURSO", "NOME_CURSO", "HORÁRIO",
                "PERFIL_DISCIPLINA",
            ])
            for badge, name, count in (("1", "Docente Homônimo", 6), ("2", "Docente Homônimo", 3), ("3", "Outro Docente", 1)):
                for index in range(count):
                    allocations.append([
                        "ALOCADA", None, "1ª", "GESTÃO", "GESTOR", "CLT EAD",
                        badge, name, "SEGUNDA" if index % 2 == 0 else "TERÇA", 2,
                        "ADM", "Administração", "19:00", "GESTÃO",
                    ])
            teachers = workbook.create_sheet("DOCENTES")
            teachers.append([
                "CHAPA", "NOME", "NM_FUNCAO", "STATUS",
                "UTILIZAÇÃO_1ª_ETAPA", "UTILIZAÇÃO_2ª_ETAPA",
            ])
            teachers.append(["1", "Docente Homônimo", "PROFESSOR", "ATIVO", .8, .4])
            teachers.append(["2", "Docente Homônimo", "PROFESSOR", "ATIVO", .5, .3])
            teachers.append(["3", "Outro Docente", "PROFESSOR", "ATIVO", .2, 0])
            workbook_path = allocation_dir / "resultado_alocacao.xlsx"
            workbook.save(workbook_path)
            workbook.close()
            _write_manifest(round_dir, {"allocation_workbook": workbook_path})
            job = {
                "id": "job-homonyms", "upload_id": "upload-1",
                "original_name": "base.xlsx", "module": 52,
                "status": "CONCLUIDA", "message": "ok",
                "validation_status": "APROVADO", "require_optimal": 1,
                "time_limit_seconds": None, "round_name": "rodada_001",
                "round_dir": str(round_dir), "exit_code": 0,
                "created_at": "2026-07-15T00:00:00+00:00",
                "updated_at": "2026-07-15T00:00:01+00:00",
            }

            result = build_insights(job)

            self.assertEqual(3, result["teacher_stats"]["used_teachers"])
            self.assertEqual(2, result["kpis"]["pareto_teacher_count_80"])
            homonyms = [
                item for item in result["teacher_distribution"]
                if item["teacher"] == "Docente Homônimo"
            ]
            self.assertEqual(2, len(homonyms))
            self.assertEqual([6, 3], [item["allocations"] for item in homonyms])
        finally:
            shutil.rmtree(round_dir, ignore_errors=True)


if __name__ == "__main__":
    unittest.main()
