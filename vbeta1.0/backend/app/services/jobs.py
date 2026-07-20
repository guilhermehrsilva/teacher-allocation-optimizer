from __future__ import annotations

import hashlib
import json
import os
import subprocess
import sys
import time
import uuid
from collections import Counter, defaultdict
from concurrent.futures import Future, ThreadPoolExecutor
from pathlib import Path
from threading import Event, Lock
from typing import Any

from openpyxl import load_workbook

from ..config import Settings
from ..database import Database, utc_now
from .coordination import LifecycleCoordinator


TERMINAL_STATES = {
    "CONCLUIDA",
    "INTERROMPIDA",
    "FALHA_ENTRADA_CONFIG",
    "FALHA_VALIDACAO",
    "VALIDACAO_REPROVADA",
    "SNAPSHOT_ALTERADO",
    "FALHA_SOLVER",
    "OTIMO_NAO_COMPROVADO",
    "FALHA_AUDITORIA",
    "AUDITORIA_REPROVADA",
    "ERRO_INTERNO",
}


def read_json(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def verified_manifest_artifact(
    round_dir: Path,
    manifest: dict[str, Any],
    artifact_key: str,
) -> Path:
    """Resolve an artifact and enforce the size/hash recorded by the pipeline."""
    round_dir = round_dir.resolve()
    artifact = manifest.get("artifacts", {}).get(artifact_key)
    if not isinstance(artifact, dict):
        raise FileNotFoundError(f"O artefato {artifact_key} não consta no manifesto.")
    relative = artifact.get("path")
    expected_size = artifact.get("size_bytes")
    expected_hash = artifact.get("sha256")
    if not relative or not isinstance(expected_size, int) or isinstance(expected_size, bool):
        raise FileNotFoundError(f"O contrato de integridade de {artifact_key} está incompleto.")
    if not isinstance(expected_hash, str) or len(expected_hash) != 64:
        raise FileNotFoundError(f"O contrato de integridade de {artifact_key} está incompleto.")
    path = (round_dir / str(relative)).resolve()
    if round_dir not in path.parents or not path.is_file():
        raise FileNotFoundError(f"O artefato {artifact_key} não está disponível.")
    if path.stat().st_size != expected_size or _sha256(path) != expected_hash.lower():
        raise FileNotFoundError(
            f"O artefato {artifact_key} diverge do manifesto auditado da rodada."
        )
    return path


def public_job(job: dict[str, Any]) -> dict[str, Any]:
    status = job["status"]
    message = job["message"]
    engine_state: str | None = None
    history: list[dict[str, Any]] = []
    round_dir = Path(job["round_dir"]) if job.get("round_dir") else None
    # Once the database has recorded a terminal state it is authoritative.
    # A stale RUNNING status.json must never resurrect an interrupted job.
    if round_dir and status not in TERMINAL_STATES:
        status_path = round_dir / "status.json"
        if status_path.exists():
            payload = read_json(status_path)
            engine_state = payload.get("state")
            history = payload.get("history", [])
            # O arquivo do motor pode registrar CONCLUIDA alguns milissegundos
            # antes de o processo terminar e o manager consolidar o banco. O
            # banco continua autoritativo para estados terminais, evitando que
            # a API libere cenários/reset enquanto o processo ainda está vivo.
            if engine_state and engine_state not in TERMINAL_STATES:
                status = engine_state
                message = payload.get("message", message)
    return {
        "id": job["id"],
        "upload_id": job["upload_id"],
        "filename": job["original_name"],
        "module": job["module"],
        "status": status,
        "message": message,
        "validation_status": job["validation_status"],
        "require_optimal": bool(job["require_optimal"]),
        "time_limit_seconds": job["time_limit_seconds"],
        "kind": job.get("kind", "PRIMARY"),
        "scenario_id": job.get("scenario_id"),
        "is_official": bool(job.get("is_official", 0)),
        "round": job.get("round_name"),
        "exit_code": job.get("exit_code"),
        "created_at": job["created_at"],
        "updated_at": job["updated_at"],
        "history": history,
        "engine_state": engine_state,
        "terminal": status in TERMINAL_STATES,
    }


class JobQueueFull(RuntimeError):
    pass


class JobManagerUnavailable(RuntimeError):
    pass


class JobManager:
    def __init__(
        self,
        settings: Settings,
        database: Database,
        coordinator: LifecycleCoordinator,
    ) -> None:
        self.settings = settings
        self.database = database
        self.coordinator = coordinator
        self.executor = ThreadPoolExecutor(max_workers=1, thread_name_prefix="solver-job")
        self._stopping = Event()
        self._state_lock = Lock()
        self._futures: dict[Future[Any], str] = {}
        self._processes: dict[str, subprocess.Popen[str]] = {}

    def shutdown(self) -> None:
        # Coordinate the stop flag and snapshots with create(), otherwise a
        # request could enqueue a future after the shutdown snapshot.
        with self.coordinator.exclusive():
            with self._state_lock:
                if self._stopping.is_set():
                    return
                self._stopping.set()
                futures = list(self._futures.items())
                processes = list(self._processes.values())

        for future, job_id in futures:
            if future.cancel():
                with self.coordinator.exclusive():
                    self.database.update_job(
                        job_id,
                        status="INTERROMPIDA",
                        message="Execução cancelada durante o encerramento seguro.",
                    )
        for process in processes:
            self._terminate_process(process)
        # No worker or child process is allowed to outlive the application.
        self.executor.shutdown(wait=True, cancel_futures=True)

    def _terminate_process(self, process: subprocess.Popen[str]) -> None:
        if process.poll() is not None:
            return
        try:
            process.terminate()
            process.wait(timeout=self.settings.shutdown_grace_seconds)
        except subprocess.TimeoutExpired:
            process.kill()
            process.wait()
        except OSError:
            # The child may have exited between poll() and terminate().
            pass

    def _forget_future(self, future: Future[Any]) -> None:
        with self._state_lock:
            self._futures.pop(future, None)

    def create(
        self,
        upload: dict[str, Any],
        require_optimal: bool,
        time_limit_seconds: float | None,
    ) -> dict[str, Any]:
        with self.coordinator.exclusive():
            if self._stopping.is_set():
                raise JobManagerUnavailable("O servidor está sendo encerrado.")
            if (
                self.database.count_inflight_jobs("PRIMARY")
                >= self.settings.max_primary_inflight_jobs
            ):
                raise JobQueueFull(
                    "A fila do motor principal atingiu o limite. "
                    "Aguarde uma execução terminar antes de reenviar."
                )
            job_id = uuid.uuid4().hex
            now = utc_now()
            self.database.create_job(
                {
                    "id": job_id,
                    "upload_id": upload["id"],
                    "status": "QUEUED",
                    "message": "Execução adicionada à fila.",
                    "module": upload["module"],
                    "require_optimal": require_optimal,
                    "time_limit_seconds": time_limit_seconds,
                    "created_at": now,
                    "updated_at": now,
                }
            )
            try:
                future = self.executor.submit(self._execute, job_id)
            except RuntimeError as exc:
                self.database.update_job(
                    job_id,
                    status="INTERROMPIDA",
                    message="O servidor foi encerrado antes de iniciar a execução.",
                )
                raise JobManagerUnavailable("O servidor está sendo encerrado.") from exc
            with self._state_lock:
                self._futures[future] = job_id
            future.add_done_callback(self._forget_future)
            return public_job(self.database.get_job(job_id) or {})

    def _job_result_dir(self, job_id: str) -> Path:
        """Isola cada execução para impedir associação cruzada de rodadas."""
        return self.settings.result_dir / job_id

    @staticmethod
    def _rounds(output_dir: Path) -> set[Path]:
        return {
            item.resolve()
            for item in output_dir.glob("rodada_*")
            if item.is_dir()
        }

    def _detect_round(self, output_dir: Path, previous: set[Path]) -> Path | None:
        created = sorted(
            self._rounds(output_dir) - previous,
            key=lambda item: item.name,
        )
        return created[0] if created else None

    def _execute(self, job_id: str) -> None:
        job = self.database.get_job(job_id)
        if not job:
            return
        if self._stopping.is_set():
            with self.coordinator.exclusive():
                self.database.update_job(
                    job_id,
                    status="INTERROMPIDA",
                    message="Execução cancelada durante o encerramento seguro.",
                )
            return
        process: subprocess.Popen[str] | None = None
        try:
            job_log_dir = self.settings.job_dir / job_id
            job_log_dir.mkdir(parents=True, exist_ok=True)
            stdout_path = job_log_dir / "stdout.log"
            stderr_path = job_log_dir / "stderr.log"
            output_dir = self._job_result_dir(job_id)
            output_dir.mkdir(parents=True, exist_ok=False)
            previous_rounds = self._rounds(output_dir)
            self.database.update_job(
                job_id,
                status="RUNNING",
                message="Preparando a execução da vbeta.",
            )
            command = [
                sys.executable,
                str(self.settings.vbeta_dir / "executar_pipeline.py"),
                "--base",
                job["stored_path"],
                "--resultado",
                str(output_dir),
                "--modulo",
                str(job["module"]),
            ]
            if job["time_limit_seconds"] is not None:
                command.extend(["--tempo-limite", str(job["time_limit_seconds"])])
            if not bool(job["require_optimal"]):
                command.append("--permitir-nao-otimo")

            creationflags = subprocess.CREATE_NO_WINDOW if os.name == "nt" else 0
            with stdout_path.open("w", encoding="utf-8") as stdout, stderr_path.open(
                "w", encoding="utf-8"
            ) as stderr:
                process = subprocess.Popen(
                    command,
                    cwd=self.settings.vbeta_dir,
                    stdout=stdout,
                    stderr=stderr,
                    text=True,
                    creationflags=creationflags,
                )
                with self._state_lock:
                    self._processes[job_id] = process
                    stop_now = self._stopping.is_set()
                if stop_now:
                    self._terminate_process(process)
                self.database.update_job(job_id, process_id=process.pid)
                round_dir: Path | None = None
                while process.poll() is None:
                    if round_dir is None:
                        round_dir = self._detect_round(output_dir, previous_rounds)
                        if round_dir:
                            self.database.update_job(
                                job_id,
                                round_name=round_dir.name,
                                round_dir=str(round_dir),
                                message="Rodada iniciada.",
                            )
                    time.sleep(0.4)
                exit_code = int(process.returncode or 0)
            round_dir = round_dir or self._detect_round(output_dir, previous_rounds)
            state = "ERRO_INTERNO"
            message = "A execução terminou sem produzir status legível."
            if round_dir:
                status_path = round_dir / "status.json"
                if status_path.exists():
                    status = read_json(status_path)
                    state = status.get("state", state)
                    message = status.get("message", message)
            if self._stopping.is_set():
                state = "INTERROMPIDA"
                message = "Execução interrompida durante o encerramento seguro."
            with self.coordinator.exclusive():
                self.database.update_job(
                    job_id,
                    status=state,
                    message=message,
                    round_name=round_dir.name if round_dir else None,
                    round_dir=str(round_dir) if round_dir else None,
                    exit_code=exit_code,
                )
        except Exception as exc:
            stopping = self._stopping.is_set()
            with self.coordinator.exclusive():
                self.database.update_job(
                    job_id,
                    status="INTERROMPIDA" if stopping else "ERRO_INTERNO",
                    message=(
                        "Execução interrompida durante o encerramento seguro."
                        if stopping
                        else f"Falha ao iniciar ou acompanhar o processo: {exc}"
                    ),
                    exit_code=50,
                )
        finally:
            if process is not None:
                with self._state_lock:
                    self._processes.pop(job_id, None)


def load_summary(job: dict[str, Any]) -> dict[str, Any]:
    if not job.get("round_dir"):
        raise FileNotFoundError("A rodada ainda não foi criada.")
    round_dir = Path(job["round_dir"]).resolve()
    manifest_path = round_dir / "manifesto.json"
    if not manifest_path.is_file():
        raise FileNotFoundError("Manifesto ainda não disponível.")
    path = verified_manifest_artifact(
        round_dir,
        read_json(manifest_path),
        "allocation_summary",
    )
    return read_json(path)


def _round_workbook(job: dict[str, Any], manifest: dict[str, Any]) -> Path:
    return verified_manifest_artifact(
        Path(job["round_dir"]),
        manifest,
        "allocation_workbook",
    )


def _sheet_records(path: Path, sheet_name: str) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        rows = workbook[sheet_name].iter_rows(values_only=True)
        headers = [str(value or "").strip() for value in next(rows)]
        return [dict(zip(headers, row, strict=False)) for row in rows]
    finally:
        workbook.close()


def _clean(value: Any) -> str:
    return str(value or "").strip()


def _number(value: Any) -> float:
    try:
        return float(value or 0)
    except (TypeError, ValueError):
        return 0.0


def _dashboard_teaching_capacity(value: Any, hours_per_transmission: int) -> float:
    """Return the CH_LETIVA usable by whole transmissions on the dashboard."""
    teaching_hours = max(0.0, _number(value))
    return (teaching_hours // hours_per_transmission) * hours_per_transmission


def _filter_values(value: Any) -> list[str]:
    """Normalize legacy scalar filters and HTTP multi-value filters."""
    raw_values = value if isinstance(value, (list, tuple, set)) else [value]
    values: list[str] = []
    seen: set[str] = set()
    for item in raw_values:
        cleaned = _clean(item)
        if cleaned and cleaned not in seen:
            seen.add(cleaned)
            values.append(cleaned)
    return values


def _filter_options(allocations: list[dict[str, Any]]) -> dict[str, list[Any]]:
    order_rank = {"1ª": 0, "2ª": 1, "ESTENDIDA": 2}
    day_rank = {
        "SEGUNDA": 0, "TERÇA": 1, "QUARTA": 2, "QUINTA": 3,
        "SEXTA": 4, "SÁBADO": 5, "DOMINGO": 6, "NSA": 7,
    }
    courses: dict[str, str] = {}
    for row in allocations:
        code = _clean(row.get("CURSO"))
        if code:
            courses[code] = _clean(row.get("NOME_CURSO")) or code
    return {
        "orders": sorted({_clean(row.get("ORDEM")) for row in allocations if _clean(row.get("ORDEM"))}, key=lambda value: order_rank.get(value, 99)),
        "courses": [{"value": code, "label": label} for code, label in sorted(courses.items(), key=lambda item: item[1])],
        "clusters": sorted({_clean(row.get("CLUSTER")) for row in allocations if _clean(row.get("CLUSTER"))}),
        "days": sorted({_clean(row.get("DIA_AULA")) for row in allocations if _clean(row.get("DIA_AULA"))}, key=lambda value: day_rank.get(value, 99)),
        "times": sorted({_clean(row.get("HORÁRIO")) for row in allocations if _clean(row.get("HORÁRIO"))}),
    }


def build_dashboard(
    job: dict[str, Any],
    summary: dict[str, Any],
    filters: dict[str, Any] | None = None,
) -> dict[str, Any]:
    round_dir = Path(job["round_dir"])
    manifest = read_json(round_dir / "manifesto.json")
    workbook_path = _round_workbook(job, manifest)
    all_allocations = _sheet_records(workbook_path, "ALOCACOES")
    teachers = _sheet_records(workbook_path, "DOCENTES")
    field_map = {
        "order": "ORDEM", "course": "CURSO", "cluster": "CLUSTER",
        "day": "DIA_AULA", "time": "HORÁRIO",
    }
    selected_filters = {
        key: values
        for key, value in (filters or {}).items()
        if key in field_map and (values := _filter_values(value))
    }
    allocations = [
        row for row in all_allocations
        if all(
            _clean(row.get(field_map[key])) in values
            for key, values in selected_filters.items()
        )
    ]
    allocated_rows = [row for row in allocations if _clean(row.get("STATUS")).upper() == "ALOCADA"]
    unassigned_rows = [row for row in allocations if _clean(row.get("STATUS")).upper() != "ALOCADA"]
    transmissions = len(allocations)
    allocated = len(allocated_rows)
    active_rows = [row for row in teachers if _clean(row.get("STATUS")).upper() == "ATIVO"]
    active_teachers = len(active_rows)
    used_badges = {_clean(row.get("CHAPA")) for row in allocated_rows if _clean(row.get("CHAPA"))}
    used_teachers = len(used_badges)
    hours_per_transmission = int(summary.get("hours_per_transmission", 2) or 2)

    discipline_by_order = Counter(_clean(row.get("ORDEM")) for row in allocations)
    first_demand = hours_per_transmission * (
        discipline_by_order.get("1ª", 0) + discipline_by_order.get("ESTENDIDA", 0)
    )
    second_demand = hours_per_transmission * (
        discipline_by_order.get("2ª", 0) + discipline_by_order.get("ESTENDIDA", 0)
    )
    active_capacity = sum(
        _dashboard_teaching_capacity(row.get("CH_LETIVA"), hours_per_transmission)
        for row in active_rows
    )

    day_data: dict[str, dict[str, Any]] = defaultdict(lambda: {"disciplines": 0, "teachers": set()})
    for row in allocations:
        day_data[_clean(row.get("DIA_AULA")) or "NÃO INFORMADO"]["disciplines"] += 1
    for row in allocated_rows:
        badge = _clean(row.get("CHAPA"))
        if badge:
            day_data[_clean(row.get("DIA_AULA")) or "NÃO INFORMADO"]["teachers"].add(badge)
    day_rank = {"SEGUNDA": 0, "TERÇA": 1, "QUARTA": 2, "QUINTA": 3, "SEXTA": 4, "SÁBADO": 5, "DOMINGO": 6, "NSA": 7}
    by_day = [
        {"day": day, "disciplines": values["disciplines"], "teachers": len(values["teachers"])}
        for day, values in sorted(day_data.items(), key=lambda item: day_rank.get(item[0], 99))
    ]

    cluster_hours = Counter(
        (_clean(row.get("CLUSTER")) or "NÃO INFORMADO") for row in allocations
    )
    cluster_items = [
        {"cluster": cluster, "hours": count * hours_per_transmission}
        for cluster, count in cluster_hours.most_common()
    ]

    stage_allocated_hours = {
        "first_stage": hours_per_transmission * sum(
            _clean(row.get("ORDEM")) in {"1ª", "ESTENDIDA"} for row in allocated_rows
        ),
        "second_stage": hours_per_transmission * sum(
            _clean(row.get("ORDEM")) in {"2ª", "ESTENDIDA"} for row in allocated_rows
        ),
    }
    validation = manifest.get("phases", {}).get("validation", {})
    audit = manifest.get("phases", {}).get("audit", {})
    return {
        "job": public_job(job),
        "kpis": {
            "coverage_pct": round(100 * allocated / transmissions, 2) if transmissions else 0,
            "allocated": allocated,
            "transmissions": transmissions,
            "unassigned": len(unassigned_rows),
            "teacher_use_pct": (
                round(100 * used_teachers / active_teachers, 2) if active_teachers else 0
            ),
            "active_teachers": active_teachers,
            "used_teachers": used_teachers,
            "zero_active_teachers": max(0, active_teachers - used_teachers),
            "disciplines_by_order": {
                "first": discipline_by_order.get("1ª", 0),
                "second": discipline_by_order.get("2ª", 0),
                "extended": discipline_by_order.get("ESTENDIDA", 0),
            },
            "first_stage_demand_hours": first_demand,
            "second_stage_demand_hours": second_demand,
            "first_stage_capacity_delta_hours": round(first_demand - active_capacity, 2),
            "second_stage_capacity_delta_hours": round(second_demand - active_capacity, 2),
            "active_teaching_capacity_hours": round(active_capacity, 2),
        },
        "unassigned_reasons": [
            {"reason": reason, "count": count}
            for reason, count in sorted(
                Counter(_clean(row.get("MOTIVO")) or "NÃO INFORMADO" for row in unassigned_rows).items(),
                key=lambda item: item[1],
                reverse=True,
            )
        ],
        "stage_hours": stage_allocated_hours,
        "charts": {"by_day": by_day, "demand_hours_by_cluster": cluster_items},
        "filters": {"selected": selected_filters, "options": _filter_options(all_allocations)},
        "guardrails": {
            "validation": validation.get("status"),
            "solver": summary.get("solver_status"),
            "audit": audit.get("status"),
        },
        "wall_time_seconds": summary.get("wall_time_seconds"),
        "metric_notes": {
            "demand": f"Cada oferta representa {hours_per_transmission}h semanais. ESTENDIDA participa das duas etapas.",
            "capacity_delta": "Demanda filtrada menos a CH letiva utilizável dos docentes ativos, arredondada para transmissões completas; valor positivo indica déficit. A compatibilidade de perfil não reduz esta capacidade bruta.",
        },
        "source_note": "Indicadores recalculados a partir da planilha de resultado e do manifesto auditado da rodada.",
    }


def artifact_path(job: dict[str, Any], artifact_key: str) -> Path:
    if not job.get("round_dir"):
        raise FileNotFoundError("A rodada ainda não possui artefatos.")
    round_dir = Path(job["round_dir"]).resolve()
    manifest_path = round_dir / "manifesto.json"
    if not manifest_path.exists():
        raise FileNotFoundError("Manifesto ainda não disponível.")
    if artifact_key == "manifest":
        path = manifest_path.resolve()
    elif artifact_key == "status":
        path = (round_dir / "status.json").resolve()
    else:
        manifest = read_json(manifest_path)
        return verified_manifest_artifact(round_dir, manifest, artifact_key)
    if round_dir not in path.parents or not path.is_file():
        raise FileNotFoundError("Caminho de artefato inválido.")
    return path
