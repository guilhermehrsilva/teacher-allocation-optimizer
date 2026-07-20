from __future__ import annotations

import hashlib
import json
import os
import shutil
import subprocess
import sys
import time
import unicodedata
import uuid
from concurrent.futures import Future, ThreadPoolExecutor
from datetime import date, datetime, time as datetime_time
from pathlib import Path
from threading import Event, Lock
from typing import Any

from openpyxl import load_workbook

from ..config import Settings
from ..database import Database, utc_now
from .coordination import LifecycleCoordinator
from .jobs import artifact_path, load_summary, public_job, read_json


SCENARIO_CHANGE_FIELDS: dict[str, dict[str, tuple[str, ...]]] = {
    "teacher": {
        "CAPACIDADE": ("STATUS", "CH_LETIVA"),
        "COMPATIBILIDADE": ("PERFIL_DISCIPLINA",),
    },
    "offer": {
        "AGENDA": ("DIA_AULA", "HORÁRIO", "ORDEM"),
        "COMPATIBILIDADE": ("PERFIL_DISCIPLINA",),
    },
}

ENTITY_SHEETS = {"teacher": "DOCENTES", "offer": "MAPA PEDAGÓGICO"}
SCENARIO_POLICY_TYPES = {"ALOCAR_CLUSTER", "PRIORIDADE", "FIXAR"}


class ScenarioError(ValueError):
    pass


class ScenarioQueueFull(ScenarioError):
    pass


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _json_default(value: Any) -> str:
    if isinstance(value, (datetime, date, datetime_time)):
        return value.isoformat()
    return str(value)


def _dump_value(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, default=_json_default)


def _load_value(value: str) -> Any:
    return json.loads(value)


def _headers(worksheet: Any) -> dict[str, int]:
    return {
        str(cell.value or "").strip(): cell.column
        for cell in next(worksheet.iter_rows(min_row=1, max_row=1))
        if str(cell.value or "").strip()
    }


def _normalize_key(value: Any) -> str:
    text = unicodedata.normalize("NFKD", str(value or ""))
    return " ".join(text.encode("ascii", "ignore").decode("ascii").lower().split())


def _normalize_change_value(field: str, value: Any) -> Any:
    if field == "CH_LETIVA":
        if isinstance(value, bool):
            raise ScenarioError("CH_LETIVA deve ser um inteiro não negativo.")
        try:
            number = int(value)
        except (TypeError, ValueError) as exc:
            raise ScenarioError("CH_LETIVA deve ser um inteiro não negativo.") from exc
        if number < 0 or str(value).strip() not in {str(number), f"{number}.0"}:
            raise ScenarioError("CH_LETIVA deve ser um inteiro não negativo.")
        return number
    if field == "STATUS":
        normalized = str(value or "").strip().upper()
        if normalized not in {"ATIVO", "DEMITIDO", "LICENÇA MATER.", "LICENÇA MATER. COMPL. 180 DIAS"}:
            raise ScenarioError("STATUS não suportado para a simulação.")
        return normalized
    if field == "ORDEM":
        normalized = str(value or "").strip().upper()
        if normalized not in {"1ª", "2ª", "ESTENDIDA"}:
            raise ScenarioError("ORDEM deve ser 1ª, 2ª ou ESTENDIDA.")
        return normalized
    if field == "DIA_AULA":
        normalized = str(value or "").strip().upper()
        if normalized not in {
            "SEGUNDA", "TERÇA", "QUARTA", "QUINTA", "SEXTA", "SÁBADO", "NSA",
        }:
            raise ScenarioError("DIA_AULA não suportado para a simulação.")
        return normalized
    if field == "HORÁRIO":
        normalized = str(value or "").strip()
        if normalized == "":
            return ""
        for pattern in ("%H:%M", "%H:%M:%S"):
            try:
                return datetime.strptime(normalized, pattern).strftime("%H:%M")
            except ValueError:
                pass
        raise ScenarioError("HORÁRIO deve usar o formato HH:MM.")
    normalized = str(value or "").strip()
    if not normalized:
        raise ScenarioError(f"{field} não pode ficar vazio no cenário.")
    return normalized


def _allowed_change(entity_type: str, change_type: str, field_name: str) -> bool:
    return field_name in SCENARIO_CHANGE_FIELDS.get(entity_type, {}).get(change_type, ())


def _baseline_source(database: Database, scenario: dict[str, Any]) -> Path:
    baseline = database.get_job(scenario["baseline_job_id"])
    if not baseline:
        raise ScenarioError("A rodada-base do cenário não foi encontrada.")
    try:
        return artifact_path(baseline, "source_copy")
    except FileNotFoundError as exc:
        raise ScenarioError(
            "A cópia imutável e auditada da rodada-base não está disponível."
        ) from exc


def _editable_scenario(database: Database, scenario_id: str) -> dict[str, Any]:
    scenario = database.get_scenario(scenario_id)
    if not scenario:
        raise ScenarioError("Cenário não encontrado.")
    if scenario["status"] in {"EXECUTANDO", "HOMOLOGADO"}:
        raise ScenarioError("Este cenário não pode mais ser editado.")
    return scenario


def read_source_value(
    source: Path,
    entity_type: str,
    row_number: int,
    field_name: str,
) -> Any:
    sheet_name = ENTITY_SHEETS.get(entity_type)
    if not sheet_name:
        raise ScenarioError("Tipo de entidade de cenário inválido.")
    workbook = load_workbook(source, read_only=True, data_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise ScenarioError(f"A aba {sheet_name} não existe na rodada-base.")
        worksheet = workbook[sheet_name]
        headers = _headers(worksheet)
        if field_name not in headers:
            raise ScenarioError(f"A coluna {field_name} não existe em {sheet_name}.")
        if row_number < 2 or row_number > worksheet.max_row:
            raise ScenarioError(f"A linha {row_number} não existe em {sheet_name}.")
        return worksheet.cell(row_number, headers[field_name]).value
    finally:
        workbook.close()


def add_scenario_change(
    database: Database,
    scenario_id: str,
    *,
    change_type: str,
    entity_type: str,
    row_number: int,
    field_name: str,
    new_value: Any,
) -> dict[str, Any]:
    scenario = _editable_scenario(database, scenario_id)
    change_type = change_type.strip().upper()
    entity_type = entity_type.strip().lower()
    field_name = field_name.strip().upper()
    if not _allowed_change(entity_type, change_type, field_name):
        raise ScenarioError("A alteração solicitada não pertence ao contrato de cenários.")
    normalized = _normalize_change_value(field_name, new_value)
    source = _baseline_source(database, scenario)
    if entity_type == "offer":
        offered_rows = {
            item["row_number"] for item in source_catalog(database, scenario_id)["offers"]
        }
        if row_number not in offered_rows:
            raise ScenarioError("A linha selecionada não pertence às ofertas tratadas pelo solver.")
    old_value = read_source_value(source, entity_type, row_number, field_name)
    payload = {
        "id": uuid.uuid4().hex,
        "scenario_id": scenario_id,
        "change_type": change_type,
        "entity_type": entity_type,
        "row_number": row_number,
        "field_name": field_name,
        "old_value_json": _dump_value(old_value),
        "new_value_json": _dump_value(normalized),
        "created_at": utc_now(),
    }
    persisted = database.upsert_scenario_change(payload)
    return public_change(persisted)


def public_change(change: dict[str, Any]) -> dict[str, Any]:
    return {
        "id": change["id"],
        "change_type": change["change_type"],
        "entity_type": change["entity_type"],
        "row_number": change["row_number"],
        "field_name": change["field_name"],
        "old_value": _load_value(change["old_value_json"]),
        "new_value": _load_value(change["new_value_json"]),
        "created_at": change["created_at"],
    }


def public_policy(policy: dict[str, Any]) -> dict[str, Any]:
    return {
        "id": policy["id"],
        "policy_type": policy["policy_type"],
        "target_type": policy["target_type"],
        "target_value": policy["target_value"],
        "configuration": json.loads(policy["configuration_json"]),
        "created_at": policy["created_at"],
    }


def delete_scenario_change(
    database: Database,
    scenario_id: str,
    change_id: str,
) -> bool:
    _editable_scenario(database, scenario_id)
    return database.delete_scenario_change(scenario_id, change_id)


def delete_scenario_policy(
    database: Database,
    scenario_id: str,
    policy_id: str,
) -> bool:
    _editable_scenario(database, scenario_id)
    return database.delete_scenario_policy(scenario_id, policy_id)


def add_scenario_policy(
    database: Database,
    scenario_id: str,
    *,
    policy_type: str,
    target_type: str,
    target_value: str,
    configuration: dict[str, Any] | None = None,
) -> dict[str, Any]:
    _editable_scenario(database, scenario_id)
    policy_type = policy_type.strip().upper()
    target_type = target_type.strip().upper()
    target_value = str(target_value or "").strip()
    configuration = dict(configuration or {})
    if policy_type not in SCENARIO_POLICY_TYPES:
        raise ScenarioError("Política de cenário não suportada.")
    allowed_targets = {
        "ALOCAR_CLUSTER": {"CLUSTER"},
        "PRIORIDADE": {"COURSE", "OFFER"},
        "FIXAR": {"OFFER"},
    }
    if target_type not in allowed_targets[policy_type] or not target_value:
        raise ScenarioError("Alvo inválido para a política selecionada.")
    catalog = source_catalog(database, scenario_id)
    if policy_type == "ALOCAR_CLUSTER":
        if target_value not in {item["name"] for item in catalog["clusters"]}:
            raise ScenarioError("O cluster selecionado não existe na baseline.")
        baseline_unassigned_rows = sorted(
            item["row_number"]
            for item in catalog["offers"]
            if item["cluster"] == target_value
            and item.get("baseline_status") != "ALOCADA"
        )
        if not baseline_unassigned_rows:
            raise ScenarioError("O cluster selecionado não possui lacunas na baseline.")
        # The secondary engine may relax profile only for these audited gaps.
        configuration["baseline_unassigned_rows"] = baseline_unassigned_rows
    if policy_type == "PRIORIDADE" and target_type == "COURSE":
        if target_value not in {item["code"] for item in catalog["courses"]}:
            raise ScenarioError("O curso selecionado não existe na baseline.")
    if target_type == "OFFER":
        try:
            row_number = int(target_value)
        except ValueError as exc:
            raise ScenarioError("A oferta da política é inválida.") from exc
        offer = next((item for item in catalog["offers"] if item["row_number"] == row_number), None)
        if not offer:
            raise ScenarioError("A oferta selecionada não existe na baseline.")
        if policy_type == "FIXAR":
            badge = str(configuration.get("teacher_badge") or "").strip()
            if not badge or badge != offer.get("baseline_teacher_badge"):
                raise ScenarioError("Só é possível fixar o docente alocado na baseline.")
    payload = {
        "id": uuid.uuid4().hex,
        "scenario_id": scenario_id,
        "policy_type": policy_type,
        "target_type": target_type,
        "target_value": target_value,
        "configuration_json": json.dumps(configuration, ensure_ascii=False),
        "created_at": utc_now(),
    }
    persisted = database.upsert_scenario_policy(payload)
    return public_policy(persisted)


def source_catalog(database: Database, scenario_id: str) -> dict[str, Any]:
    scenario = database.get_scenario(scenario_id)
    if not scenario:
        raise ScenarioError("Cenário não encontrado.")
    source = _baseline_source(database, scenario)
    baseline = database.get_job(scenario["baseline_job_id"])
    baseline_decisions = {}
    if baseline and baseline.get("round_dir"):
        try:
            baseline_decisions = _decision_map(load_summary(baseline))
        except (FileNotFoundError, json.JSONDecodeError, KeyError, ValueError) as exc:
            raise ScenarioError(
                "O resumo auditado da rodada-base não está disponível."
            ) from exc
    workbook = load_workbook(source, read_only=True, data_only=True)
    try:
        teachers_ws = workbook["DOCENTES"]
        teacher_headers = _headers(teachers_ws)
        teachers = []
        for row_number, row in enumerate(
            teachers_ws.iter_rows(min_row=2, values_only=True), start=2
        ):
            if not any(value not in (None, "") for value in row):
                continue
            def teacher_value(field: str) -> Any:
                column = teacher_headers.get(field)
                return row[column - 1] if column else None
            teachers.append({
                "row_number": row_number,
                "badge": str(teacher_value("CHAPA") or ""),
                "name": str(teacher_value("NOME") or ""),
                "status": str(teacher_value("STATUS") or ""),
                "job_function": str(teacher_value("NM_FUNCAO") or ""),
                "teaching_capacity": teacher_value("CH_LETIVA") or 0,
                "profile": str(teacher_value("PERFIL_DISCIPLINA") or ""),
            })

        offers_ws = workbook["MAPA PEDAGÓGICO"]
        offer_headers = _headers(offers_ws)
        offers = []
        for row_number, row in enumerate(
            offers_ws.iter_rows(min_row=2, values_only=True), start=2
        ):
            if not any(value not in (None, "") for value in row):
                continue
            def offer_value(field: str) -> Any:
                column = offer_headers.get(field)
                return row[column - 1] if column else None
            if baseline_decisions:
                if row_number not in baseline_decisions:
                    continue
            else:
                synergy = _normalize_key(offer_value("SINERGIA"))
                lesson_format = _normalize_key(offer_value("FORMATO_AULA"))
                if synergy not in {"curso unico", "curso responsavel"} or lesson_format != "ao vivo":
                    continue
            baseline_assignment = baseline_decisions.get(row_number, {})
            course_code = str(offer_value("CURSO") or "")
            course_name = str(offer_value("NOME_CURSO") or course_code)
            offers.append({
                "row_number": row_number,
                "course": course_name,
                "course_code": course_code,
                "course_name": course_name,
                "discipline_code": str(offer_value("COD_DISCIPLINA") or ""),
                "discipline_name": str(offer_value("NOME_DISCIPLINA") or ""),
                "day": str(offer_value("DIA_AULA") or ""),
                "time": str(offer_value("HORÁRIO") or ""),
                "order": str(offer_value("ORDEM") or ""),
                "cluster": str(offer_value("CLUSTER") or ""),
                "profile": str(offer_value("PERFIL_DISCIPLINA") or ""),
                "baseline_status": str(baseline_assignment.get("status") or ""),
                "baseline_teacher_badge": str(baseline_assignment.get("teacher_badge") or ""),
                "baseline_teacher_name": str(baseline_assignment.get("teacher_name") or ""),
            })
    finally:
        workbook.close()
    courses = sorted(
        {
            (str(item["course_code"]), str(item["course_name"]))
            for item in offers if item["course_code"]
        },
        key=lambda item: item[1],
    )
    cluster_totals: dict[str, dict[str, int]] = {}
    for item in offers:
        cluster = str(item.get("cluster") or "").strip()
        if not cluster:
            continue
        metrics = cluster_totals.setdefault(cluster, {"total_offers": 0, "unassigned_offers": 0})
        metrics["total_offers"] += 1
        if item.get("baseline_status") != "ALOCADA":
            metrics["unassigned_offers"] += 1
    return {
        "teachers": teachers,
        "offers": offers,
        "courses": [{"code": code, "name": name} for code, name in courses],
        "clusters": [
            {"name": name, **metrics}
            for name, metrics in sorted(
                cluster_totals.items(),
                key=lambda item: (-item[1]["unassigned_offers"], item[0]),
            )
        ],
    }


def materialize_scenario(
    settings: Settings,
    database: Database,
    scenario_id: str,
    job_id: str,
) -> tuple[Path, Path]:
    scenario = database.get_scenario(scenario_id)
    if not scenario:
        raise ScenarioError("Cenário não encontrado.")
    source = _baseline_source(database, scenario)
    changes = database.list_scenario_changes(scenario_id)
    policies = database.list_scenario_policies(scenario_id)
    if not changes and not policies:
        raise ScenarioError("Adicione ao menos uma alteração ou política antes de simular.")
    job_dir = settings.scenario_job_dir / job_id
    job_dir.mkdir(parents=True, exist_ok=False)
    destination = job_dir / "entrada_cenario.xlsx"
    try:
        shutil.copy2(source, destination)
        workbook = load_workbook(destination)
        try:
            for change in changes:
                sheet_name = ENTITY_SHEETS[change["entity_type"]]
                worksheet = workbook[sheet_name]
                headers = _headers(worksheet)
                field = change["field_name"]
                if field not in headers:
                    raise ScenarioError(f"A coluna {field} não existe em {sheet_name}.")
                current = worksheet.cell(change["row_number"], headers[field]).value
                expected = _load_value(change["old_value_json"])
                current_comparable = _json_default(current) if isinstance(
                    current, (datetime, date, datetime_time)
                ) else current
                if current_comparable != expected:
                    raise ScenarioError(
                        f"A origem mudou em {sheet_name}, linha {change['row_number']}, coluna {field}."
                    )
                worksheet.cell(
                    change["row_number"], headers[field], _load_value(change["new_value_json"])
                )
            workbook.save(destination)
        finally:
            workbook.close()

        snapshot_path = job_dir / "alteracoes.json"
        snapshot = {
            "schema_version": 1,
            "scenario_id": scenario_id,
            "baseline_job_id": scenario["baseline_job_id"],
            "scenario_engine": str(settings.scenario_engine_dir),
            "source_sha256": _sha256(source),
            "materialized_sha256": _sha256(destination),
            "changes": [public_change(change) for change in changes],
            "policies": [public_policy(policy) for policy in policies],
            "created_at": utc_now(),
        }
        temporary_snapshot = snapshot_path.with_suffix(".json.tmp")
        temporary_snapshot.write_text(
            json.dumps(snapshot, ensure_ascii=False, indent=2, default=_json_default),
            encoding="utf-8",
        )
        temporary_snapshot.replace(snapshot_path)
        return destination, snapshot_path
    except Exception:
        # Uma materialização que não virou job não pode deixar fonte ou
        # premissas órfãs contendo dados docentes.
        shutil.rmtree(job_dir)
        raise


def public_scenario(database: Database, scenario: dict[str, Any]) -> dict[str, Any]:
    changes = [
        public_change(item) for item in database.list_scenario_changes(scenario["id"])
    ]
    latest_job = database.latest_scenario_job(scenario["id"])
    policies = [
        public_policy(item) for item in database.list_scenario_policies(scenario["id"])
    ]
    return {
        "id": scenario["id"],
        "baseline_job_id": scenario["baseline_job_id"],
        "baseline_round": scenario.get("baseline_round"),
        "baseline_filename": scenario.get("baseline_filename"),
        "module": scenario["module"],
        "name": scenario["name"],
        "description": scenario["description"],
        "status": scenario["status"],
        "created_at": scenario["created_at"],
        "updated_at": scenario["updated_at"],
        "promoted_at": scenario.get("promoted_at"),
        "official_job_id": scenario.get("official_job_id"),
        "changes": changes,
        "policies": policies,
        "latest_job": public_job(latest_job) if latest_job else None,
    }


def _decision_map(summary: dict[str, Any]) -> dict[int, dict[str, Any]]:
    return {
        int(item["source_row"]): item
        for item in summary.get("decisions", [])
        if item.get("source_row") is not None
    }


def compare_scenario(
    database: Database,
    scenario_id: str,
) -> dict[str, Any]:
    scenario = database.get_scenario(scenario_id)
    if not scenario:
        raise ScenarioError("Cenário não encontrado.")
    if scenario["status"] not in {"CONCLUIDO", "HOMOLOGADO"}:
        raise ScenarioError(
            "As premissas atuais ainda não possuem uma simulação concluída."
        )
    scenario_job = database.latest_scenario_job(scenario_id)
    baseline_job = database.get_job(scenario["baseline_job_id"])
    if not scenario_job or scenario_job["status"] != "CONCLUIDA":
        raise ScenarioError("O cenário ainda não possui uma simulação concluída.")
    if not baseline_job or baseline_job["status"] != "CONCLUIDA":
        raise ScenarioError("A rodada-base não está concluída.")
    baseline_summary = load_summary(baseline_job)
    scenario_summary = load_summary(scenario_job)
    baseline_decisions = _decision_map(baseline_summary)
    scenario_decisions = _decision_map(scenario_summary)
    recovered: list[dict[str, Any]] = []
    lost: list[dict[str, Any]] = []
    reassigned: list[dict[str, Any]] = []
    unchanged = 0
    for row_number in sorted(set(baseline_decisions) | set(scenario_decisions)):
        before = baseline_decisions.get(row_number, {})
        after = scenario_decisions.get(row_number, {})
        before_allocated = before.get("status") == "ALOCADA"
        after_allocated = after.get("status") == "ALOCADA"
        record = {
            "source_row": row_number,
            "discipline_code": after.get("discipline_code") or before.get("discipline_code"),
            "discipline_name": after.get("discipline_name") or before.get("discipline_name"),
            "before_teacher": before.get("teacher_name"),
            "after_teacher": after.get("teacher_name"),
        }
        if not before_allocated and after_allocated:
            recovered.append(record)
        elif before_allocated and not after_allocated:
            lost.append(record)
        elif before_allocated and after_allocated and before.get("teacher_badge") != after.get("teacher_badge"):
            reassigned.append(record)
        elif before.get("status") == after.get("status") and before.get("teacher_badge") == after.get("teacher_badge"):
            unchanged += 1

    transmissions = max(
        int(baseline_summary.get("transmissions", 0)),
        int(scenario_summary.get("transmissions", 0)),
        1,
    )
    baseline_allocated = int(baseline_summary.get("allocated", 0))
    scenario_allocated = int(scenario_summary.get("allocated", 0))
    baseline_stage = baseline_summary.get("allocated_stage_hours", {})
    scenario_stage = scenario_summary.get("allocated_stage_hours", {})
    hours_per_transmission = int(scenario_summary.get(
        "hours_per_transmission",
        baseline_summary.get("hours_per_transmission", 2),
    ))
    def contract_hours(summary: dict[str, Any], internal: bool) -> int:
        count = sum(
            1 for item in summary.get("decisions", [])
            if item.get("status") == "ALOCADA"
            and str(item.get("contract_model") or "").upper().startswith("CLT") is internal
        )
        return count * hours_per_transmission
    baseline_internal = contract_hours(baseline_summary, True)
    scenario_internal = contract_hours(scenario_summary, True)
    baseline_external = contract_hours(baseline_summary, False)
    scenario_external = contract_hours(scenario_summary, False)
    manifest = read_json(Path(scenario_job["round_dir"]) / "manifesto.json")
    return {
        "scenario": public_scenario(database, scenario),
        "baseline_job": public_job(baseline_job),
        "scenario_job": public_job(scenario_job),
        "kpis": {
            "baseline_coverage_pct": round(100 * baseline_allocated / transmissions, 2),
            "scenario_coverage_pct": round(100 * scenario_allocated / transmissions, 2),
            "coverage_delta_pp": round(100 * (scenario_allocated - baseline_allocated) / transmissions, 2),
            "allocated_delta": scenario_allocated - baseline_allocated,
            "unassigned_hours_delta": hours_per_transmission * (
                int(scenario_summary.get("unassigned", 0))
                - int(baseline_summary.get("unassigned", 0))
            ),
            "used_teachers_delta": int(scenario_summary.get("used_teachers", 0))
            - int(baseline_summary.get("used_teachers", 0)),
            "internal_allocated_hours_delta": scenario_internal - baseline_internal,
            "external_allocated_hours_delta": scenario_external - baseline_external,
            "assignment_stability_pct": round(100 * unchanged / transmissions, 2),
            "first_stage_hours_delta": int(scenario_stage.get("first_stage", 0))
            - int(baseline_stage.get("first_stage", 0)),
            "second_stage_hours_delta": int(scenario_stage.get("second_stage", 0))
            - int(baseline_stage.get("second_stage", 0)),
        },
        "differences": {
            "recovered": recovered,
            "lost": lost,
            "reassigned": reassigned,
        },
        "guardrails": {
            "validation": manifest.get("phases", {}).get("validation", {}).get("status"),
            "solver": scenario_summary.get("solver_status"),
            "audit": manifest.get("phases", {}).get("audit", {}).get("status"),
            "eligible_for_promotion": (
                scenario_summary.get("solver_status") == "OPTIMAL"
                and manifest.get("phases", {}).get("audit", {}).get("status") == "APROVADO"
                and manifest.get("phases", {}).get("validation", {}).get("status")
                in {"APROVADO", "APROVADO_COM_RESSALVAS"}
            ),
        },
    }


class ScenarioManager:
    """Executa apenas o motor secundário; nunca chama o pipeline de `vbeta`."""

    def __init__(
        self,
        settings: Settings,
        database: Database,
        coordinator: LifecycleCoordinator,
    ) -> None:
        self.settings = settings
        self.database = database
        self.coordinator = coordinator
        self.executor = ThreadPoolExecutor(max_workers=1, thread_name_prefix="scenario-solver")
        self._stopping = Event()
        self._state_lock = Lock()
        self._futures: dict[Future[Any], tuple[str, str]] = {}
        self._processes: dict[str, subprocess.Popen[str]] = {}

    def exclusive(self):
        return self.coordinator.exclusive()

    def shutdown(self) -> None:
        with self.coordinator.exclusive():
            with self._state_lock:
                if self._stopping.is_set():
                    return
                self._stopping.set()
                futures = list(self._futures.items())
                processes = list(self._processes.values())
        for future, (scenario_id, job_id) in futures:
            if future.cancel():
                with self.coordinator.exclusive():
                    self.database.update_job(
                        job_id,
                        status="INTERROMPIDA",
                        message="Simulação cancelada durante o encerramento seguro.",
                    )
                    self.database.update_scenario(scenario_id, status="FALHA")
        for process in processes:
            self._terminate_process(process)
        self.executor.shutdown(wait=True, cancel_futures=True)

    def _terminate_process(self, process: subprocess.Popen[str]) -> None:
        if process.poll() is not None:
            return
        try:
            process.terminate()
            process.wait(timeout=self.settings.shutdown_grace_seconds)
        except subprocess.TimeoutExpired:
            process.kill()
            process.wait()
        except OSError:
            pass

    def _forget_future(self, future: Future[Any]) -> None:
        with self._state_lock:
            self._futures.pop(future, None)

    def create_run(self, scenario_id: str) -> dict[str, Any]:
        # The web package intentionally runs as a single process. This lock
        # makes the check-and-reserve sequence atomic within that contract.
        with self.coordinator.exclusive():
            return self._reserve_run(scenario_id)

    def _reserve_run(self, scenario_id: str) -> dict[str, Any]:
        if self._stopping.is_set():
            raise ScenarioError("O servidor está sendo encerrado.")
        scenario = self.database.get_scenario(scenario_id)
        if not scenario:
            raise ScenarioError("Cenário não encontrado.")
        if not self.settings.scenario_engine_dir.is_dir():
            raise ScenarioError("O motor secundário de cenários não está disponível.")
        if scenario["status"] == "HOMOLOGADO":
            raise ScenarioError("Um cenário homologado não pode ser reprocessado.")
        latest = self.database.latest_scenario_job(scenario_id)
        if latest and latest["status"] in {"QUEUED", "RUNNING"}:
            raise ScenarioError("Este cenário já está em execução.")
        if (
            self.database.count_inflight_jobs("SCENARIO")
            >= self.settings.max_scenario_inflight_jobs
        ):
            raise ScenarioQueueFull(
                "A fila de cenários atingiu o limite. "
                "Aguarde uma simulação terminar antes de reenviar."
            )
        job_id = uuid.uuid4().hex
        input_path, changes_path = materialize_scenario(
            self.settings, self.database, scenario_id, job_id
        )
        now = utc_now()
        self.database.create_job({
            "id": job_id,
            "upload_id": scenario["upload_id"],
            "status": "QUEUED",
            "message": "Simulação adicionada à fila do motor secundário.",
            "module": scenario["module"],
            "require_optimal": True,
            "time_limit_seconds": None,
            "kind": "SCENARIO",
            "scenario_id": scenario_id,
            "created_at": now,
            "updated_at": now,
        })
        self.database.create_scenario_run({
            "scenario_id": scenario_id,
            "job_id": job_id,
            "input_path": str(input_path),
            "changes_path": str(changes_path),
            "created_at": now,
        })
        self.database.update_scenario(scenario_id, status="EXECUTANDO")
        try:
            future = self.executor.submit(self._execute, scenario_id, job_id, input_path)
        except RuntimeError as exc:
            self.database.update_job(
                job_id,
                status="INTERROMPIDA",
                message="O servidor foi encerrado antes de iniciar a simulação.",
            )
            self.database.update_scenario(scenario_id, status="FALHA")
            raise ScenarioError("O servidor está sendo encerrado.") from exc
        with self._state_lock:
            self._futures[future] = (scenario_id, job_id)
        future.add_done_callback(self._forget_future)
        return public_job(self.database.get_job(job_id) or {})

    def _execute(self, scenario_id: str, job_id: str, input_path: Path) -> None:
        job = self.database.get_job(job_id)
        if not job:
            return
        if self._stopping.is_set():
            with self.coordinator.exclusive():
                self.database.update_job(
                    job_id,
                    status="INTERROMPIDA",
                    message="Simulação cancelada durante o encerramento seguro.",
                )
                self.database.update_scenario(scenario_id, status="FALHA")
            return
        process: subprocess.Popen[str] | None = None
        try:
            log_dir = self.settings.scenario_job_dir / job_id
            stdout_path = log_dir / "stdout.log"
            stderr_path = log_dir / "stderr.log"
            output_dir = self.settings.scenario_result_dir / scenario_id / job_id
            output_dir.mkdir(parents=True, exist_ok=True)
            command = [
                sys.executable,
                str(self.settings.scenario_engine_dir / "executar_cenario.py"),
                "--base", str(input_path),
                "--resultado", str(output_dir),
                "--modulo", str(job["module"]),
                "--politicas", str(log_dir / "alteracoes.json"),
            ]
            creationflags = subprocess.CREATE_NO_WINDOW if os.name == "nt" else 0
            self.database.update_job(
                job_id, status="RUNNING", message="Executando o motor secundário de cenários."
            )
            with stdout_path.open("w", encoding="utf-8") as stdout, stderr_path.open(
                "w", encoding="utf-8"
            ) as stderr:
                process = subprocess.Popen(
                    command,
                    cwd=self.settings.scenario_engine_dir,
                    stdout=stdout,
                    stderr=stderr,
                    text=True,
                    creationflags=creationflags,
                )
                with self._state_lock:
                    self._processes[job_id] = process
                    stop_now = self._stopping.is_set()
                if stop_now:
                    self._terminate_process(process)
                self.database.update_job(job_id, process_id=process.pid)
                round_dir: Path | None = None
                while process.poll() is None:
                    if round_dir is None:
                        rounds = sorted(output_dir.glob("rodada_*"))
                        if rounds:
                            round_dir = rounds[0]
                            self.database.update_job(
                                job_id,
                                round_name=round_dir.name,
                                round_dir=str(round_dir),
                                message="Simulação iniciada no motor secundário.",
                            )
                    time.sleep(0.4)
                exit_code = int(process.returncode or 0)
            rounds = sorted(output_dir.glob("rodada_*"))
            round_dir = round_dir or (rounds[0] if rounds else None)
            state = "ERRO_INTERNO"
            message = "A simulação terminou sem produzir status legível."
            if round_dir and (round_dir / "status.json").is_file():
                status = read_json(round_dir / "status.json")
                state = status.get("state", state)
                message = status.get("message", message)
            if self._stopping.is_set():
                state = "INTERROMPIDA"
                message = "Simulação interrompida durante o encerramento seguro."
            with self.coordinator.exclusive():
                self.database.update_job(
                    job_id,
                    status=state,
                    message=message,
                    round_name=round_dir.name if round_dir else None,
                    round_dir=str(round_dir) if round_dir else None,
                    exit_code=exit_code,
                )
                self.database.update_scenario(
                    scenario_id,
                    status="CONCLUIDO" if state == "CONCLUIDA" else "FALHA",
                )
        except Exception as exc:
            stopping = self._stopping.is_set()
            with self.coordinator.exclusive():
                self.database.update_job(
                    job_id,
                    status="INTERROMPIDA" if stopping else "ERRO_INTERNO",
                    message=(
                        "Simulação interrompida durante o encerramento seguro."
                        if stopping
                        else f"Falha no motor secundário: {exc}"
                    ),
                    exit_code=50,
                )
                self.database.update_scenario(scenario_id, status="FALHA")
        finally:
            if process is not None:
                with self._state_lock:
                    self._processes.pop(job_id, None)
