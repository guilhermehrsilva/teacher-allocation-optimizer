from __future__ import annotations

import json
import shutil
import sys
import uuid
from datetime import datetime, timezone
from pathlib import Path
from typing import BinaryIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

from ..config import Settings
from ..database import Database, utc_now


class UploadValidationError(ValueError):
    pass


class UploadQuotaError(UploadValidationError):
    pass


_DETAIL_LABELS = {
    "active_profile_token_count": "Perfis ativos disponíveis para comparação",
    "duplicates": "Cabeçalhos duplicados",
    "error": "Falha encontrada",
    "extra": "Colunas não previstas",
    "missing": "Colunas obrigatórias ausentes",
    "profile": "Perfil",
    "profiles_without_active_teacher": "Perfis sem docente ativo compatível",
    "rows": "Linhas que precisam de revisão",
}


def _human_detail_value(value: object) -> str:
    """Render validation metadata without exposing Python/JSON notation."""
    if value in (None, ""):
        return "não informado"
    if isinstance(value, bool):
        return "sim" if value else "não"
    if isinstance(value, dict):
        parts = [
            f"{_DETAIL_LABELS.get(str(key), str(key).replace('_', ' '))}: "
            f"{_human_detail_value(item)}"
            for key, item in value.items()
        ]
        return "; ".join(parts)
    if isinstance(value, (list, tuple, set)):
        return "; ".join(_human_detail_value(item) for item in value) or "não informado"
    return str(value)


def _excel_safe(value: object) -> object:
    """Prevent spreadsheet formula execution in every user-controlled cell."""
    if isinstance(value, str) and value.lstrip().startswith(("=", "+", "-", "@")):
        return f"'{value}"
    return value


def _excel_safe_row(values: tuple[object, ...]) -> tuple[object, ...]:
    return tuple(_excel_safe(value) for value in values)


def _format_actionable_details(issue: dict) -> str:
    details = issue.get("details") or {}
    if not details:
        return ""

    profiles = details.get("profiles_without_active_teacher")
    if profiles:
        lines = ["Perfis sem docente ativo compatível:"]
        for item in profiles:
            affected_rows = ", ".join(map(str, item.get("rows", []))) or "não informadas"
            lines.append(f"• {item.get('profile', 'Perfil não informado')} — revisar linhas {affected_rows}.")
        active_count = details.get("active_profile_token_count")
        if active_count is not None:
            lines.append(f"A base possui {active_count} perfil(is) ativo(s) cadastrado(s) para comparação.")
        lines.append(
            "Ação sugerida: revisar o PERFIL_DISCIPLINA das ofertas e dos docentes ativos "
            "ou disponibilizar um docente com perfil compatível."
        )
        return "\n".join(lines)

    if missing := details.get("missing"):
        return f"Ação sugerida: incluir as colunas obrigatórias ausentes: {', '.join(map(str, missing))}."
    if extra := details.get("extra"):
        return f"Ação sugerida: revisar ou remover as colunas não previstas: {', '.join(map(str, extra))}."
    if duplicates := details.get("duplicates"):
        return f"Ação sugerida: manter apenas um cabeçalho para cada coluna duplicada: {', '.join(map(str, duplicates))}."
    if error := details.get("error"):
        return f"Falha identificada ao abrir o arquivo: {error}. Ação sugerida: verificar se a planilha está íntegra e no formato .xlsx."

    readable_items = []
    for key, value in details.items():
        label = _DETAIL_LABELS.get(str(key), str(key).replace("_", " ").strip().capitalize())
        rendered = _human_detail_value(value)
        readable_items.append(f"{label}: {rendered}.")
    readable_items.append("Ação sugerida: revisar os itens indicados e registrar a correção na primeira coluna.")
    return "\n".join(readable_items)


def write_validation_workbook(
    report: dict,
    output_dir: Path,
    source_name: str,
) -> Path:
    output_path = output_dir / "pendencias_validacao.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "PENDENCIAS"
    worksheet.sheet_view.showGridLines = False
    headers = (
        "STATUS DA CORREÇÃO", "OBSERVAÇÃO", "SEVERIDADE", "BLOQUEANTE",
        "CÓDIGO", "PLANILHA", "COLUNA", "QUANTIDADE", "LINHAS", "ORIENTAÇÃO", "INSIGHT ACIONÁVEL",
    )
    worksheet.append(headers)
    for issue in report.get("issues", []):
        worksheet.append(_excel_safe_row((
            "PENDENTE",
            "",
            issue.get("severity", ""),
            "SIM" if issue.get("blocking") else "NÃO",
            issue.get("code", ""),
            issue.get("sheet", ""),
            issue.get("column", ""),
            issue.get("count", 0),
            "; ".join(map(str, issue.get("rows", []))),
            issue.get("message", ""),
            _format_actionable_details(issue),
        )))

    header_fill = PatternFill("solid", fgColor="005F86")
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    worksheet.row_dimensions[1].height = 32
    widths = (22, 28, 13, 12, 30, 24, 20, 12, 22, 58, 42)
    for index, width in enumerate(widths, start=1):
        worksheet.column_dimensions[chr(64 + index)].width = width
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    if worksheet.max_row >= 2:
        status_validation = DataValidation(
            type="list",
            formula1='"PENDENTE,RESOLVIDO,NÃO SE APLICA"',
            allow_blank=False,
        )
        worksheet.add_data_validation(status_validation)
        status_validation.add(f"A2:A{worksheet.max_row}")

    summary = workbook.create_sheet("RESUMO")
    summary.sheet_view.showGridLines = False
    metadata = report.get("metadata", {})
    validation_summary = report.get("summary", {})
    summary_rows = (
        ("RESULTADO DA VALIDAÇÃO", ""),
        ("Arquivo de origem", source_name),
        ("Status", report.get("status", "")),
        ("Módulo esperado", metadata.get("expected_module", "")),
        ("Ofertas alocáveis", metadata.get("allocating_rows", "")),
        ("Grupos de pendência", validation_summary.get("issue_groups", 0)),
        ("Grupos bloqueantes", validation_summary.get("blocking_issue_groups", 0)),
    )
    for row in summary_rows:
        summary.append(_excel_safe_row(row))
    summary["A1"].fill = header_fill
    summary["A1"].font = Font(color="FFFFFF", bold=True, size=14)
    summary.merge_cells("A1:B1")
    summary.column_dimensions["A"].width = 28
    summary.column_dimensions["B"].width = 55
    for row in summary.iter_rows(min_row=2):
        row[0].font = Font(bold=True, color="17324D")
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    workbook.save(output_path)
    return output_path


def safe_filename(filename: str | None) -> str:
    name = Path(filename or "base.xlsx").name
    if Path(name).suffix.lower() != ".xlsx":
        raise UploadValidationError("Envie um arquivo Excel no formato .xlsx.")
    return name


def _directory_size(path: Path) -> int:
    total = 0
    if not path.is_dir():
        return total
    for item in path.rglob("*"):
        try:
            if item.is_file():
                total += item.stat().st_size
        except OSError:
            # A concurrently removed file no longer consumes the quota.
            continue
    return total


def _record_upload_dir(settings: Settings, upload: dict) -> Path | None:
    root = settings.upload_dir.resolve()
    candidate = Path(upload["stored_path"]).resolve().parent
    if candidate.parent != root or candidate.name != upload["id"]:
        return None
    return candidate


def cleanup_upload_storage(
    settings: Settings,
    database: Database,
    *,
    now: datetime | None = None,
) -> dict[str, int]:
    """Remove abandoned filesystem entries and expired uploads without jobs.

    The caller serializes this operation with upload creation and resets through
    the process-wide lifecycle coordinator.
    """
    now = now or datetime.now(timezone.utc)
    records = database.list_uploads_with_job_count()
    referenced_ids = {str(upload["id"]) for upload in records}
    removed_directories = 0
    removed_records = 0

    for child in settings.upload_dir.iterdir():
        if child.is_dir() and child.name not in referenced_ids:
            shutil.rmtree(child, ignore_errors=True)
            removed_directories += int(not child.exists())

    retention = max(0, settings.orphan_upload_retention_seconds)
    for upload in records:
        if int(upload.get("job_count") or 0) > 0:
            continue
        try:
            created_at = datetime.fromisoformat(str(upload["created_at"]))
            if created_at.tzinfo is None:
                created_at = created_at.replace(tzinfo=timezone.utc)
            age_seconds = (now - created_at.astimezone(timezone.utc)).total_seconds()
        except (TypeError, ValueError):
            age_seconds = float("inf")
        if age_seconds < retention:
            continue
        upload_dir = _record_upload_dir(settings, upload)
        if not database.delete_upload_if_unused(str(upload["id"])):
            continue
        removed_records += 1
        if upload_dir and upload_dir.exists():
            shutil.rmtree(upload_dir, ignore_errors=True)
            removed_directories += int(not upload_dir.exists())

    return {
        "removed_records": removed_records,
        "removed_directories": removed_directories,
        "usage_bytes": _directory_size(settings.upload_dir),
    }


def save_upload_stream(
    stream: BinaryIO,
    filename: str | None,
    module: int,
    settings: Settings,
    database: Database,
) -> dict:
    cleanup = cleanup_upload_storage(settings, database)
    current_usage = cleanup["usage_bytes"]
    original_name = safe_filename(filename)
    upload_id = uuid.uuid4().hex
    upload_dir = settings.upload_dir / upload_id
    upload_dir.mkdir(parents=True, exist_ok=False)
    stored_path = upload_dir / "fonte.xlsx"

    total = 0
    try:
        with stored_path.open("wb") as target:
            while chunk := stream.read(1024 * 1024):
                total += len(chunk)
                if total > settings.max_upload_bytes:
                    raise UploadValidationError("O arquivo excede o limite de 50 MB.")
                if current_usage + total > settings.max_upload_storage_bytes:
                    raise UploadQuotaError(
                        "A cota de armazenamento de uploads foi atingida. "
                        "Aguarde a limpeza de bases antigas antes de tentar novamente."
                    )
                target.write(chunk)
    except Exception:
        shutil.rmtree(upload_dir, ignore_errors=True)
        raise

    if total == 0:
        shutil.rmtree(upload_dir, ignore_errors=True)
        raise UploadValidationError("O arquivo enviado está vazio.")

    src_path = str(settings.vbeta_dir / "src")
    if src_path not in sys.path:
        sys.path.insert(0, src_path)
    from alocacao_docente.validation import validate_workbook

    try:
        validation = validate_workbook(stored_path, expected_module=module)
        validation_path, _validation_csv_path = validation.write(upload_dir / "validacao")
        report = json.loads(validation_path.read_text(encoding="utf-8"))
        validation_workbook_path = write_validation_workbook(
            report,
            upload_dir / "validacao",
            original_name,
        )
    except Exception as exc:
        shutil.rmtree(upload_dir, ignore_errors=True)
        raise UploadValidationError(f"Não foi possível validar a planilha: {exc}") from exc

    if current_usage + _directory_size(upload_dir) > settings.max_upload_storage_bytes:
        shutil.rmtree(upload_dir, ignore_errors=True)
        raise UploadQuotaError(
            "A cota de armazenamento de uploads foi atingida durante a validação. "
            "Aguarde a limpeza de bases antigas antes de tentar novamente."
        )

    created_at = utc_now()
    try:
        database.create_upload(
            {
                "id": upload_id,
                "original_name": original_name,
                "stored_path": str(stored_path),
                "module": module,
                "validation_status": validation.status,
                "validation_path": str(validation_path),
                # Mantém a coluna legada para uploads existentes.
                "validation_csv_path": str(validation_workbook_path),
                "created_at": created_at,
            }
        )
    except Exception:
        shutil.rmtree(upload_dir, ignore_errors=True)
        raise

    # A fonte reprovada nunca pode alimentar um job. Preserve somente os
    # relatórios acionáveis e remova o maior artefato imediatamente.
    if validation.status == "REPROVADO":
        try:
            stored_path.unlink(missing_ok=True)
        except OSError as exc:
            database.delete_upload_if_unused(upload_id)
            shutil.rmtree(upload_dir, ignore_errors=True)
            raise UploadValidationError(
                "A base foi reprovada e não pôde ser descartada com segurança."
            ) from exc
    report["source"] = original_name
    return {
        "id": upload_id,
        "filename": original_name,
        "size_bytes": total,
        "module": module,
        "created_at": created_at,
        "validation": report,
    }
