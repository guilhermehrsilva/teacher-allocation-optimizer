from __future__ import annotations

import unicodedata
from datetime import datetime, time
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .domain import (
    CONTRACT_UNKNOWN,
    Problem,
    Teacher,
    Transmission,
    module_stages_for_order,
)


MAPA_SHEET = "MAPA PEDAGÓGICO"
DOCENTES_SHEET = "DOCENTES"


def normalize(value: Any) -> str:
    return " ".join(str(value or "").strip().casefold().split())


def normalize_key(value: Any) -> str:
    text = unicodedata.normalize("NFKD", normalize(value))
    return "".join(char for char in text if not unicodedata.combining(char))


def profile_tokens(value: Any) -> frozenset[str]:
    return frozenset(normalize(item) for item in str(value or "").split(",") if item.strip())


def as_time(value: Any) -> time | None:
    if isinstance(value, datetime):
        return value.time()
    if isinstance(value, time):
        return value
    if isinstance(value, str):
        for pattern in ("%H:%M", "%H:%M:%S"):
            try:
                return datetime.strptime(value.strip(), pattern).time()
            except ValueError:
                pass
    return None


def _rows(worksheet: Any) -> tuple[dict[str, int], list[tuple[int, tuple[Any, ...]]]]:
    header = tuple(cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1)))
    names = [str(name) for name in header if name is not None]
    duplicates = sorted({name for name in names if names.count(name) > 1})
    if duplicates:
        raise ValueError(f"Cabeçalhos duplicados em {worksheet.title}: {', '.join(duplicates)}")
    indexes = {str(name): index for index, name in enumerate(header) if name is not None}
    rows = [
        (excel_row, tuple(values))
        for excel_row, values in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2)
        if any(value is not None and str(value).strip() for value in values)
    ]
    return indexes, rows


def _value(row: tuple[Any, ...], indexes: dict[str, int], column: str) -> Any:
    if column not in indexes:
        raise ValueError(f"Coluna obrigatória ausente: {column}")
    return row[indexes[column]]


def _nonnegative_integer(value: Any, column: str, excel_row: int) -> int:
    valid = (
        isinstance(value, int)
        and not isinstance(value, bool)
        or isinstance(value, float)
        and value.is_integer()
    )
    if not valid or value < 0:
        raise ValueError(
            f"{column} inválida na linha {excel_row} de DOCENTES: "
            "esperado inteiro não negativo"
        )
    return int(value)


def load_problem(path: str | Path) -> Problem:
    source = Path(path)
    workbook = load_workbook(source, data_only=True, read_only=True)
    try:
        for sheet in (MAPA_SHEET, DOCENTES_SHEET):
            if sheet not in workbook.sheetnames:
                raise ValueError(f"Aba obrigatória ausente: {sheet}")
        mapa_indexes, mapa_rows = _rows(workbook[MAPA_SHEET])
        docente_indexes, docente_rows = _rows(workbook[DOCENTES_SHEET])
    finally:
        workbook.close()

    teachers: list[Teacher] = []
    for excel_row, row in docente_rows:
        capacity = _nonnegative_integer(
            _value(row, docente_indexes, "CH_LETIVA"), "CH_LETIVA", excel_row
        )
        contracted = _nonnegative_integer(
            _value(row, docente_indexes, "CH_CONTRATADA"), "CH_CONTRATADA", excel_row
        )
        profile_text = str(_value(row, docente_indexes, "PERFIL_DISCIPLINA") or "").strip()
        teacher = Teacher(
                id=len(teachers),
                name=str(_value(row, docente_indexes, "NOME") or "").strip(),
                badge=str(_value(row, docente_indexes, "CHAPA") or "").strip(),
                job_function=str(_value(row, docente_indexes, "NM_FUNCAO") or "").strip(),
                contracted_capacity=contracted,
                teaching_capacity=capacity,
                manager=str(_value(row, docente_indexes, "GESTOR") or "").strip(),
                status=str(_value(row, docente_indexes, "STATUS") or "").strip(),
                profiles=profile_tokens(profile_text),
                profile_text=profile_text,
            )
        if teacher.contract_family == CONTRACT_UNKNOWN:
            raise ValueError(
                f"NM_FUNCAO não suportada na linha {excel_row} de DOCENTES: "
                f"{teacher.job_function}"
            )
        teachers.append(teacher)

    transmissions: list[Transmission] = []
    for excel_row, row in mapa_rows:
        synergy = str(_value(row, mapa_indexes, "SINERGIA") or "").strip()
        synergy_key = normalize_key(synergy)
        live_key = normalize_key(_value(row, mapa_indexes, "FORMATO_AULA"))
        transmits = synergy_key in {"curso unico", "curso responsavel"}
        if not transmits or live_key != "ao vivo":
            continue
        day = str(_value(row, mapa_indexes, "DIA_AULA") or "").strip().upper()
        profile_text = str(_value(row, mapa_indexes, "PERFIL_DISCIPLINA") or "").strip()
        order = str(_value(row, mapa_indexes, "ORDEM") or "").strip()
        try:
            module_stages_for_order(order)
        except ValueError as exc:
            raise ValueError(
                f"ORDEM inválida na linha {excel_row} de {MAPA_SHEET}: {order!r}"
            ) from exc
        transmission = Transmission(
                id=len(transmissions),
                excel_row=excel_row,
                course=str(_value(row, mapa_indexes, "CURSO") or "").strip(),
                course_name=str(_value(row, mapa_indexes, "NOME_CURSO") or "").strip(),
                curriculum=str(_value(row, mapa_indexes, "CURRÍCULO") or "").strip(),
                discipline_code=str(_value(row, mapa_indexes, "COD_DISCIPLINA") or "").strip(),
                discipline_name=str(_value(row, mapa_indexes, "NOME_DISCIPLINA") or "").strip(),
                profiles=profile_tokens(profile_text),
                profile_text=profile_text,
                synergy=synergy,
                day=day,
                start_time=as_time(_value(row, mapa_indexes, "HORÁRIO")),
                order=order,
                cluster=str(_value(row, mapa_indexes, "CLUSTER") or "").strip(),
                coordinator=str(_value(row, mapa_indexes, "COORDENADOR") or "").strip(),
            )
        transmissions.append(transmission)

    return Problem(source, tuple(teachers), tuple(transmissions))
