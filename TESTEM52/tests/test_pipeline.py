from __future__ import annotations

import hashlib
import json
import shutil
import sys
import unittest
import uuid
from pathlib import Path

from openpyxl import Workbook, load_workbook


ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from pipeline import (  # noqa: E402
    EXIT_INPUT_CONFIG,
    EXIT_SUCCESS,
    InputConfigurationError,
    PipelineConfig,
    discover_workbook,
    run_pipeline,
)
from alocacao_docente.validation import (  # noqa: E402
    DOCENTES_HEADERS,
    DOCENTES_SHEET,
    MAPA_HEADERS,
    MAPA_SHEET,
)


def write_valid_m52_workbook(path: Path) -> None:
    workbook = Workbook()
    mapa = workbook.active
    mapa.title = MAPA_SHEET
    mapa.append(MAPA_HEADERS)
    rows = []
    for identifier, order, methodology, proof, day, hour in (
        (1, "1ª", "5 SEMANAS", "5", "SEGUNDA", "19:00"),
        (2, "2ª", "5 SEMANAS", "10", "SEGUNDA", "19:00"),
        (3, "ESTENDIDA", "ESTENDIDA", "10", "TERÇA", "20:40"),
    ):
        rows.append(
            [
                f"CURSO_{identifier}",
                "CURSO TESTE",
                f"CURSO_{identifier}_2026",
                f"DISC_{identifier}",
                f"DISCIPLINA TESTE {identifier}",
                "Gestão - Administração",
                2026,
                52,
                "1º",
                order,
                "Validado",
                methodology,
                "ONLINE",
                proof,
                "N",
                "CLUSTER",
                "GESTÃO E NEGÓCIOS",
                "GESTOR TESTE",
                "CLT STRICTO",
                "Curso Único",
                day,
                hour,
                "AO VIVO",
            ]
        )
    for row in rows:
        mapa.append(row)
    docentes = workbook.create_sheet(DOCENTES_SHEET)
    docentes.append(DOCENTES_HEADERS)
    docentes.append(
        [
            "DOCENTE TESTE",
            "012345678901",
            "PROFESSOR DE ENSINO SUPERIOR EAD",
            40,
            4,
            "GESTOR TESTE",
            "ATIVO",
            "Gestão - Administração",
        ]
    )
    workbook.save(path)
    workbook.close()


class PipelineTests(unittest.TestCase):
    def make_workspace(self) -> Path:
        workspace = ROOT / "tests" / f".pipeline-{uuid.uuid4().hex}"
        workspace.mkdir()
        self.addCleanup(shutil.rmtree, workspace, True)
        return workspace

    def test_discovery_rejects_zero_workbooks(self):
        workspace = self.make_workspace()
        input_dir = workspace / "entrada"
        output_dir = workspace / "resultado"
        input_dir.mkdir()
        (input_dir / "~$bloqueio.xlsx").write_bytes(b"temporario")

        with self.assertRaisesRegex(InputConfigurationError, "encontrados 0"):
            discover_workbook(input_dir)
        run = run_pipeline(PipelineConfig(input_dir=input_dir, output_dir=output_dir))
        self.assertEqual(EXIT_INPUT_CONFIG, run.exit_code)
        self.assertIsNotNone(run.status_path)
        status = json.loads(run.status_path.read_text(encoding="utf-8"))
        self.assertEqual("FALHA_ENTRADA_CONFIG", status["state"])
        self.assertEqual(EXIT_INPUT_CONFIG, status["exit_code"])

    def test_discovery_rejects_multiple_workbooks(self):
        workspace = self.make_workspace()
        input_dir = workspace / "entrada"
        output_dir = workspace / "resultado"
        input_dir.mkdir()
        (input_dir / "a.xlsx").write_bytes(b"a")
        (input_dir / "b.XLSX").write_bytes(b"b")
        (input_dir / "~$bloqueio.xlsx").write_bytes(b"temporario")

        with self.assertRaisesRegex(InputConfigurationError, "encontrados 2"):
            discover_workbook(input_dir)
        run = run_pipeline(PipelineConfig(input_dir=input_dir, output_dir=output_dir))
        self.assertEqual(EXIT_INPUT_CONFIG, run.exit_code)
        self.assertIsNotNone(run.status_path)
        status = json.loads(run.status_path.read_text(encoding="utf-8"))
        self.assertEqual("FALHA_ENTRADA_CONFIG", status["state"])
        self.assertEqual(EXIT_INPUT_CONFIG, status["exit_code"])

    def test_valid_m52_pipeline_is_optimal_and_audited(self):
        workspace = self.make_workspace()
        input_dir = workspace / "entrada"
        output_dir = workspace / "resultado"
        input_dir.mkdir()
        source = input_dir / "base_m52.xlsx"
        write_valid_m52_workbook(source)
        (input_dir / "~$base_m52.xlsx").write_bytes(b"temporario")

        run = run_pipeline(PipelineConfig(input_dir=input_dir, output_dir=output_dir))

        self.assertEqual(EXIT_SUCCESS, run.exit_code, run.message)
        self.assertEqual("CONCLUIDA", run.state)
        self.assertIsNotNone(run.round_dir)
        round_dir = run.round_dir
        assert round_dir is not None
        status = json.loads((round_dir / "status.json").read_text(encoding="utf-8"))
        manifest = json.loads((round_dir / "manifesto.json").read_text(encoding="utf-8"))
        audit = json.loads(
            (round_dir / "auditoria" / "auditoria_alocacao.json").read_text(encoding="utf-8")
        )
        summary = json.loads(
            (round_dir / "alocacao" / "resumo_alocacao.json").read_text(encoding="utf-8")
        )

        self.assertEqual(0, status["exit_code"])
        self.assertEqual("CONCLUIDA", status["state"])
        self.assertEqual(52, manifest["config"]["expected_module"])
        self.assertEqual(42, manifest["config"]["random_seed"])
        self.assertEqual(8, manifest["config"]["workers"])
        self.assertEqual(0, manifest["config"]["grasp_iterations"])
        self.assertEqual(
            64,
            len(manifest["code_fingerprints"]["solver"]["sha256"]),
        )
        source_sha256 = hashlib.sha256(source.read_bytes()).hexdigest()
        self.assertEqual(source_sha256, manifest["source"]["sha256"])
        self.assertEqual(
            source_sha256,
            manifest["phases"]["validation"]["source_sha256"],
        )
        self.assertEqual("APROVADO", manifest["phases"]["validation"]["status"])
        self.assertEqual("OPTIMAL", manifest["phases"]["solver"]["solver_status"])
        self.assertEqual(
            "OPTIMAL",
            manifest["phases"]["solver"]["diagnostics"]["canonicalization_status"],
        )
        self.assertEqual(
            1,
            manifest["phases"]["solver"]["diagnostics"]["canonicalization_workers"],
        )
        self.assertEqual("APROVADO", manifest["phases"]["audit"]["status"])
        self.assertEqual("APROVADO", audit["status"])
        self.assertFalse(
            manifest["phases"]["solver"]["diagnostics"][
                "historical_contract_restricts_eligibility"
            ]
        )
        self.assertEqual("CLT STRICTO", summary["decisions"][0]["original_contract_model"])
        self.assertEqual("CLT EAD", summary["decisions"][0]["suggested_contract_model"])
        self.assertEqual(3, summary["allocated"])
        self.assertEqual(
            {"first_stage": 4, "second_stage": 4},
            summary["allocated_stage_hours"],
        )
        self.assertEqual(4, summary["average_weekly_allocated_hours"])
        self.assertTrue((round_dir / "fonte" / source.name).is_file())
        allocation_path = round_dir / "alocacao" / "resultado_alocacao.xlsx"
        self.assertTrue(allocation_path.is_file())
        self.assertTrue((round_dir / "alocacao" / "resumo_alocacao.json").is_file())
        workbook = load_workbook(allocation_path, read_only=True, data_only=True)
        try:
            worksheet = workbook["ALOCACOES"]
            headers = [cell.value for cell in worksheet[1]]
            values = next(worksheet.iter_rows(min_row=2, values_only=True))
            allocation = dict(zip(headers, values))
            self.assertEqual("CLT STRICTO", allocation["MODELO_CONTRATO_ORIGEM"])
            self.assertEqual("CLT EAD", allocation["MODELO_CONTRATO"])
            docentes = workbook["DOCENTES"]
            teacher_headers = [cell.value for cell in docentes[1]]
            teacher_values = next(docentes.iter_rows(min_row=2, values_only=True))
            teacher_row = dict(zip(teacher_headers, teacher_values))
            self.assertEqual(17, len(teacher_headers))
            self.assertNotIn("MODELO_CONTRATO", teacher_headers)
            self.assertEqual(3, teacher_row["QTD_TRANSMISSÕES"])
            self.assertEqual(2, teacher_row["QTD_DISCIPLINAS_1ª_ETAPA"])
            self.assertEqual(1, teacher_row["UTILIZAÇÃO_1ª_ETAPA"])
            self.assertEqual(2, teacher_row["QTD_DISCIPLINAS_2ª_ETAPA"])
            self.assertEqual(1, teacher_row["UTILIZAÇÃO_2ª_ETAPA"])
            self.assertEqual(1, teacher_row["UTILIZAÇÃO"])
        finally:
            workbook.close()


if __name__ == "__main__":
    unittest.main()
