from __future__ import annotations

import shutil
import json
import unittest
import uuid
from io import BytesIO
from pathlib import Path

from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook

from app.config import Settings
from app.main import create_app
from app.services.jobs import build_dashboard
from app.services.insights import build_insights
from app.services.scenarios import materialize_scenario
from app.services.uploads import write_validation_workbook


REPO_ROOT = Path(__file__).resolve().parents[3]


class BackendApiTests(unittest.TestCase):
    def setUp(self) -> None:
        data_dir = REPO_ROOT / "tela" / "data" / "test_runs" / uuid.uuid4().hex
        data_dir.mkdir(parents=True, exist_ok=False)
        self.data_dir = data_dir
        self.settings = Settings(
            repo_root=REPO_ROOT,
            tela_dir=REPO_ROOT / "tela",
            vbeta_dir=REPO_ROOT / "vbeta",
            data_dir=data_dir,
        )
        self.app = create_app(self.settings)
        self.client_context = TestClient(self.app)
        self.client = self.client_context.__enter__()

    def tearDown(self) -> None:
        self.client_context.__exit__(None, None, None)
        shutil.rmtree(self.data_dir, ignore_errors=True)

    def test_health_reports_vbeta_and_data_directory(self) -> None:
        response = self.client.get("/api/health")
        self.assertEqual(200, response.status_code)
        self.assertEqual("ok", response.json()["status"])
        self.assertTrue(response.json()["vbeta_available"])
        self.assertTrue(response.json()["scenario_engine_available"])
        self.assertTrue(response.json()["data_dir_ready"])

    def test_scenario_materialization_is_isolated_from_primary_source(self) -> None:
        source = self.data_dir / "baseline.xlsx"
        workbook = Workbook()
        teachers = workbook.active
        teachers.title = "DOCENTES"
        teachers.append(["CHAPA", "NOME", "STATUS", "CH_LETIVA", "PERFIL_DISCIPLINA"])
        teachers.append(["100", "Docente Teste", "ATIVO", 8, "GESTÃO"])
        offers = workbook.create_sheet("MAPA PEDAGÓGICO")
        offers.append([
            "CURSO", "NOME_CURSO", "COD_DISCIPLINA", "NOME_DISCIPLINA",
            "DIA_AULA", "HORÁRIO", "ORDEM", "PERFIL_DISCIPLINA", "CLUSTER",
            "SINERGIA", "FORMATO_AULA",
        ])
        offers.append(["ADM", "Administração", "ADM01", "Gestão", "SEGUNDA", "19:00", "1ª", "GESTÃO", "GESTÃO E NEGÓCIOS", "Curso único", "Ao vivo"])
        offers.append(["ADM", "Administração", "ADM02", "Linha informativa", "SEGUNDA", "20:40", "1ª", "GESTÃO", "GESTÃO E NEGÓCIOS", "Sinérgica", "Ao vivo"])
        workbook.save(source)
        workbook.close()

        now = "2026-07-15T00:00:00+00:00"
        database = self.app.state.database
        database.create_upload({
            "id": "upload-scenario", "original_name": source.name,
            "stored_path": str(source), "module": 52,
            "validation_status": "APROVADO",
            "validation_path": str(self.data_dir / "validation.json"),
            "validation_csv_path": str(self.data_dir / "validation.xlsx"),
            "created_at": now,
        })
        database.create_job({
            "id": "primary-baseline", "upload_id": "upload-scenario",
            "status": "CONCLUIDA", "message": "ok", "module": 52,
            "require_optimal": True, "time_limit_seconds": None,
            "created_at": now, "updated_at": now,
        })
        database.update_job(
            "primary-baseline", round_name="rodada_001", round_dir=str(self.data_dir / "rodada_001"),
        )

        created = self.client.post("/api/scenarios", json={
            "baseline_job_id": "primary-baseline",
            "name": "Capacidade futura",
            "description": "Simular aumento de capacidade sem tocar no solver principal.",
        })
        self.assertEqual(201, created.status_code)
        scenario_id = created.json()["id"]
        catalog = self.client.get(f"/api/scenarios/{scenario_id}/catalog")
        self.assertEqual(200, catalog.status_code)
        self.assertEqual([2], [item["row_number"] for item in catalog.json()["offers"]])
        change = self.client.post(f"/api/scenarios/{scenario_id}/changes", json={
            "change_type": "CAPACIDADE", "entity_type": "teacher",
            "row_number": 2, "field_name": "CH_LETIVA", "new_value": 12,
        })
        self.assertEqual(201, change.status_code)
        rejected_supporting_row = self.client.post(f"/api/scenarios/{scenario_id}/changes", json={
            "change_type": "AGENDA", "entity_type": "offer",
            "row_number": 3, "field_name": "DIA_AULA", "new_value": "TERÇA",
        })
        self.assertEqual(409, rejected_supporting_row.status_code)
        self.assertIn("ofertas tratadas pelo solver", rejected_supporting_row.json()["detail"])
        policy = self.client.post(f"/api/scenarios/{scenario_id}/policies", json={
            "policy_type": "ALOCAR_CLUSTER", "target_type": "CLUSTER",
            "target_value": "GESTÃO E NEGÓCIOS", "configuration": {},
        })
        self.assertEqual(201, policy.status_code)
        materialized, snapshot = materialize_scenario(
            self.settings, database, scenario_id, "scenario-job-test",
        )

        original_book = load_workbook(source, data_only=True)
        scenario_book = load_workbook(materialized, data_only=True)
        try:
            self.assertEqual(8, original_book["DOCENTES"]["D2"].value)
            self.assertEqual(12, scenario_book["DOCENTES"]["D2"].value)
        finally:
            original_book.close()
            scenario_book.close()
        self.assertTrue(snapshot.is_file())
        snapshot_payload = json.loads(snapshot.read_text(encoding="utf-8"))
        self.assertEqual("ALOCAR_CLUSTER", snapshot_payload["policies"][0]["policy_type"])
        self.assertEqual("PRIMARY", self.client.get("/api/jobs").json()[0]["kind"])

        reset_scenario = self.client.delete("/api/scenarios", params={"scope": "latest"})
        self.assertEqual(200, reset_scenario.status_code)
        self.assertEqual(1, reset_scenario.json()["deleted_scenarios"])
        self.assertEqual([], self.client.get("/api/scenarios").json())
        self.assertTrue(source.is_file())

        reset_baseline = self.client.delete("/api/jobs/primary", params={"scope": "latest"})
        self.assertEqual(200, reset_baseline.status_code)
        self.assertEqual(1, reset_baseline.json()["deleted_rounds"])
        self.assertEqual([], self.client.get("/api/jobs").json())

    def test_frontend_html_disables_browser_cache(self) -> None:
        response = self.client.get("/")
        self.assertEqual(200, response.status_code)
        self.assertEqual(
            "no-store, no-cache, must-revalidate, max-age=0",
            response.headers["cache-control"],
        )

    def test_upload_rejects_non_xlsx_file(self) -> None:
        response = self.client.post(
            "/api/uploads",
            data={"module": "52"},
            files={"file": ("base.csv", b"x,y", "text/csv")},
        )
        self.assertEqual(422, response.status_code)
        self.assertIn(".xlsx", response.json()["detail"])

    def test_upload_accepts_only_supported_modules(self) -> None:
        response = self.client.post(
            "/api/uploads",
            data={"module": "50"},
            files={"file": ("base.xlsx", b"not-needed", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        self.assertEqual(422, response.status_code)
        self.assertIn("greater than or equal to 51", str(response.json()["detail"]))

    def test_validation_issues_download_is_actionable_xlsx(self) -> None:
        report = {
            "status": "APROVADO_COM_RESSALVAS",
            "metadata": {"expected_module": 52, "allocating_rows": 10},
            "summary": {"issue_groups": 1, "blocking_issue_groups": 0},
            "issues": [{
                "severity": "ALTA",
                "blocking": False,
                "code": "PERFIL_SEM_DOCENTE_ATIVO",
                "sheet": "MAPA PEDAGÓGICO",
                "column": "PERFIL_DISCIPLINA",
                "count": 2,
                "rows": [4, 8],
                "message": "Duas ofertas não possuem docente ativo compatível.",
                "details": {
                    "profiles_without_active_teacher": [{
                        "profile": "Comunicação - Relações Públicas",
                        "rows": [4, 8],
                    }],
                    "active_profile_token_count": 111,
                },
            }],
        }
        workbook_path = write_validation_workbook(report, self.data_dir, "base.xlsx")
        self.app.state.database.create_upload({
            "id": "upload-xlsx", "original_name": "base.xlsx",
            "stored_path": str(self.data_dir / "base.xlsx"), "module": 52,
            "validation_status": "APROVADO_COM_RESSALVAS",
            "validation_path": str(self.data_dir / "validation.json"),
            "validation_csv_path": str(workbook_path),
            "created_at": "2026-07-14T00:00:00+00:00",
        })

        response = self.client.get("/api/uploads/upload-xlsx/validation.xlsx")

        self.assertEqual(200, response.status_code)
        self.assertIn("pendencias_validacao.xlsx", response.headers["content-disposition"])
        workbook = load_workbook(BytesIO(response.content))
        try:
            self.assertEqual(["PENDENCIAS", "RESUMO"], workbook.sheetnames)
            worksheet = workbook["PENDENCIAS"]
            self.assertEqual("STATUS DA CORREÇÃO", worksheet["A1"].value)
            self.assertEqual("PENDENTE", worksheet["A2"].value)
            self.assertEqual("MAPA PEDAGÓGICO", worksheet["F2"].value)
            self.assertEqual("4; 8", worksheet["I2"].value)
            self.assertEqual("INSIGHT ACIONÁVEL", worksheet["K1"].value)
            self.assertIn("Comunicação - Relações Públicas", worksheet["K2"].value)
            self.assertIn("Ação sugerida", worksheet["K2"].value)
            self.assertNotIn("profiles_without_active_teacher", worksheet["K2"].value)
            self.assertNotIn("{", worksheet["K2"].value)
            self.assertEqual(1, len(worksheet.data_validations.dataValidation))
        finally:
            workbook.close()

    def test_dashboard_route_accepts_repeated_filters(self) -> None:
        round_dir = self.data_dir / "fixture_round"
        allocation_dir = round_dir / "alocacao"
        allocation_dir.mkdir(parents=True)

        workbook = Workbook()
        allocations = workbook.active
        allocations.title = "ALOCACOES"
        allocations.append([
            "STATUS", "MOTIVO", "ORDEM", "CURSO", "NOME_CURSO",
            "CLUSTER", "DIA_AULA", "HORÁRIO", "CHAPA",
        ])
        allocations.append(["ALOCADA", None, "1ª", "ADM", "Administração", "GESTÃO", "SEGUNDA", "19:00", "1"])
        allocations.append(["ALOCADA", None, "2ª", "ADM", "Administração", "GESTÃO", "TERÇA", "19:00", "1"])
        allocations.append(["ALOCADA", None, "ESTENDIDA", "ENG", "Engenharia", "ENGENHARIAS", "SEGUNDA", "20:40", "2"])
        allocations.append(["NAO_ALOCADA", "SEM_CANDIDATO", "1ª", "ADM", "Administração", "GESTÃO", "QUARTA", "19:00", None])
        teachers = workbook.create_sheet("DOCENTES")
        teachers.append(["STATUS", "NM_FUNCAO", "CH_LETIVA"])
        teachers.append(["ATIVO", "PROFESSOR REGENTE", 20])
        workbook.save(allocation_dir / "resultado_alocacao.xlsx")

        (round_dir / "manifesto.json").write_text(
            '{"artifacts":{"allocation_workbook":{"path":"alocacao/resultado_alocacao.xlsx"}},'
            '"phases":{"validation":{"status":"APROVADO"},"audit":{"status":"APROVADO"}}}',
            encoding="utf-8",
        )
        (allocation_dir / "resumo_alocacao.json").write_text(
            '{"hours_per_transmission":2,"solver_status":"OPTIMAL","unassigned":1}',
            encoding="utf-8",
        )

        database = self.app.state.database
        now = "2026-07-14T00:00:00+00:00"
        database.create_upload({
            "id": "upload-multifilter", "original_name": "base.xlsx",
            "stored_path": str(self.data_dir / "base.xlsx"), "module": 52,
            "validation_status": "APROVADO",
            "validation_path": str(self.data_dir / "validation.json"),
            "validation_csv_path": str(self.data_dir / "validation.csv"),
            "created_at": now,
        })
        database.create_job({
            "id": "job-multifilter", "upload_id": "upload-multifilter",
            "status": "CONCLUIDA", "message": "ok", "module": 52,
            "require_optimal": True, "time_limit_seconds": None,
            "created_at": now, "updated_at": now,
        })
        database.update_job(
            "job-multifilter", round_name="rodada_001", round_dir=str(round_dir),
        )

        response = self.client.get(
            "/api/dashboard/job-multifilter",
            params=[("order", "1ª"), ("order", "2ª"), ("course", "ADM")],
        )

        self.assertEqual(200, response.status_code)
        payload = response.json()
        self.assertEqual(3, payload["kpis"]["transmissions"])
        self.assertEqual(
            {"order": ["1ª", "2ª"], "course": ["ADM"]},
            payload["filters"]["selected"],
        )


class DashboardContractTests(unittest.TestCase):
    def test_dashboard_reconciles_core_metrics(self) -> None:
        round_dir = REPO_ROOT / "tela" / "data" / "test_runs" / uuid.uuid4().hex
        allocation_dir = round_dir / "alocacao"
        allocation_dir.mkdir(parents=True, exist_ok=False)
        try:
            workbook = Workbook()
            allocations = workbook.active
            allocations.title = "ALOCACOES"
            allocations.append(["STATUS", "MOTIVO", "ORDEM", "CURSO", "NOME_CURSO", "CLUSTER", "DIA_AULA", "HORÁRIO", "CHAPA"])
            allocations.append(["ALOCADA", None, "1ª", "ADM", "Administração", "GESTÃO", "SEGUNDA", "19:00", "1"])
            allocations.append(["ALOCADA", None, "2ª", "ADM", "Administração", "GESTÃO", "TERÇA", "19:00", "1"])
            allocations.append(["ALOCADA", None, "ESTENDIDA", "ENG", "Engenharia", "ENGENHARIAS", "SEGUNDA", "20:40", "2"])
            allocations.append(["NAO_ALOCADA", "CHOQUE_DE_HORARIO", "1ª", "ADM", "Administração", "GESTÃO", "SEGUNDA", "19:00", None])
            for index in range(3, 8):
                allocations.append(["NAO_ALOCADA", "CHOQUE_DE_HORARIO", "1ª", "ADM", "Administração", f"CLUSTER {index}", "SEGUNDA", "19:00", None])
            teachers = workbook.create_sheet("DOCENTES")
            teachers.append(["STATUS", "NM_FUNCAO", "CH_LETIVA"])
            teachers.append(["ATIVO", "PROFESSOR REGENTE", 20])
            teachers.append(["ATIVO", "PROFESSOR DE ENSINO SUPERIOR PRESENCIAL", 9])
            workbook.save(allocation_dir / "resultado_alocacao.xlsx")
            (round_dir / "manifesto.json").write_text(
                """
                {
                  "artifacts": {"allocation_workbook": {"path": "alocacao/resultado_alocacao.xlsx"}},
                  "phases": {
                    "validation": {"status": "APROVADO"},
                    "audit": {"status": "APROVADO"}
                  }
                }
                """,
                encoding="utf-8",
            )
            job = {
                "id": "job-1",
                "upload_id": "upload-1",
                "original_name": "base.xlsx",
                "module": 52,
                "status": "CONCLUIDA",
                "message": "ok",
                "validation_status": "APROVADO",
                "require_optimal": 1,
                "time_limit_seconds": None,
                "round_name": "rodada_001",
                "round_dir": str(round_dir),
                "exit_code": 0,
                "created_at": "2026-07-14T00:00:00+00:00",
                "updated_at": "2026-07-14T00:00:01+00:00",
            }
            summary = {
                "hours_per_transmission": 2,
                "solver_status": "OPTIMAL",
                "wall_time_seconds": 12.5,
            }
            dashboard = build_dashboard(job, summary)
            self.assertEqual(33.33, dashboard["kpis"]["coverage_pct"])
            self.assertEqual(100.0, dashboard["kpis"]["teacher_use_pct"])
            self.assertEqual(6, dashboard["unassigned_reasons"][0]["count"])
            self.assertEqual(16, dashboard["kpis"]["first_stage_demand_hours"])
            self.assertEqual(28, dashboard["kpis"]["active_teaching_capacity_hours"])
            self.assertEqual(-12, dashboard["kpis"]["first_stage_capacity_delta_hours"])
            self.assertEqual(7, len(dashboard["charts"]["demand_hours_by_cluster"]))
            self.assertNotIn("OUTROS", {item["cluster"] for item in dashboard["charts"]["demand_hours_by_cluster"]})
            filtered = build_dashboard(job, summary, {"course": "ENG"})
            self.assertEqual(1, filtered["kpis"]["transmissions"])
            self.assertEqual(2, filtered["charts"]["demand_hours_by_cluster"][0]["hours"])
            self.assertEqual({"course": ["ENG"]}, filtered["filters"]["selected"])
            self.assertEqual("APROVADO", dashboard["guardrails"]["audit"])
        finally:
            shutil.rmtree(round_dir, ignore_errors=True)


class InsightsContractTests(unittest.TestCase):
    def test_insights_reconcile_pareto_risk_and_stage_exposure(self) -> None:
        round_dir = REPO_ROOT / "tela" / "data" / "test_runs" / uuid.uuid4().hex
        allocation_dir = round_dir / "alocacao"
        allocation_dir.mkdir(parents=True, exist_ok=False)
        try:
            workbook = Workbook()
            allocations = workbook.active
            allocations.title = "ALOCACOES"
            allocations.append([
                "STATUS", "MOTIVO", "ORDEM", "CLUSTER", "COORDENADOR",
                "MODELO_CONTRATO", "DOCENTE", "DIA_AULA", "CANDIDATOS_ELEGÍVEIS",
                "CURSO", "NOME_CURSO", "HORÁRIO", "PERFIL_DISCIPLINA",
            ])
            for teacher, order, day, candidates in [
                ("A", "1ª", "SEGUNDA", 1), ("A", "1ª", "SEGUNDA", 2),
                ("A", "2ª", "TERÇA", 2), ("A", "ESTENDIDA", "QUARTA", 2),
                ("B", "1ª", "SEGUNDA", 2), ("B", "2ª", "TERÇA", 2),
                ("C", "1ª", "SEGUNDA", 2), ("C", "2ª", "TERÇA", 2),
                ("D", "1ª", "QUARTA", 2), ("E", "2ª", "QUINTA", 2),
            ]:
                allocations.append(["ALOCADA", None, order, "CLUSTER 1", "GESTOR", "CLT EAD", teacher, day, candidates, "CURSO 1", "Curso 1", "19:00", "PERFIL 1"])
            allocations.append(["NAO_ALOCADA", "SEM_CANDIDATO", "1ª", "CLUSTER 2", "GESTOR", "CLT EAD", None, "SEXTA", 0, "CURSO 2", "Curso 2", "21:00", "PERFIL 2"])

            teachers = workbook.create_sheet("DOCENTES")
            teachers.append([
                "NOME", "NM_FUNCAO", "STATUS", "UTILIZAÇÃO_1ª_ETAPA", "UTILIZAÇÃO_2ª_ETAPA",
            ])
            for name, p1, p2 in [("A", 1, .5), ("B", .5, .5), ("C", .4, .4), ("D", .2, 0), ("E", 0, .2)]:
                teachers.append([name, "PROFESSOR", "ATIVO", p1, p2])
            workbook.save(allocation_dir / "resultado_alocacao.xlsx")
            (round_dir / "manifesto.json").write_text(
                '{"artifacts":{"allocation_workbook":{"path":"alocacao/resultado_alocacao.xlsx"}}}',
                encoding="utf-8",
            )
            job = {
                "id": "job-insights", "upload_id": "upload-1", "original_name": "base.xlsx",
                "module": 52, "status": "CONCLUIDA", "message": "ok",
                "validation_status": "APROVADO", "require_optimal": 1,
                "time_limit_seconds": None, "round_name": "rodada_001",
                "round_dir": str(round_dir), "exit_code": 0,
                "created_at": "2026-07-14T00:00:00+00:00",
                "updated_at": "2026-07-14T00:00:01+00:00",
            }

            result = build_insights(job)

            self.assertEqual(40.0, result["kpis"]["top_20_teacher_share_pct"])
            self.assertEqual(3, result["kpis"]["pareto_teacher_count_80"])
            self.assertEqual(1, result["kpis"]["single_candidate_allocations"])
            self.assertEqual(6, result["stage_load"]["first_stage_allocations"])
            self.assertEqual(5, result["stage_load"]["second_stage_allocations"])
            self.assertEqual("SEM_CANDIDATO", result["breakdowns"]["unassigned_reasons"][0]["label"])
            capacity_opportunity = next(
                item for item in result["opportunities"] if item["kind"] == "capacity"
            )
            self.assertEqual("5 docentes sem alocação", capacity_opportunity["metric"])
            self.assertNotIn("saldo médio", capacity_opportunity["metric"].lower())
            self.assertEqual(
                [{"day": "SEXTA", "count": 1}],
                result["coverage"]["courses"][0]["gap_days"],
            )
            congested_slot = result["diagnostics"]["most_congested_unassigned_slots"][0]
            self.assertEqual("SEXTA · 21:00", congested_slot["slot"])
            self.assertEqual(1, congested_slot["affected_courses"])
            self.assertEqual("Curso 2", congested_slot["top_course"])
            self.assertEqual("PERFIL 2", congested_slot["top_profile"])
        finally:
            shutil.rmtree(round_dir, ignore_errors=True)


if __name__ == "__main__":
    unittest.main()
