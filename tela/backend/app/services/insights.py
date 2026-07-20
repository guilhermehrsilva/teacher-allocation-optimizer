from __future__ import annotations

import math
import statistics
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook

from .jobs import public_job, read_json


_DAY_RANK = {
    "SEGUNDA": 1,
    "TERÇA": 2,
    "QUARTA": 3,
    "QUINTA": 4,
    "SEXTA": 5,
    "SÁBADO": 6,
    "DOMINGO": 7,
}


def _number(value: Any) -> float:
    if value in (None, ""):
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def _percent(value: Any) -> float:
    number = _number(value)
    return round(number * 100 if abs(number) <= 1 else number, 2)


def _text(value: Any, fallback: str = "Não informado") -> str:
    text = str(value or "").strip()
    return text or fallback


def _records(path: Path, sheet_name: str) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook[sheet_name]
        rows = worksheet.iter_rows(values_only=True)
        headers = [str(value or "").strip() for value in next(rows)]
        return [dict(zip(headers, row, strict=False)) for row in rows]
    finally:
        workbook.close()


def _allocation_workbook(job: dict[str, Any]) -> Path:
    round_dir = Path(job["round_dir"]).resolve()
    manifest = read_json(round_dir / "manifesto.json")
    relative = manifest.get("artifacts", {}).get("allocation_workbook", {}).get("path")
    if not relative:
        raise FileNotFoundError("A planilha de resultado não consta no manifesto da rodada.")
    path = (round_dir / relative).resolve()
    if round_dir not in path.parents or not path.is_file():
        raise FileNotFoundError("A planilha de resultado da rodada não está disponível.")
    return path


def _breakdown(rows: Iterable[dict[str, Any]], field: str, limit: int = 12) -> list[dict[str, Any]]:
    counts = Counter(_text(row.get(field)) for row in rows)
    total = sum(counts.values())
    return [
        {"label": label, "count": count, "share_pct": round(100 * count / total, 2) if total else 0}
        for label, count in counts.most_common(limit)
    ]


def _outlier_names(values: list[tuple[str, int]]) -> set[str]:
    if len(values) < 4:
        return set()
    ordered = sorted(value for _, value in values)
    lower = ordered[: len(ordered) // 2]
    upper = ordered[(len(ordered) + 1) // 2 :]
    q1 = statistics.median(lower)
    q3 = statistics.median(upper)
    iqr = q3 - q1
    low_limit = q1 - 1.5 * iqr
    high_limit = q3 + 1.5 * iqr
    mean = statistics.fmean(ordered)
    deviation = statistics.pstdev(ordered)
    return {
        name
        for name, value in values
        if value < low_limit
        or value > high_limit
        or (deviation > 0 and abs((value - mean) / deviation) >= 2)
    }


def _risk_class(score: int) -> str:
    if score <= 29:
        return "Baixo"
    if score <= 59:
        return "Médio"
    if score <= 79:
        return "Alto"
    return "Crítico"


def _coverage_rows(
    rows: list[dict[str, Any]],
    dimensions: tuple[str, ...],
) -> list[dict[str, Any]]:
    groups: dict[tuple[str, ...], list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        groups[tuple(_text(row.get(field)) for field in dimensions)].append(row)
    result: list[dict[str, Any]] = []
    for key, items in groups.items():
        allocated = sum(_text(item.get("STATUS"), "").upper() == "ALOCADA" for item in items)
        total = len(items)
        unassigned_items = [
            item for item in items if _text(item.get("STATUS"), "").upper() != "ALOCADA"
        ]
        gap_days = Counter(_text(item.get("DIA_AULA")) for item in unassigned_items)
        record = {field.lower(): value for field, value in zip(dimensions, key, strict=False)}
        record.update(
            {
                "total": total,
                "allocated": allocated,
                "unassigned": total - allocated,
                "coverage_pct": round(100 * allocated / total, 2) if total else 0,
                "demand_hours": total * 2,
                "single_candidate": sum(
                    _number(item.get("CANDIDATOS_ELEGÍVEIS")) <= 1
                    for item in items
                    if _text(item.get("STATUS"), "").upper() == "ALOCADA"
                ),
                "gap_days": [
                    {"day": day, "count": count}
                    for day, count in sorted(
                        gap_days.items(),
                        key=lambda value: (_DAY_RANK.get(value[0].upper(), 99), value[0]),
                    )
                ],
            }
        )
        result.append(record)
    return sorted(
        result,
        key=lambda item: (-item["unassigned"], item["coverage_pct"], -item["total"]),
    )


def build_insights(job: dict[str, Any]) -> dict[str, Any]:
    workbook_path = _allocation_workbook(job)
    allocations = _records(workbook_path, "ALOCACOES")
    teachers = _records(workbook_path, "DOCENTES")
    allocated = [row for row in allocations if _text(row.get("STATUS"), "").upper() == "ALOCADA"]
    unassigned = [row for row in allocations if _text(row.get("STATUS"), "").upper() != "ALOCADA"]

    allocation_counts = Counter(_text(row.get("DOCENTE")) for row in allocated)
    teacher_lookup = {_text(row.get("NOME")): row for row in teachers}
    teacher_days: dict[str, Counter[str]] = defaultdict(Counter)
    for row in allocated:
        teacher_days[_text(row.get("DOCENTE"))][_text(row.get("DIA_AULA"))] += 1

    ranked_counts = sorted(allocation_counts.items(), key=lambda item: (-item[1], item[0]))
    total_allocated = sum(allocation_counts.values())
    top_twenty_size = max(1, math.ceil(len(ranked_counts) * 0.2)) if ranked_counts else 0
    top_twenty_share = (
        100 * sum(value for _, value in ranked_counts[:top_twenty_size]) / total_allocated
        if total_allocated
        else 0
    )
    cumulative = 0
    pareto_count = 0
    teacher_distribution: list[dict[str, Any]] = []
    for name, count in ranked_counts:
        cumulative += count
        cumulative_pct = 100 * cumulative / total_allocated if total_allocated else 0
        if cumulative_pct <= 80 or pareto_count == 0:
            pareto_count += 1
        row = teacher_lookup.get(name, {})
        teacher_distribution.append(
            {
                "teacher": name,
                "role": _text(row.get("NM_FUNCAO")),
                "allocations": count,
                "share_pct": round(100 * count / total_allocated, 2) if total_allocated else 0,
                "cumulative_pct": round(cumulative_pct, 2),
                "stage_1_utilization_pct": _percent(row.get("UTILIZAÇÃO_1ª_ETAPA")),
                "stage_2_utilization_pct": _percent(row.get("UTILIZAÇÃO_2ª_ETAPA")),
            }
        )

    active_teachers = [row for row in teachers if _text(row.get("STATUS"), "").upper() == "ATIVO"]
    peak_utilizations = [
        max(_percent(row.get("UTILIZAÇÃO_1ª_ETAPA")), _percent(row.get("UTILIZAÇÃO_2ª_ETAPA")))
        for row in active_teachers
    ]
    p90_count = 0
    if ranked_counts:
        ordered_counts = sorted(count for _, count in ranked_counts)
        p90_count = ordered_counts[min(len(ordered_counts) - 1, math.ceil(0.9 * len(ordered_counts)) - 1)]
    outliers = _outlier_names(ranked_counts)

    risk_teachers: list[dict[str, Any]] = []
    for name, count in ranked_counts:
        row = teacher_lookup.get(name, {})
        stage_1 = _percent(row.get("UTILIZAÇÃO_1ª_ETAPA"))
        stage_2 = _percent(row.get("UTILIZAÇÃO_2ª_ETAPA"))
        peak = max(stage_1, stage_2)
        day_peak = max(teacher_days[name].values(), default=0)
        day_concentration = round(100 * day_peak / count, 2) if count else 0
        active = _text(row.get("STATUS"), "").upper() == "ATIVO"
        score = 0
        if peak >= 90:
            score += 25
        if peak >= 100:
            score += 20
        if day_concentration >= 60:
            score += 15
        if not active:
            score += 25
        if p90_count and count >= p90_count:
            score += 10
        score = min(score, 100)
        risk_teachers.append(
            {
                "teacher": name,
                "role": _text(row.get("NM_FUNCAO")),
                "allocations": count,
                "stage_1_utilization_pct": stage_1,
                "stage_2_utilization_pct": stage_2,
                "peak_utilization_pct": peak,
                "day_concentration_pct": day_concentration,
                "score": score,
                "risk_class": _risk_class(score),
                "load_outlier": name in outliers,
            }
        )
    risk_teachers.sort(key=lambda item: (-item["score"], -item["allocations"], item["teacher"]))

    scarce_allocations = sum(1 for row in allocated if _number(row.get("CANDIDATOS_ELEGÍVEIS")) <= 1)
    critical_or_high = sum(1 for row in risk_teachers if row["risk_class"] in {"Alto", "Crítico"})
    stage_1_count = sum(1 for row in allocated if _text(row.get("ORDEM"), "").upper() in {"1ª", "ESTENDIDA"})
    stage_2_count = sum(1 for row in allocated if _text(row.get("ORDEM"), "").upper() in {"2ª", "ESTENDIDA"})
    clusters = _breakdown(allocated, "CLUSTER")
    top_cluster = clusters[0] if clusters else {"label": "Não informado", "count": 0, "share_pct": 0}

    course_coverage = _coverage_rows(allocations, ("CURSO", "NOME_CURSO"))
    discipline_coverage = _coverage_rows(
        allocations,
        ("CURSO", "NOME_CURSO", "COD_DISCIPLINA", "NOME_DISCIPLINA"),
    )
    courses_below_90 = sum(item["coverage_pct"] < 90 for item in course_coverage)
    disciplines_uncovered = sum(item["coverage_pct"] == 0 for item in discipline_coverage)
    rpa_rows = [
        row for row in allocated
        if "RPA" in _text(row.get("MODELO_CONTRATO"), "").upper()
    ]
    idle_active_teachers = [
        row for row in teachers
        if _text(row.get("STATUS"), "").upper() == "ATIVO"
        and _number(row.get("QTD_TRANSMISSÕES")) == 0
    ]
    internal_idle_hours = sum(
        max(0, _number(row.get("CH_DISPONÍVEL")))
        for row in teachers
        if _text(row.get("STATUS"), "").upper() == "ATIVO"
    )
    affected_courses = len({_text(row.get("CURSO")) for row in unassigned})
    gap_reasons = _breakdown(unassigned, "MOTIVO")
    congested_slots: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in unassigned:
        slot = f"{_text(row.get('DIA_AULA'))} · {_text(row.get('HORÁRIO'))}"
        congested_slots[slot].append(row)
    top_gap_profile = Counter(
        _text(row.get("PERFIL_DISCIPLINA")) for row in unassigned
    ).most_common(1)
    opportunities = [
        {
            "priority": "Alta" if unassigned else "Monitorar",
            "kind": "coverage",
            "title": "Recuperar ofertas não alocadas",
            "metric": f"{len(unassigned) * 2}h de demanda",
            "impact": f"{len(unassigned)} ofertas em {affected_courses} cursos ainda estão sem cobertura.",
            "action": "Atacar primeiro os perfis e horários com maior recorrência de falha.",
        },
        {
            "priority": "Alta" if scarce_allocations else "Monitorar",
            "kind": "scarcity",
            "title": "Proteger alocações de candidato único",
            "metric": f"{scarce_allocations} ofertas sensíveis",
            "impact": "Qualquer afastamento ou reajuste pode transformar estas alocações em lacunas.",
            "action": "Criar segunda opção de perfil ou preservar agenda e CH do docente atual.",
        },
        {
            "priority": "Média",
            "kind": "capacity",
            "title": "Reavaliar docentes ativos sem alocação",
            "metric": f"{len(idle_active_teachers)} docentes sem alocação",
            "impact": "Estes docentes não receberam ofertas nesta rodada, mas podem ter restrições de compatibilidade.",
            "action": "Verificar perfil, agenda e saldo de CH antes de buscar capacidade externa.",
        },
        {
            "priority": "Alta" if rpa_rows else "Monitorar",
            "kind": "cost",
            "title": "Revisar exposição a RPA/NF",
            "metric": f"{len(rpa_rows) * 2}h alocadas",
            "impact": "Horas externas são um proxy de custo; o valor financeiro ainda não existe no contrato atual.",
            "action": "Simular internalização apenas quando houver docente compatível e saldo na etapa correta.",
        },
    ]

    automatic_insights = [
        {
            "tone": "attention" if top_twenty_share >= 50 else "neutral",
            "title": "Concentração da distribuição",
            "text": f"Os {top_twenty_size} docentes do grupo superior (20%) concentram {top_twenty_share:.1f}% das ofertas alocadas.",
        },
        {
            "tone": "critical" if critical_or_high else "positive",
            "title": "Risco operacional docente",
            "text": f"Foram identificados {critical_or_high} docentes em risco alto ou crítico pelo pico de utilização, concentração diária e volume relativo.",
        },
        {
            "tone": "attention" if scarce_allocations else "positive",
            "title": "Escassez de alternativas",
            "text": f"{scarce_allocations} alocações dependeram de um único candidato elegível e merecem proteção em reajustes futuros.",
        },
        {
            "tone": "neutral",
            "title": "Maior concentração temática",
            "text": f"{top_cluster['label']} reúne {top_cluster['count']} ofertas ({top_cluster['share_pct']:.1f}% das alocadas).",
        },
    ]

    return {
        "job": public_job(job),
        "kpis": {
            "coverage_pct": round(100 * len(allocated) / len(allocations), 2) if allocations else 0,
            "courses_below_90_pct": courses_below_90,
            "disciplines_uncovered": disciplines_uncovered,
            "unassigned_demand_hours": len(unassigned) * 2,
            "rpa_allocated_hours": len(rpa_rows) * 2,
            "internal_idle_hours": round(internal_idle_hours, 2),
            "top_20_teacher_share_pct": round(top_twenty_share, 2),
            "median_peak_utilization_pct": round(statistics.median(peak_utilizations), 2) if peak_utilizations else 0,
            "high_or_critical_risk_teachers": critical_or_high,
            "load_outliers": len(outliers),
            "single_candidate_allocations": scarce_allocations,
            "pareto_teacher_count_80": pareto_count,
        },
        "stage_load": {
            "first_stage_allocations": stage_1_count,
            "second_stage_allocations": stage_2_count,
            "difference": stage_2_count - stage_1_count,
        },
        "teacher_stats": {
            "used_teachers": len(ranked_counts),
            "mean_allocations": round(statistics.fmean(value for _, value in ranked_counts), 2) if ranked_counts else 0,
            "median_allocations": round(statistics.median(value for _, value in ranked_counts), 2) if ranked_counts else 0,
            "pareto_count_80": pareto_count,
            "top_20_count": top_twenty_size,
            "top_20_share_pct": round(top_twenty_share, 2),
        },
        "teacher_distribution": teacher_distribution[:20],
        "risk_teachers": risk_teachers[:30],
        "breakdowns": {
            "clusters": clusters,
            "days": _breakdown(allocated, "DIA_AULA", 7),
            "coordinators": _breakdown(allocated, "COORDENADOR"),
            "contracts": _breakdown(allocated, "MODELO_CONTRATO"),
            "unassigned_reasons": _breakdown(unassigned, "MOTIVO"),
        },
        "automatic_insights": automatic_insights,
        "coverage": {
            "courses": course_coverage,
            "discipline_gaps": discipline_coverage,
            "gap_reasons": gap_reasons,
        },
        "opportunities": opportunities,
        "diagnostics": {
            "most_congested_unassigned_slots": [
                {
                    "slot": slot,
                    "count": len(slot_rows),
                    "affected_courses": len({
                        (_text(row.get("CURSO")), _text(row.get("NOME_CURSO")))
                        for row in slot_rows
                    }),
                    "top_course": Counter(
                        _text(row.get("NOME_CURSO"), _text(row.get("CURSO")))
                        for row in slot_rows
                    ).most_common(1)[0][0],
                    "top_course_count": Counter(
                        _text(row.get("NOME_CURSO"), _text(row.get("CURSO")))
                        for row in slot_rows
                    ).most_common(1)[0][1],
                    "top_profile": Counter(
                        _text(row.get("PERFIL_DISCIPLINA")) for row in slot_rows
                    ).most_common(1)[0][0],
                }
                for slot, slot_rows in sorted(
                    congested_slots.items(), key=lambda item: (-len(item[1]), item[0])
                )
            ],
            "top_gap_profile": top_gap_profile[0][0] if top_gap_profile else None,
        },
        "methodology": [
            "Pareto e concentração calculados sobre a quantidade de ofertas alocadas por docente.",
            "Pico de utilização é o maior valor entre a 1ª e a 2ª etapa; capacidades das etapas não são somadas.",
            "Outliers combinam intervalo interquartil (1,5 × IQR) e z-score absoluto maior ou igual a 2.",
            "O score de risco adapta a regra do notebook estatístico: utilização, concentração diária, vínculo ativo e volume no percentil 90.",
        ],
        "limitations": [
            "Esta rodada representa a alocação planejada. Comparações planejado × realizado exigem a futura ingestão do Registro de execução.",
            "Indicadores financeiros de RPA dependem de campos de valor que ainda não fazem parte do contrato da aplicação.",
        ],
        "source_note": "Indicadores calculados da planilha de resultado publicada e do manifesto auditado da rodada.",
    }
