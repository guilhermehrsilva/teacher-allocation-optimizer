from __future__ import annotations

import json
import sys
import uuid
from pathlib import Path
from typing import BinaryIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

from ..config import Settings
from ..database import Database, utc_now


class UploadValidationError(ValueError):
    pass


def _format_actionable_details(issue: dict) -> str:
    details = issue.get("details") or {}
    if not details:
        return ""

    profiles = details.get("profiles_without_active_teacher")
    if profiles:
        lines = ["Perfis sem docente ativo compatível:"]
        for item in profiles:
            affected_rows = ", ".join(map(str, item.get("rows", []))) or "não informadas"
            lines.append(f"• {item.get('profile', 'Perfil não informado')} — revisar linhas {affected_rows}.")
        active_count = details.get("active_profile_token_count")
        if active_count is not None:
            lines.append(f"A base possui {active_count} perfil(is) ativo(s) cadastrado(s) para comparação.")
        lines.append(
            "Ação sugerida: revisar o PERFIL_DISCIPLINA das ofertas e dos docentes ativos "
            "ou disponibilizar um docente com perfil compatível."
        )
        return "\n".join(lines)

    if missing := details.get("missing"):
        return f"Ação sugerida: incluir as colunas obrigatórias ausentes: {', '.join(map(str, missing))}."
    if extra := details.get("extra"):
        return f"Ação sugerida: revisar ou remover as colunas não previstas: {', '.join(map(str, extra))}."
    if duplicates := details.get("duplicates"):
        return f"Ação sugerida: manter apenas um cabeçalho para cada coluna duplicada: {', '.join(map(str, duplicates))}."
    if error := details.get("error"):
        return f"Falha identificada ao abrir o arquivo: {error}. Ação sugerida: verificar se a planilha está íntegra e no formato .xlsx."

    readable_items = []
    for key, value in details.items():
        label = str(key).replace("_", " ").strip().capitalize()
        if isinstance(value, list):
            rendered = ", ".join(map(str, value))
        else:
            rendered = str(value)
        readable_items.append(f"{label}: {rendered}.")
    return "\n".join(readable_items)


def write_validation_workbook(
    report: dict,
    output_dir: Path,
    source_name: str,
) -> Path:
    output_path = output_dir / "pendencias_validacao.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "PENDENCIAS"
    worksheet.sheet_view.showGridLines = False
    headers = (
        "STATUS DA CORREÇÃO", "OBSERVAÇÃO", "SEVERIDADE", "BLOQUEANTE",
        "CÓDIGO", "PLANILHA", "COLUNA", "QUANTIDADE", "LINHAS", "ORIENTAÇÃO", "INSIGHT ACIONÁVEL",
    )
    worksheet.append(headers)
    for issue in report.get("issues", []):
        worksheet.append((
            "PENDENTE",
            "",
            issue.get("severity", ""),
            "SIM" if issue.get("blocking") else "NÃO",
            issue.get("code", ""),
            issue.get("sheet", ""),
            issue.get("column", ""),
            issue.get("count", 0),
            "; ".join(map(str, issue.get("rows", []))),
            issue.get("message", ""),
            _format_actionable_details(issue),
        ))

    header_fill = PatternFill("solid", fgColor="005F86")
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    worksheet.row_dimensions[1].height = 32
    widths = (22, 28, 13, 12, 30, 24, 20, 12, 22, 58, 42)
    for index, width in enumerate(widths, start=1):
        worksheet.column_dimensions[chr(64 + index)].width = width
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    if worksheet.max_row >= 2:
        status_validation = DataValidation(
            type="list",
            formula1='"PENDENTE,RESOLVIDO,NÃO SE APLICA"',
            allow_blank=False,
        )
        worksheet.add_data_validation(status_validation)
        status_validation.add(f"A2:A{worksheet.max_row}")

    summary = workbook.create_sheet("RESUMO")
    summary.sheet_view.showGridLines = False
    metadata = report.get("metadata", {})
    validation_summary = report.get("summary", {})
    summary_rows = (
        ("RESULTADO DA VALIDAÇÃO", ""),
        ("Arquivo de origem", source_name),
        ("Status", report.get("status", "")),
        ("Módulo esperado", metadata.get("expected_module", "")),
        ("Ofertas alocáveis", metadata.get("allocating_rows", "")),
        ("Grupos de pendência", validation_summary.get("issue_groups", 0)),
        ("Grupos bloqueantes", validation_summary.get("blocking_issue_groups", 0)),
    )
    for row in summary_rows:
        summary.append(row)
    summary["A1"].fill = header_fill
    summary["A1"].font = Font(color="FFFFFF", bold=True, size=14)
    summary.merge_cells("A1:B1")
    summary.column_dimensions["A"].width = 28
    summary.column_dimensions["B"].width = 55
    for row in summary.iter_rows(min_row=2):
        row[0].font = Font(bold=True, color="17324D")
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    workbook.save(output_path)
    return output_path


def safe_filename(filename: str | None) -> str:
    name = Path(filename or "base.xlsx").name
    if Path(name).suffix.lower() != ".xlsx":
        raise UploadValidationError("Envie um arquivo Excel no formato .xlsx.")
    return name


def save_upload_stream(
    stream: BinaryIO,
    filename: str | None,
    module: int,
    settings: Settings,
    database: Database,
) -> dict:
    original_name = safe_filename(filename)
    upload_id = uuid.uuid4().hex
    upload_dir = settings.upload_dir / upload_id
    upload_dir.mkdir(parents=True, exist_ok=False)
    stored_path = upload_dir / "fonte.xlsx"

    total = 0
    with stored_path.open("wb") as target:
        while chunk := stream.read(1024 * 1024):
            total += len(chunk)
            if total > settings.max_upload_bytes:
                target.close()
                stored_path.unlink(missing_ok=True)
                raise UploadValidationError("O arquivo excede o limite de 50 MB.")
            target.write(chunk)

    if total == 0:
        stored_path.unlink(missing_ok=True)
        raise UploadValidationError("O arquivo enviado está vazio.")

    src_path = str(settings.vbeta_dir / "src")
    if src_path not in sys.path:
        sys.path.insert(0, src_path)
    from alocacao_docente.validation import validate_workbook

    try:
        validation = validate_workbook(stored_path, expected_module=module)
        validation_path, _validation_csv_path = validation.write(upload_dir / "validacao")
        report = json.loads(validation_path.read_text(encoding="utf-8"))
        validation_workbook_path = write_validation_workbook(
            report,
            upload_dir / "validacao",
            original_name,
        )
    except Exception as exc:
        raise UploadValidationError(f"Não foi possível validar a planilha: {exc}") from exc

    created_at = utc_now()
    database.create_upload(
        {
            "id": upload_id,
            "original_name": original_name,
            "stored_path": str(stored_path),
            "module": module,
            "validation_status": validation.status,
            "validation_path": str(validation_path),
            # Mantém a coluna legada do banco para compatibilidade com uploads existentes.
            "validation_csv_path": str(validation_workbook_path),
            "created_at": created_at,
        }
    )
    report["source"] = original_name
    return {
        "id": upload_id,
        "filename": original_name,
        "size_bytes": total,
        "module": module,
        "created_at": created_at,
        "validation": report,
    }
