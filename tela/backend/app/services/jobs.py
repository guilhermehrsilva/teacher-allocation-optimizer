from __future__ import annotations

import json
import os
import subprocess
import sys
import time
import uuid
from collections import Counter, defaultdict
from concurrent.futures import ThreadPoolExecutor
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from ..config import Settings
from ..database import Database, utc_now


TERMINAL_STATES = {
    "CONCLUIDA",
    "INTERROMPIDA",
    "FALHA_ENTRADA_CONFIG",
    "FALHA_VALIDACAO",
    "VALIDACAO_REPROVADA",
    "SNAPSHOT_ALTERADO",
    "FALHA_SOLVER",
    "OTIMO_NAO_COMPROVADO",
    "FALHA_AUDITORIA",
    "AUDITORIA_REPROVADA",
    "ERRO_INTERNO",
}


def read_json(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def public_job(job: dict[str, Any]) -> dict[str, Any]:
    status = job["status"]
    message = job["message"]
    history: list[dict[str, Any]] = []
    round_dir = Path(job["round_dir"]) if job.get("round_dir") else None
    if round_dir:
        status_path = round_dir / "status.json"
        if status_path.exists():
            payload = read_json(status_path)
            status = payload.get("state", status)
            message = payload.get("message", message)
            history = payload.get("history", [])
    return {
        "id": job["id"],
        "upload_id": job["upload_id"],
        "filename": job["original_name"],
        "module": job["module"],
        "status": status,
        "message": message,
        "validation_status": job["validation_status"],
        "require_optimal": bool(job["require_optimal"]),
        "time_limit_seconds": job["time_limit_seconds"],
        "kind": job.get("kind", "PRIMARY"),
        "scenario_id": job.get("scenario_id"),
        "is_official": bool(job.get("is_official", 0)),
        "round": job.get("round_name"),
        "exit_code": job.get("exit_code"),
        "created_at": job["created_at"],
        "updated_at": job["updated_at"],
        "history": history,
        "terminal": status in TERMINAL_STATES,
    }


class JobManager:
    def __init__(self, settings: Settings, database: Database) -> None:
        self.settings = settings
        self.database = database
        self.executor = ThreadPoolExecutor(max_workers=1, thread_name_prefix="solver-job")

    def shutdown(self) -> None:
        self.executor.shutdown(wait=False, cancel_futures=False)

    def create(
        self,
        upload: dict[str, Any],
        require_optimal: bool,
        time_limit_seconds: float | None,
    ) -> dict[str, Any]:
        job_id = uuid.uuid4().hex
        now = utc_now()
        self.database.create_job(
            {
                "id": job_id,
                "upload_id": upload["id"],
                "status": "QUEUED",
                "message": "Execução adicionada à fila.",
                "module": upload["module"],
                "require_optimal": require_optimal,
                "time_limit_seconds": time_limit_seconds,
                "created_at": now,
                "updated_at": now,
            }
        )
        self.executor.submit(self._execute, job_id)
        return public_job(self.database.get_job(job_id) or {})

    def _rounds(self) -> set[Path]:
        return {item.resolve() for item in self.settings.result_dir.glob("rodada_*") if item.is_dir()}

    def _detect_round(self, previous: set[Path]) -> Path | None:
        created = sorted(self._rounds() - previous, key=lambda item: item.name)
        return created[0] if created else None

    def _execute(self, job_id: str) -> None:
        job = self.database.get_job(job_id)
        if not job:
            return
        job_log_dir = self.settings.job_dir / job_id
        job_log_dir.mkdir(parents=True, exist_ok=True)
        stdout_path = job_log_dir / "stdout.log"
        stderr_path = job_log_dir / "stderr.log"
        previous_rounds = self._rounds()
        self.database.update_job(
            job_id,
            status="RUNNING",
            message="Preparando a execução da vbeta.",
        )
        command = [
            sys.executable,
            str(self.settings.vbeta_dir / "executar_pipeline.py"),
            "--base",
            job["stored_path"],
            "--resultado",
            str(self.settings.result_dir),
            "--modulo",
            str(job["module"]),
        ]
        if job["time_limit_seconds"] is not None:
            command.extend(["--tempo-limite", str(job["time_limit_seconds"])])
        if not bool(job["require_optimal"]):
            command.append("--permitir-nao-otimo")

        creationflags = subprocess.CREATE_NO_WINDOW if os.name == "nt" else 0
        try:
            with stdout_path.open("w", encoding="utf-8") as stdout, stderr_path.open(
                "w", encoding="utf-8"
            ) as stderr:
                process = subprocess.Popen(
                    command,
                    cwd=self.settings.vbeta_dir,
                    stdout=stdout,
                    stderr=stderr,
                    text=True,
                    creationflags=creationflags,
                )
                self.database.update_job(job_id, process_id=process.pid)
                round_dir: Path | None = None
                while process.poll() is None:
                    if round_dir is None:
                        round_dir = self._detect_round(previous_rounds)
                        if round_dir:
                            self.database.update_job(
                                job_id,
                                round_name=round_dir.name,
                                round_dir=str(round_dir),
                                message="Rodada iniciada.",
                            )
                    time.sleep(0.4)
                exit_code = int(process.returncode or 0)
            round_dir = round_dir or self._detect_round(previous_rounds)
            state = "ERRO_INTERNO"
            message = "A execução terminou sem produzir status legível."
            if round_dir:
                status_path = round_dir / "status.json"
                if status_path.exists():
                    status = read_json(status_path)
                    state = status.get("state", state)
                    message = status.get("message", message)
            self.database.update_job(
                job_id,
                status=state,
                message=message,
                round_name=round_dir.name if round_dir else None,
                round_dir=str(round_dir) if round_dir else None,
                exit_code=exit_code,
            )
        except Exception as exc:
            self.database.update_job(
                job_id,
                status="ERRO_INTERNO",
                message=f"Falha ao iniciar ou acompanhar o processo: {exc}",
                exit_code=50,
            )


def load_summary(job: dict[str, Any]) -> dict[str, Any]:
    if not job.get("round_dir"):
        raise FileNotFoundError("A rodada ainda não foi criada.")
    path = Path(job["round_dir"]) / "alocacao" / "resumo_alocacao.json"
    if not path.exists():
        raise FileNotFoundError("O resumo da alocação ainda não está disponível.")
    return read_json(path)


def _round_workbook(job: dict[str, Any], manifest: dict[str, Any]) -> Path:
    round_dir = Path(job["round_dir"]).resolve()
    relative = manifest.get("artifacts", {}).get("allocation_workbook", {}).get("path")
    if not relative:
        raise FileNotFoundError("A planilha de resultado não consta no manifesto da rodada.")
    path = (round_dir / relative).resolve()
    if round_dir not in path.parents or not path.is_file():
        raise FileNotFoundError("A planilha de resultado da rodada não está disponível.")
    return path


def _sheet_records(path: Path, sheet_name: str) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        rows = workbook[sheet_name].iter_rows(values_only=True)
        headers = [str(value or "").strip() for value in next(rows)]
        return [dict(zip(headers, row, strict=False)) for row in rows]
    finally:
        workbook.close()


def _clean(value: Any) -> str:
    return str(value or "").strip()


def _number(value: Any) -> float:
    try:
        return float(value or 0)
    except (TypeError, ValueError):
        return 0.0


def _dashboard_teaching_capacity(value: Any, hours_per_transmission: int) -> float:
    """Return the CH_LETIVA usable by whole transmissions on the dashboard."""
    teaching_hours = max(0.0, _number(value))
    return (teaching_hours // hours_per_transmission) * hours_per_transmission


def _filter_values(value: Any) -> list[str]:
    """Normalize legacy scalar filters and HTTP multi-value filters."""
    raw_values = value if isinstance(value, (list, tuple, set)) else [value]
    values: list[str] = []
    seen: set[str] = set()
    for item in raw_values:
        cleaned = _clean(item)
        if cleaned and cleaned not in seen:
            seen.add(cleaned)
            values.append(cleaned)
    return values


def _filter_options(allocations: list[dict[str, Any]]) -> dict[str, list[Any]]:
    order_rank = {"1ª": 0, "2ª": 1, "ESTENDIDA": 2}
    day_rank = {
        "SEGUNDA": 0, "TERÇA": 1, "QUARTA": 2, "QUINTA": 3,
        "SEXTA": 4, "SÁBADO": 5, "DOMINGO": 6, "NSA": 7,
    }
    courses: dict[str, str] = {}
    for row in allocations:
        code = _clean(row.get("CURSO"))
        if code:
            courses[code] = _clean(row.get("NOME_CURSO")) or code
    return {
        "orders": sorted({_clean(row.get("ORDEM")) for row in allocations if _clean(row.get("ORDEM"))}, key=lambda value: order_rank.get(value, 99)),
        "courses": [{"value": code, "label": label} for code, label in sorted(courses.items(), key=lambda item: item[1])],
        "clusters": sorted({_clean(row.get("CLUSTER")) for row in allocations if _clean(row.get("CLUSTER"))}),
        "days": sorted({_clean(row.get("DIA_AULA")) for row in allocations if _clean(row.get("DIA_AULA"))}, key=lambda value: day_rank.get(value, 99)),
        "times": sorted({_clean(row.get("HORÁRIO")) for row in allocations if _clean(row.get("HORÁRIO"))}),
    }


def build_dashboard(
    job: dict[str, Any],
    summary: dict[str, Any],
    filters: dict[str, Any] | None = None,
) -> dict[str, Any]:
    round_dir = Path(job["round_dir"])
    manifest = read_json(round_dir / "manifesto.json")
    workbook_path = _round_workbook(job, manifest)
    all_allocations = _sheet_records(workbook_path, "ALOCACOES")
    teachers = _sheet_records(workbook_path, "DOCENTES")
    field_map = {
        "order": "ORDEM", "course": "CURSO", "cluster": "CLUSTER",
        "day": "DIA_AULA", "time": "HORÁRIO",
    }
    selected_filters = {
        key: values
        for key, value in (filters or {}).items()
        if key in field_map and (values := _filter_values(value))
    }
    allocations = [
        row for row in all_allocations
        if all(
            _clean(row.get(field_map[key])) in values
            for key, values in selected_filters.items()
        )
    ]
    allocated_rows = [row for row in allocations if _clean(row.get("STATUS")).upper() == "ALOCADA"]
    unassigned_rows = [row for row in allocations if _clean(row.get("STATUS")).upper() != "ALOCADA"]
    transmissions = len(allocations)
    allocated = len(allocated_rows)
    active_rows = [row for row in teachers if _clean(row.get("STATUS")).upper() == "ATIVO"]
    active_teachers = len(active_rows)
    used_badges = {_clean(row.get("CHAPA")) for row in allocated_rows if _clean(row.get("CHAPA"))}
    used_teachers = len(used_badges)
    hours_per_transmission = int(summary.get("hours_per_transmission", 2) or 2)

    discipline_by_order = Counter(_clean(row.get("ORDEM")) for row in allocations)
    first_demand = hours_per_transmission * (
        discipline_by_order.get("1ª", 0) + discipline_by_order.get("ESTENDIDA", 0)
    )
    second_demand = hours_per_transmission * (
        discipline_by_order.get("2ª", 0) + discipline_by_order.get("ESTENDIDA", 0)
    )
    active_capacity = sum(
        _dashboard_teaching_capacity(row.get("CH_LETIVA"), hours_per_transmission)
        for row in active_rows
    )

    day_data: dict[str, dict[str, Any]] = defaultdict(lambda: {"disciplines": 0, "teachers": set()})
    for row in allocations:
        day_data[_clean(row.get("DIA_AULA")) or "NÃO INFORMADO"]["disciplines"] += 1
    for row in allocated_rows:
        badge = _clean(row.get("CHAPA"))
        if badge:
            day_data[_clean(row.get("DIA_AULA")) or "NÃO INFORMADO"]["teachers"].add(badge)
    day_rank = {"SEGUNDA": 0, "TERÇA": 1, "QUARTA": 2, "QUINTA": 3, "SEXTA": 4, "SÁBADO": 5, "DOMINGO": 6, "NSA": 7}
    by_day = [
        {"day": day, "disciplines": values["disciplines"], "teachers": len(values["teachers"])}
        for day, values in sorted(day_data.items(), key=lambda item: day_rank.get(item[0], 99))
    ]

    cluster_hours = Counter(
        (_clean(row.get("CLUSTER")) or "NÃO INFORMADO") for row in allocations
    )
    cluster_items = [
        {"cluster": cluster, "hours": count * hours_per_transmission}
        for cluster, count in cluster_hours.most_common()
    ]

    stage_allocated_hours = {
        "first_stage": hours_per_transmission * sum(
            _clean(row.get("ORDEM")) in {"1ª", "ESTENDIDA"} for row in allocated_rows
        ),
        "second_stage": hours_per_transmission * sum(
            _clean(row.get("ORDEM")) in {"2ª", "ESTENDIDA"} for row in allocated_rows
        ),
    }
    validation = manifest.get("phases", {}).get("validation", {})
    audit = manifest.get("phases", {}).get("audit", {})
    return {
        "job": public_job(job),
        "kpis": {
            "coverage_pct": round(100 * allocated / transmissions, 2) if transmissions else 0,
            "allocated": allocated,
            "transmissions": transmissions,
            "unassigned": int(summary.get("unassigned", 0)),
            "teacher_use_pct": (
                round(100 * used_teachers / active_teachers, 2) if active_teachers else 0
            ),
            "active_teachers": active_teachers,
            "used_teachers": used_teachers,
            "zero_active_teachers": max(0, active_teachers - used_teachers),
            "disciplines_by_order": {
                "first": discipline_by_order.get("1ª", 0),
                "second": discipline_by_order.get("2ª", 0),
                "extended": discipline_by_order.get("ESTENDIDA", 0),
            },
            "first_stage_demand_hours": first_demand,
            "second_stage_demand_hours": second_demand,
            "first_stage_capacity_delta_hours": round(first_demand - active_capacity, 2),
            "second_stage_capacity_delta_hours": round(second_demand - active_capacity, 2),
            "active_teaching_capacity_hours": round(active_capacity, 2),
        },
        "unassigned_reasons": [
            {"reason": reason, "count": count}
            for reason, count in sorted(
                Counter(_clean(row.get("MOTIVO")) or "NÃO INFORMADO" for row in unassigned_rows).items(),
                key=lambda item: item[1],
                reverse=True,
            )
        ],
        "stage_hours": stage_allocated_hours,
        "charts": {"by_day": by_day, "demand_hours_by_cluster": cluster_items},
        "filters": {"selected": selected_filters, "options": _filter_options(all_allocations)},
        "guardrails": {
            "validation": validation.get("status"),
            "solver": summary.get("solver_status"),
            "audit": audit.get("status"),
        },
        "wall_time_seconds": summary.get("wall_time_seconds"),
        "metric_notes": {
            "demand": f"Cada oferta representa {hours_per_transmission}h semanais. ESTENDIDA participa das duas etapas.",
            "capacity_delta": "Demanda filtrada menos a CH letiva utilizável dos docentes ativos, arredondada para transmissões completas; valor positivo indica déficit. A compatibilidade de perfil não reduz esta capacidade bruta.",
        },
        "source_note": "Indicadores recalculados a partir da planilha de resultado e do manifesto auditado da rodada.",
    }


def artifact_path(job: dict[str, Any], artifact_key: str) -> Path:
    if not job.get("round_dir"):
        raise FileNotFoundError("A rodada ainda não possui artefatos.")
    round_dir = Path(job["round_dir"]).resolve()
    manifest_path = round_dir / "manifesto.json"
    if not manifest_path.exists():
        raise FileNotFoundError("Manifesto ainda não disponível.")
    if artifact_key == "manifest":
        path = manifest_path.resolve()
    elif artifact_key == "status":
        path = (round_dir / "status.json").resolve()
    else:
        manifest = read_json(manifest_path)
        artifact = manifest.get("artifacts", {}).get(artifact_key)
        if not artifact:
            raise FileNotFoundError("Artefato não encontrado.")
        path = (round_dir / artifact["path"]).resolve()
    if round_dir not in path.parents or not path.is_file():
        raise FileNotFoundError("Caminho de artefato inválido.")
    return path
